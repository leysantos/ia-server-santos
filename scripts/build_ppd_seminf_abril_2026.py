#!/usr/bin/env python3
"""Gera planilhas-exemplos/ppd_seminf_abril_2026.xlsm a partir do template v8.1 SEMINF.

Adiciona abas ORC_SINTETICO, ORC_ANALITICO, CRONOGRAMA e ESP_TECNICA (marcadores 2026),
preservando VBA, Base de preços e fórmulas MCQ do modelo oficial.
"""
from __future__ import annotations

import argparse
import re
import shutil
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
PLANILHAS = ROOT / "planilhas-exemplos"
SOURCE_V81 = PLANILHAS / "00_MOD_MC_OR_R00-Nivel-1-2-Abril2026-10-06-2026v8.1.xlsm"
OUTPUT = PLANILHAS / "ppd_seminf_abril_2026.xlsm"

# Abas auxiliares do v8.1 — ocultas no pacote 2026 (não apagar: macros podem referenciar)
HIDE_SHEETS = frozenset(
    {
        "CURVA_ABC",
        "FUNDACAO RASA",
        "FUNDACAO PROFUNDA",
        "SUPRAESTRUTURA",
        "PAREDES E REVEST",
        "PISOS",
    }
)

PLANILHA_REF_RE = re.compile(r"PLANILHA!", re.IGNORECASE)

# Abas onde substituir fórmulas (evita varrer CURVA_ABC com 100k+ linhas)
REF_SHEETS = frozenset({"MCQ", "PLANILHA", "ORC_SINTETICO", "ORC_ANALITICO", "ETAPAS"})


def _replace_planilha_refs(wb) -> int:
    """Substitui PLANILHA! → ORC_SINTETICO! em fórmulas e named ranges."""
    changed = 0
    for ws in wb.worksheets:
        if ws.title not in REF_SHEETS and not ws.title.startswith("ORC_"):
            continue
        max_row = min(ws.max_row or 1, 2500)
        max_col = min(ws.max_column or 1, 40)
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.startswith("=") and PLANILHA_REF_RE.search(val):
                    cell.value = PLANILHA_REF_RE.sub("ORC_SINTETICO!", val)
                    changed += 1

    for name in list(wb.defined_names):
        dn = wb.defined_names[name]
        attr = getattr(dn, "attr_text", None) or ""
        if attr and PLANILHA_REF_RE.search(attr):
            dn.attr_text = PLANILHA_REF_RE.sub("ORC_SINTETICO!", attr)
            changed += 1
    return changed


def _copy_sheet(wb, source: str, target: str) -> None:
    if target in wb.sheetnames:
        del wb[target]
    ws = wb.copy_worksheet(wb[source])
    ws.title = target


def _seed_cronograma(ws) -> None:
    """Layout mínimo compatível com sync_cronograma_worksheet."""
    ws.cell(1, 13, "OBJETO:")
    ws.cell(2, 13, "LOCAL:")
    ws.cell(3, 13, "ORÇAMENTO:")
    ws.cell(4, 12, "PRAZO DE EXECUÇAO:")
    ws.cell(8, 1, "ETAPA")
    ws.cell(8, 2, "DESCRIÇÃO")
    ws.cell(8, 3, "VALOR (R$)")
    ws.page_setup.orientation = "landscape"


def _seed_esp_tecnica(ws) -> None:
    ws.cell(1, 1, "PROJETO")
    ws.cell(2, 1, "ESPECIFICAÇÃO TÉCNICA")
    ws.cell(
        4,
        1,
        "(Conteúdo não gerado — use a aba Especificação no orçamento para gerar o texto)",
    )


def build_ppd_seminf_2026(
    source: Path = SOURCE_V81,
    output: Path = OUTPUT,
    *,
    force: bool = False,
) -> Path:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl necessário — rode make setup-backend") from exc

    if not source.exists():
        raise FileNotFoundError(f"Template v8.1 ausente: {source}")
    if output.exists() and not force:
        raise FileExistsError(f"{output} já existe — use --force para sobrescrever")

    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        output.unlink()
    shutil.copy(source, output)

    wb = openpyxl.load_workbook(output, keep_vba=True)

    if "PLANILHA" not in wb.sheetnames:
        raise ValueError("Template v8.1 sem aba PLANILHA")

    _copy_sheet(wb, "PLANILHA", "ORC_SINTETICO")
    _copy_sheet(wb, "PLANILHA", "ORC_ANALITICO")

    refs = _replace_planilha_refs(wb)

    if "CRONOGRAMA" not in wb.sheetnames:
        ws_cron = wb.create_sheet("CRONOGRAMA")
        _seed_cronograma(ws_cron)
    if "ESP_TECNICA" not in wb.sheetnames:
        ws_esp = wb.create_sheet("ESP_TECNICA")
        _seed_esp_tecnica(ws_esp)

    del wb["PLANILHA"]

    for name in HIDE_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"

    wb.save(output)
    wb.close()

    # Validação
    sys.path.insert(0, str(ROOT / "backend"))
    from pricing.budget.ppd_workbook_init import is_seminf_2026_workbook, workbook_sheetnames

    names = workbook_sheetnames(output)
    if not is_seminf_2026_workbook(names):
        raise RuntimeError(f"Workbook gerado inválido — abas: {names}")

    print(f"✓ Gerado: {output}")
    print(f"  Abas: {names}")
    print(f"  Fórmulas PLANILHA! → ORC_SINTETICO!: {refs}")
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=SOURCE_V81)
    parser.add_argument("--output", type=Path, default=OUTPUT)
    parser.add_argument("--force", action="store_true", help="Sobrescrever arquivo existente")
    args = parser.parse_args()
    build_ppd_seminf_2026(args.source, args.output, force=args.force)


if __name__ == "__main__":
    main()
