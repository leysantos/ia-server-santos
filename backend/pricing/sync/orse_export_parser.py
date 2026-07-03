"""Parser de exports Excel/CSV do ORSE 2 → price_bank."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from pricing.budget.orse_bundle_detect import is_foreign_price_base_file
from pricing.budget.price_bank_store import (
    CompositionClosed,
    CompositionItem,
    CompositionOpen,
    InsumoRecord,
)
from pricing.providers._tabular import _norm_header, _pick_column, _rows_from_matrix

_CODE_KEYS = ("codigo", "code", "código", "cod", "item", "codinsumo", "codservico")
_DESC_KEYS = ("descricao", "descrição", "description", "desc", "servico", "serviço")
_UNIT_KEYS = ("unidade", "unit", "und", "un")
_PRICE_KEYS = (
    "preco",
    "preço",
    "price",
    "valor",
    "custo",
    "total",
    "precounitario",
    "preco unitario",
    "preco unitário",
    "precounit",
)
_TYPE_KEYS = ("tipo", "classificacao", "classificação", "categoria")
_COEF_KEYS = ("coeficiente", "coef", "qtde", "quantidade", "qtd")


@dataclass
class OrseExportBundle:
    closed: list[CompositionClosed]
    open_map: dict[str, CompositionOpen]
    insumos: list[InsumoRecord]
    metadata: dict[str, Any]


def _float_cell(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip()
    if not text:
        return 0.0
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _load_workbook_matrix(path: Path) -> list[tuple[str, list[list]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        import csv

        with open(path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
        if not rows:
            return []
        return [("Sheet1", rows)]
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl necessário: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, list[list]]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        matrix = [list(row) for row in ws.iter_rows(values_only=True)]
        if matrix:
            out.append((name, matrix))
    wb.close()
    return out


def _find_header_row(matrix: list[list], *, max_scan: int = 40) -> tuple[int, list[str]] | None:
    for idx, row in enumerate(matrix[:max_scan]):
        headers = [_norm_header(c) for c in row]
        if _pick_column(headers, _CODE_KEYS) is not None and (
            _pick_column(headers, _DESC_KEYS) is not None or _pick_column(headers, _PRICE_KEYS) is not None
        ):
            return idx, [str(c or "") for c in row]
    return None


def _parse_tabular_sheet(matrix: list[list]) -> list[dict[str, Any]]:
    found = _find_header_row(matrix)
    if not found:
        return []
    header_idx, headers = found
    rows = _rows_from_matrix(headers, matrix[header_idx + 1 :])
    return rows


def _rows_to_closed(rows: list[dict[str, Any]]) -> list[CompositionClosed]:
    closed: list[CompositionClosed] = []
    seen: set[str] = set()
    for row in rows:
        code = str(row.get("code") or "").strip()
        if not code or code in seen:
            continue
        desc = str(row.get("description") or "").strip()
        if not desc:
            continue
        price = float(row.get("price") or 0)
        if price <= 0:
            continue
        seen.add(code)
        closed.append(
            CompositionClosed(
                code=code,
                description=desc,
                unit=str(row.get("unit") or "un").strip() or "un",
                price=price,
                price_sem_desoneracao=price,
            )
        )
    return closed


def _rows_to_insumos(rows: list[dict[str, Any]]) -> list[InsumoRecord]:
    insumos: list[InsumoRecord] = []
    seen: set[str] = set()
    for row in rows:
        code = str(row.get("code") or "").strip()
        if not code or code in seen:
            continue
        desc = str(row.get("description") or "").strip()
        if not desc:
            continue
        price = float(row.get("price") or 0)
        seen.add(code)
        insumos.append(
            InsumoRecord(
                code=code,
                description=desc,
                unit=str(row.get("unit") or "un").strip() or "un",
                price=price,
                price_sem_desoneracao=price,
                origin="ORSE",
            )
        )
    return insumos


def _parse_analytical_workbook(path: Path) -> dict[str, CompositionOpen]:
    """Tenta montar CPUs a partir de export analítico ORSE (composição + itens)."""
    open_map: dict[str, CompositionOpen] = {}
    for sheet_name, matrix in _load_workbook_matrix(path):
        found = _find_header_row(matrix)
        if not found:
            continue
        header_idx, headers = found
        headers_norm = [_norm_header(h) for h in headers]
        code_i = _pick_column(headers_norm, _CODE_KEYS)
        desc_i = _pick_column(headers_norm, _DESC_KEYS)
        unit_i = _pick_column(headers_norm, _UNIT_KEYS)
        price_i = _pick_column(headers_norm, _PRICE_KEYS)
        coef_i = _pick_column(headers_norm, _COEF_KEYS)
        type_i = _pick_column(headers_norm, _TYPE_KEYS)
        if code_i is None:
            continue

        current: CompositionOpen | None = None
        for row in matrix[header_idx + 1 :]:
            if not row:
                continue
            cells = list(row)
            code = str(cells[code_i] if code_i < len(cells) else "").strip()
            desc = str(cells[desc_i] if desc_i is not None and desc_i < len(cells) else "").strip()
            if not code and not desc:
                continue

            coef = _float_cell(cells[coef_i]) if coef_i is not None and coef_i < len(cells) else 0.0
            unit_price = _float_cell(cells[price_i]) if price_i is not None and price_i < len(cells) else 0.0
            unit = str(cells[unit_i] if unit_i is not None and unit_i < len(cells) else "un").strip() or "un"
            item_type_raw = str(cells[type_i] if type_i is not None and type_i < len(cells) else "").lower()

            is_header = coef <= 0 and unit_price > 0 and re.match(r"^\d", code)
            is_item = coef > 0 or (unit_price > 0 and not is_header)

            if is_header or (code and desc and not is_item and unit_price > 0):
                current = CompositionOpen(
                    code=code,
                    description=desc,
                    unit=unit,
                    total_price=unit_price,
                    total_price_sem=unit_price,
                    items=[],
                )
                open_map[code] = current
                continue

            if current and is_item and code:
                item_type = "insumo"
                if "mao" in item_type_raw or "m.o" in item_type_raw:
                    item_type = "mao_obra"
                elif "equip" in item_type_raw:
                    item_type = "equipamento"
                partial = unit_price * coef if coef else unit_price
                current.items.append(
                    CompositionItem(
                        item_type=item_type,
                        code=code,
                        description=desc,
                        unit=unit,
                        coefficient=coef or 1.0,
                        unit_price=unit_price,
                        partial_cost=partial,
                        unit_price_sem=unit_price,
                        partial_cost_sem=partial,
                    )
                )
                current.total_price = sum(i.partial_cost for i in current.items)
                current.total_price_sem = current.total_price
        if open_map:
            break
    return open_map


def parse_orse_composicoes_file(path: Path) -> list[CompositionClosed]:
    closed: list[CompositionClosed] = []
    for _sheet, matrix in _load_workbook_matrix(path):
        closed.extend(_rows_to_closed(_parse_tabular_sheet(matrix)))
    return closed


def parse_orse_insumos_file(path: Path) -> list[InsumoRecord]:
    insumos: list[InsumoRecord] = []
    for _sheet, matrix in _load_workbook_matrix(path):
        insumos.extend(_rows_to_insumos(_parse_tabular_sheet(matrix)))
    return insumos


def _validate_orse_bundle(
    bundle: OrseExportBundle,
    *,
    composicoes_path: Path,
    insumos_path: Path | None,
    analitico_path: Path | None,
) -> None:
    for path, role in (
        (composicoes_path, "composições"),
        (insumos_path, "insumos"),
        (analitico_path, "analítico"),
    ):
        if path and path.is_file() and is_foreign_price_base_file(path):
            raise ValueError(
                f"Arquivo '{path.name}' não é export ORSE ({role}) — parece ser SEMINF/PPD/SINAPI. "
                "Exporte planilhas do ORSE 2 (Relatórios → Cadastrais) e use 'Importar pasta export ORSE'."
            )

    seminf_hits = sum(
        1 for c in bundle.closed if ".SEMINF" in c.code.upper() or c.code.upper().endswith("SEMINF")
    )
    if seminf_hits > 0:
        raise ValueError(
            f"As composições parecem ser da base SEMINF ({seminf_hits} códigos .SEMINF), não ORSE. "
            "Exclua a referência incorreta e reimporte com os Excel exportados do ORSE 2."
        )

    if not bundle.insumos:
        raise ValueError(
            "Importação ORSE incompleta: planilha de Insumos ausente ou vazia. "
            "No ORSE 2 exporte Relatórios → Cadastrais → Insumos (Excel)."
        )

    if not bundle.open_map:
        raise ValueError(
            "Importação ORSE incompleta: nenhuma composição aberta (CPU) encontrada. "
            "Exporte Relatórios → Cadastrais → Analítico de Composições (Excel) "
            "e inclua na pasta de importação."
        )


def parse_orse_export_bundle(
    *,
    composicoes_path: Path,
    insumos_path: Path | None = None,
    analitico_path: Path | None = None,
) -> OrseExportBundle:
    closed = parse_orse_composicoes_file(composicoes_path)
    if not closed:
        raise ValueError(
            f"Nenhuma composição parseada em {composicoes_path.name}. "
            "Verifique se o Excel contém colunas Código, Descrição, Unidade e Preço."
        )

    insumos: list[InsumoRecord] = []
    if insumos_path and insumos_path.is_file():
        insumos = parse_orse_insumos_file(insumos_path)

    open_map: dict[str, CompositionOpen] = {}
    if analitico_path and analitico_path.is_file():
        open_map = _parse_analytical_workbook(analitico_path)

    bundle = OrseExportBundle(
        closed=closed,
        open_map=open_map,
        insumos=insumos,
        metadata={
            "composicoes_file": str(composicoes_path.resolve()),
            "insumos_file": str(insumos_path.resolve()) if insumos_path else "",
            "analitico_file": str(analitico_path.resolve()) if analitico_path else "",
        },
    )
    _validate_orse_bundle(
        bundle,
        composicoes_path=composicoes_path,
        insumos_path=insumos_path,
        analitico_path=analitico_path,
    )
    return bundle
