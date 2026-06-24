"""Parser SINAPI Referência — insumos, composições fechadas (sintéticas) e abertas (analíticas/CPU)."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from pricing.budget.price_bank_store import (
    CompositionClosed,
    CompositionItem,
    CompositionOpen,
    InsumoRecord,
)

BRAZIL_UFS = frozenset(
    {
        "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG",
        "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO",
    }
)

_NATIONAL_DATA_START = 10
_NATIONAL_UF_ROW = 8
_NATIONAL_HEADER_ROW = 9

_COMPOSITION_CLOSED_SHEETS = frozenset({"csd", "ccd", "cse", "cce"})
_INSUMO_SHEETS = frozenset({"isd", "icd", "ise", "ice"})
_ANALYTICAL_SHEETS = frozenset(
    {
        "analítico",
        "analitico",
        "analítico com custo",
        "analitico com custo",
        "analitico c custo",
    }
)

_CODE_KEYS = ("codigo", "código", "cod", "code")
_COMP_CODE_KEYS = ("codigo da composição", "código da composição", "codigo composicao", "cod composicao")
_DESC_KEYS = ("descricao", "descrição", "description", "desc")
_COMP_DESC_KEYS = (
    "descricao da composição",
    "descrição da composição",
    "descricao composicao",
)
_ITEM_DESC_KEYS = ("descricao do item", "descrição do item", "descricao item")
_UNIT_KEYS = ("unidade", "unit", "und")
_ITEM_UNIT_KEYS = ("unidade do item", "unidade item")
_PRICE_KEYS = ("custo total", "custo", "preco", "preço", "price", "valor", "total")
_UNIT_PRICE_KEYS = ("preço unitário", "preco unitario", "preço unitario", "preco unitário")
_COEF_KEYS = ("coeficiente", "coef", "quantidade")
_ITEM_CODE_KEYS = ("codigo do item", "código do item", "codigo item", "cod item")
_TYPE_KEYS = ("tipo item", "tipo", "tipo do item")
_PARTIAL_KEYS = ("custo do item", "custo parcial", "custo item")


def _norm(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _pick_col(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    for idx, header in enumerate(headers):
        h = _norm(header)
        if h in candidates or any(c in h for c in candidates):
            return idx
    return None


def _float_cell(value: object, default: float = 0.0) -> float:
    if value is None:
        return default
    text = str(value).strip()
    if not text or text in {"-", "—", "–"}:
        return default
    try:
        return float(text.replace(",", "."))
    except (TypeError, ValueError):
        return default


def _pct_cell(value: object, default: float = 0.0) -> float:
    """%AS na planilha SINAPI: fração decimal (0.0204 = 2,04%) ou inteiro percentual."""
    if value is None:
        return default
    text = str(value).strip()
    if not text or text in {"-", "—", "–"}:
        return default
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        num = float(text.replace(",", "."))
    except (TypeError, ValueError):
        return default
    if num > 1:
        return num / 100.0
    return num


def _str_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_national_matrix(matrix: list[tuple]) -> bool:
    if len(matrix) <= _NATIONAL_UF_ROW:
        return False
    row = matrix[_NATIONAL_UF_ROW]
    ufs = {str(c).strip().upper() for c in row if c and str(c).strip().upper() in BRAZIL_UFS}
    return len(ufs) >= 10


def _uf_columns_from_row(row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        uf = str(cell).strip().upper()
        if uf in BRAZIL_UFS:
            mapping[uf] = idx
    return mapping


def _insumo_uf_columns(matrix: list[tuple]) -> dict[str, int]:
    """ISD/ICD: UFs na linha de cabeçalho (geralmente índice 3 ou 9)."""
    for row_idx in (3, _NATIONAL_HEADER_ROW, _NATIONAL_UF_ROW):
        if row_idx < len(matrix):
            cols = _uf_columns_from_row(matrix[row_idx])
            if cols:
                return cols
    return {}


def _closed_uf_cost_pct_columns(matrix: list[tuple]) -> dict[str, dict[str, int]]:
    """CSD/CCD: cada UF tem coluna Custo (R$) e %AS adjacente."""
    uf_row = matrix[_NATIONAL_UF_ROW] if len(matrix) > _NATIONAL_UF_ROW else ()
    header_row = matrix[_NATIONAL_HEADER_ROW] if len(matrix) > _NATIONAL_HEADER_ROW else ()
    mapping: dict[str, dict[str, int]] = {}
    for idx, cell in enumerate(uf_row):
        if cell is None:
            continue
        uf = str(cell).strip().upper()
        if uf not in BRAZIL_UFS:
            continue
        cost_col = idx
        pct_col = idx + 1
        if pct_col < len(header_row):
            h_next = _norm(header_row[pct_col])
            if "%as" not in h_next and h_next not in {"as", "% as"}:
                pct_col = idx + 1
        mapping[uf] = {"cost_col": cost_col, "pct_as_col": pct_col}
    return mapping


def _composition_codes_from_analitico(matrix: list[tuple]) -> list[tuple[str, str, str, str]]:
    """Linhas-pai do Analítico: (código, descrição, unidade, grupo) na ordem do CSD."""
    out: list[tuple[str, str, str, str]] = []
    for row in matrix[_NATIONAL_DATA_START:]:
        if not row or len(row) < 6:
            continue
        code = row[1]
        tipo = row[2]
        if code and not tipo:
            out.append(
                (
                    str(int(code)) if isinstance(code, (int, float)) else str(code).strip(),
                    str(row[4] or "").strip(),
                    str(row[5] or "un").strip(),
                    _str_cell(row[0]),
                )
            )
    return out


def _parse_closed_national_all_ufs(
    matrix: list[tuple],
    *,
    comp_codes: list[tuple[str, str, str, str]],
) -> list[tuple[str, str, str, str, dict[str, dict[str, float]]]]:
    """Retorna (code, desc, unit, grupo, {uf: {price, pct_as}}) por linha — aba CSD ou CCD."""
    uf_cols = _closed_uf_cost_pct_columns(matrix)
    rows: list[tuple[str, str, str, str, dict[str, dict[str, float]]]] = []
    for idx, row in enumerate(matrix[_NATIONAL_DATA_START:]):
        if not row:
            continue
        meta = comp_codes[idx] if idx < len(comp_codes) else None
        grupo = meta[3] if meta else _str_cell(row[0])
        code = meta[0] if meta else str(row[1] or "").strip()
        desc = meta[1] if meta else str(row[2] or "").strip()
        unit = meta[2] if meta else str(row[3] or "un").strip()
        if not desc:
            continue
        regional: dict[str, dict[str, float]] = {}
        for uf, cols in uf_cols.items():
            cost_col = cols["cost_col"]
            pct_col = cols["pct_as_col"]
            price = _float_cell(row[cost_col]) if cost_col < len(row) else 0.0
            pct_as = _pct_cell(row[pct_col]) if pct_col < len(row) else 0.0
            regional[uf] = {"price": price, "pct_as": pct_as}
        rows.append((code, desc, unit, grupo, regional))
    return rows


def _parse_insumos_national_all_ufs(
    matrix: list[tuple],
    *,
    origin: str,
) -> list[tuple[str, str, str, str, str, dict[str, float]]]:
    """Retorna (code, desc, unit, classificacao, origem_preco, {uf: price}) por insumo."""
    uf_cols = _insumo_uf_columns(matrix)
    rows: list[tuple[str, str, str, str, str, dict[str, float]]] = []
    for row in matrix[_NATIONAL_DATA_START:]:
        if not row:
            continue
        code_raw = row[1]
        desc = str(row[2] or "").strip()
        if not code_raw or not desc:
            continue
        code = str(int(code_raw)) if isinstance(code_raw, (int, float)) else str(code_raw).strip()
        unit = str(row[3] or "un").strip()
        classificacao = _str_cell(row[0])
        origem_preco = _str_cell(row[4]) if len(row) > 4 else ""
        regional: dict[str, float] = {}
        for uf, col in uf_cols.items():
            if col < len(row):
                regional[uf] = _float_cell(row[col])
        rows.append((code, desc, unit, classificacao, origem_preco, regional))
    return rows


def _parse_closed_national(
    matrix: list[tuple],
    *,
    uf: str,
    comp_codes: list[tuple[str, str, str, str]],
) -> list[tuple[str, str, str, str, float]]:
    """Retorna (code, desc, unit, grupo, price) por linha."""
    uf = uf.upper()
    uf_cols = _closed_uf_cost_pct_columns(matrix)
    cols = uf_cols.get(uf)
    if cols is None:
        raise ValueError(f"UF {uf} não encontrada na planilha SINAPI nacional")
    col = cols["cost_col"]

    rows: list[tuple[str, str, str, str, float]] = []
    for idx, row in enumerate(matrix[_NATIONAL_DATA_START:]):
        if not row or len(row) <= col:
            continue
        meta = comp_codes[idx] if idx < len(comp_codes) else None
        grupo = meta[3] if meta else _str_cell(row[0])
        code = meta[0] if meta else str(row[1] or "").strip()
        desc = meta[1] if meta else str(row[2] or "").strip()
        unit = meta[2] if meta else str(row[3] or "un").strip()
        if not desc:
            continue
        rows.append((code, desc, unit, grupo, _float_cell(row[col])))
    return rows


def _parse_insumos_national(
    matrix: list[tuple],
    *,
    uf: str,
    origin: str,
) -> list[tuple[str, str, str, str, str, float]]:
    """Retorna (code, desc, unit, classificacao, origem_preco, price) por insumo."""
    uf = uf.upper()
    uf_cols = _insumo_uf_columns(matrix)
    col = uf_cols.get(uf)
    if col is None:
        return []

    rows: list[tuple[str, str, str, str, str, float]] = []
    for row in matrix[_NATIONAL_DATA_START:]:
        if not row or len(row) <= col:
            continue
        code_raw = row[1]
        desc = str(row[2] or "").strip()
        if not code_raw or not desc:
            continue
        code = str(int(code_raw)) if isinstance(code_raw, (int, float)) else str(code_raw).strip()
        unit = str(row[3] or "un").strip()
        classificacao = _str_cell(row[0])
        origem_preco = _str_cell(row[4]) if len(row) > 4 else ""
        rows.append((code, desc, unit, classificacao, origem_preco, _float_cell(row[col])))
    return rows


def _price_maps(
    closed: list[CompositionClosed],
    insumos: list[InsumoRecord],
) -> tuple[dict[str, float], dict[str, float], dict[str, float], dict[str, float]]:
    comp_com = {c.code: c.price for c in closed}
    comp_sem = {c.code: c.price_sem_desoneracao or c.price for c in closed}
    ins_com = {i.code: i.price for i in insumos}
    ins_sem = {i.code: i.price_sem_desoneracao or i.price for i in insumos}
    return comp_com, comp_sem, ins_com, ins_sem


def _lookup_item_prices(
    item_type: str,
    code: str,
    comp_com: dict[str, float],
    comp_sem: dict[str, float],
    ins_com: dict[str, float],
    ins_sem: dict[str, float],
) -> tuple[float, float]:
    t = _norm(item_type)
    if "compos" in t or t == "composicao":
        return comp_com.get(code, 0.0), comp_sem.get(code, 0.0)
    return ins_com.get(code, 0.0), ins_sem.get(code, 0.0)


def _parse_analytical_national(
    matrix: list[tuple],
    *,
    comp_com: dict[str, float],
    comp_sem: dict[str, float],
    ins_com: dict[str, float],
    ins_sem: dict[str, float],
    insumo_meta: dict[str, dict[str, str]] | None = None,
    closed_grupos: dict[str, str] | None = None,
) -> dict[str, CompositionOpen]:
    compositions: dict[str, CompositionOpen] = {}
    current_code = ""
    current_desc = ""
    current_unit = "un"
    current_grupo = ""
    meta_by_code = insumo_meta or {}
    grupos_by_code = closed_grupos or {}

    for row in matrix[_NATIONAL_DATA_START:]:
        if not row or len(row) < 6:
            continue
        comp_code_raw = row[1]
        tipo = row[2]
        item_code_raw = row[3]
        desc = str(row[4] or "").strip()
        unit = str(row[5] or "un").strip()
        coef = _float_cell(row[6] if len(row) > 6 else 0)
        situacao = _str_cell(row[7]) if len(row) > 7 else ""

        if comp_code_raw and not tipo:
            current_code = (
                str(int(comp_code_raw))
                if isinstance(comp_code_raw, (int, float))
                else str(comp_code_raw).strip()
            )
            current_desc = desc
            current_unit = unit
            current_grupo = _str_cell(row[0]) or grupos_by_code.get(current_code, "")
            if current_code not in compositions:
                price_com = comp_com.get(current_code, 0.0)
                price_sem = comp_sem.get(current_code, 0.0)
                compositions[current_code] = CompositionOpen(
                    code=current_code,
                    description=current_desc,
                    unit=current_unit,
                    total_price=price_com,
                    total_price_sem=price_sem,
                    grupo=current_grupo,
                    items=[],
                )
            continue

        if not current_code or not tipo:
            continue

        item_code = (
            str(int(item_code_raw))
            if isinstance(item_code_raw, (int, float))
            else str(item_code_raw or "").strip()
        )
        if not item_code and not desc:
            continue

        type_raw = str(tipo).strip()
        item_type = _classify_item_type(type_raw, item_code, desc)
        unit_price, unit_price_sem = _lookup_item_prices(
            item_type, item_code, comp_com, comp_sem, ins_com, ins_sem
        )
        partial = round(coef * unit_price, 6) if coef and unit_price else 0.0
        partial_sem = round(coef * unit_price_sem, 6) if coef and unit_price_sem else 0.0

        ins_meta = meta_by_code.get(item_code, {})
        classificacao = ins_meta.get("classificacao", "")
        origem_preco = ins_meta.get("origem_preco", "")

        comp = compositions.setdefault(
            current_code,
            CompositionOpen(
                code=current_code,
                description=current_desc or current_code,
                unit=current_unit,
                total_price=comp_com.get(current_code, 0.0),
                total_price_sem=comp_sem.get(current_code, 0.0),
                grupo=current_grupo or grupos_by_code.get(current_code, ""),
                items=[],
            ),
        )
        if current_grupo and not comp.grupo:
            comp.grupo = current_grupo
        comp.items.append(
            CompositionItem(
                item_type=item_type,
                code=item_code or desc[:20],
                description=desc or item_code,
                unit=unit,
                coefficient=coef,
                unit_price=unit_price,
                partial_cost=partial,
                unit_price_sem=unit_price_sem,
                partial_cost_sem=partial_sem,
                classificacao=classificacao,
                origem_preco=origem_preco,
                situacao=situacao,
            )
        )

    return compositions


def _parse_labor_charges_sheet(
    matrix: list[tuple],
    *,
    paired_uf_cols: bool,
) -> dict[str, dict[str, float | str]]:
    """Extrai Horista/Mensalista por UF das linhas 4–7 da planilha nacional."""
    if len(matrix) < 7:
        return {}
    uf_row = matrix[3]
    loc_row = matrix[4]
    horista_row = matrix[5]
    mens_row = matrix[6]
    out: dict[str, dict[str, float | str]] = {}

    def _scan_ufs(row: tuple) -> list[tuple[str, int]]:
        found: list[tuple[str, int]] = []
        for idx, cell in enumerate(row):
            if cell is None:
                continue
            uf = str(cell).strip().upper()
            if uf in BRAZIL_UFS:
                found.append((uf, idx))
        return found

    for uf, idx in _scan_ufs(uf_row):
        out[uf] = {
            "localidade": _str_cell(loc_row[idx] if idx < len(loc_row) else ""),
            "horista": _float_cell(horista_row[idx] if idx < len(horista_row) else 0),
            "mensalista": _float_cell(mens_row[idx] if idx < len(mens_row) else 0),
        }
    return out


def _merge_labor_charges(
    semd: dict[str, dict[str, float | str]],
    comd: dict[str, dict[str, float | str]],
) -> dict[str, dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for uf in set(semd) | set(comd):
        s = semd.get(uf, {})
        c = comd.get(uf, {})
        merged[uf] = {
            "localidade": str(s.get("localidade") or c.get("localidade") or ""),
            "horista_semd": float(s.get("horista") or 0),
            "mensalista_semd": float(s.get("mensalista") or 0),
            "horista_comd": float(c.get("horista") or 0),
            "mensalista_comd": float(c.get("mensalista") or 0),
        }
    return merged


def _apply_sinapi_tp2(
    closed: list[CompositionClosed],
    open_map: dict[str, CompositionOpen],
) -> None:
    from pricing.budget.tp2_as import apply_tp2_to_items, composition_tp2_from_regional, merge_tp2

    closed_by_code = {c.code: c for c in closed}
    for comp in closed:
        comp.tp2 = composition_tp2_from_regional(comp.regional)

    for code, comp in open_map.items():
        closed_row = closed_by_code.get(code)
        comp_tp2 = ""
        pct_as = 0.0
        if closed_row:
            comp_tp2 = closed_row.tp2 or composition_tp2_from_regional(closed_row.regional)
            for entry in (closed_row.regional or {}).values():
                if isinstance(entry, dict):
                    pct_as = max(
                        pct_as,
                        float(entry.get("pct_as_comd") or 0),
                        float(entry.get("pct_as_semd") or 0),
                    )
        comp.tp2 = merge_tp2(comp_tp2, pct_as)
        items_dicts = apply_tp2_to_items(
            [i.to_dict() for i in comp.items],
            composition_tp2=comp.tp2,
            pct_as=pct_as,
        )
        comp.items = [CompositionItem(**item) for item in items_dicts]


def _parse_national_workbook(wb: Any, *, uf: str) -> dict[str, Any]:
    sheets = {_norm(name): name for name in wb.sheetnames}

    def _matrix(sheet_key: str) -> list[tuple]:
        name = sheets.get(sheet_key)
        if not name:
            return []
        ws = wb[name]
        return list(ws.iter_rows(values_only=True))

    csd = _matrix("csd")
    ccd = _matrix("ccd")
    isd = _matrix("isd")
    icd = _matrix("icd")
    analitico = _matrix("analítico") or _matrix("analitico")

    if not _is_national_matrix(csd):
        raise ValueError("Planilha não está no formato nacional SINAPI 2025+")

    comp_codes = _composition_codes_from_analitico(analitico) if analitico else []
    csd_all = _parse_closed_national_all_ufs(csd, comp_codes=comp_codes)
    ccd_all = _parse_closed_national_all_ufs(ccd, comp_codes=comp_codes) if ccd else []
    ccd_by_idx = {i: reg for i, (_, _, _, _, reg) in enumerate(ccd_all)}

    primary = uf.upper()
    closed: list[CompositionClosed] = []
    closed_grupos: dict[str, str] = {}
    for idx, (code, desc, unit, grupo, sem_reg) in enumerate(csd_all):
        com_reg = ccd_by_idx.get(idx, sem_reg)
        merged: dict[str, dict[str, float]] = {}
        for u in set(com_reg) | set(sem_reg):
            sem_entry = sem_reg.get(u, {})
            com_entry = com_reg.get(u, {})
            merged[u] = {
                "comd": float(com_entry.get("price", 0.0)),
                "semd": float(sem_entry.get("price", 0.0)),
                "pct_as_comd": float(com_entry.get("pct_as", 0.0)),
                "pct_as_semd": float(sem_entry.get("pct_as", 0.0)),
            }
        price_com = merged.get(primary, {}).get("comd", 0.0)
        price_sem = merged.get(primary, {}).get("semd", price_com)
        if grupo:
            closed_grupos[code] = grupo
        closed.append(
            CompositionClosed(
                code=code,
                description=desc,
                unit=unit,
                price=price_com,
                price_sem_desoneracao=price_sem,
                regional=merged,
                grupo=grupo,
            )
        )

    isd_all = _parse_insumos_national_all_ufs(isd, origin="ISD") if isd else []
    icd_all = _parse_insumos_national_all_ufs(icd, origin="ICD") if icd else []
    icd_by_code = {code: reg for code, _, _, _, _, reg in icd_all}
    insumo_meta: dict[str, dict[str, str]] = {}

    insumos: list[InsumoRecord] = []
    for code, desc, unit, classificacao, origem_preco, sem_reg in isd_all:
        com_reg = icd_by_code.get(code, sem_reg)
        merged = {
            u: {"comd": com_reg.get(u, sem_reg.get(u, 0.0)), "semd": sem_reg.get(u, 0.0)}
            for u in set(com_reg) | set(sem_reg)
        }
        price_com = merged.get(primary, {}).get("comd", 0.0)
        price_sem = merged.get(primary, {}).get("semd", price_com)
        insumo_meta[code] = {
            "classificacao": classificacao,
            "origem_preco": origem_preco,
        }
        insumos.append(
            InsumoRecord(
                code=code,
                description=desc,
                unit=unit,
                price=price_com,
                price_sem_desoneracao=price_sem,
                origin="ICD" if code in icd_by_code else "ISD",
                regional=merged,
                classificacao=classificacao,
                origem_preco=origem_preco,
            )
        )

    comp_com = {c.code: c.price for c in closed}
    comp_sem = {c.code: c.price_sem_desoneracao or c.price for c in closed}
    ins_com = {i.code: i.price for i in insumos}
    ins_sem = {i.code: i.price_sem_desoneracao or i.price for i in insumos}
    open_map = (
        _parse_analytical_national(
            analitico,
            comp_com=comp_com,
            comp_sem=comp_sem,
            ins_com=ins_com,
            ins_sem=ins_sem,
            insumo_meta=insumo_meta,
            closed_grupos=closed_grupos,
        )
        if analitico
        else {}
    )

    for code, comp in open_map.items():
        row = next((c for c in closed if c.code == code), None)
        if row:
            if not comp.total_price:
                comp.total_price = row.price_for_uf(primary, sem=False)
            if not comp.total_price_sem:
                comp.total_price_sem = row.price_for_uf(primary, sem=True)

    _apply_sinapi_tp2(closed, open_map)

    labor_semd = _parse_labor_charges_sheet(csd, paired_uf_cols=True)
    labor_comd = _parse_labor_charges_sheet(ccd, paired_uf_cols=True) if ccd else {}
    labor_charges = _merge_labor_charges(labor_semd, labor_comd)

    return {
        "closed": closed,
        "open": open_map,
        "insumos": insumos,
        "format": "national",
        "all_ufs": True,
        "labor_charges": labor_charges,
    }


def _float_cell_legacy(value: object, default: float = 0.0) -> float:
    try:
        return float(str(value).replace(",", ".").strip() or default)
    except (TypeError, ValueError):
        return default


def _rows_from_matrix(headers: list, data_rows: list) -> list[dict[str, Any]]:
    headers_norm = [_norm(h) for h in headers]
    code_i = _pick_col(headers_norm, _CODE_KEYS)
    desc_i = _pick_col(headers_norm, _DESC_KEYS)
    unit_i = _pick_col(headers_norm, _UNIT_KEYS)
    price_i = _pick_col(headers_norm, _PRICE_KEYS)

    if desc_i is None and code_i is None:
        return []

    rows: list[dict[str, Any]] = []
    for row in data_rows:
        if not row:
            continue
        cells = list(row)
        code = cells[code_i] if code_i is not None and code_i < len(cells) else ""
        desc = cells[desc_i] if desc_i is not None and desc_i < len(cells) else ""
        if not str(desc or code).strip():
            continue
        unit = cells[unit_i] if unit_i is not None and unit_i < len(cells) else "un"
        price_raw = cells[price_i] if price_i is not None and price_i < len(cells) else 0
        rows.append(
            {
                "code": str(code or "").strip(),
                "description": str(desc or "").strip(),
                "unit": str(unit or "un").strip(),
                "price": _float_cell_legacy(price_raw),
            }
        )
    return rows


def _classify_item_type(type_raw: str, code: str, description: str) -> str:
    t = _norm(type_raw)
    blob = f"{t} {code} {description}".lower()
    if "compos" in t or "compos" in blob:
        return "composicao"
    if any(k in blob for k in ("mao de obra", "mão de obra", " mao ", "mod ")):
        return "mao_obra"
    if any(k in blob for k in ("equip", "equipamento")):
        return "equipamento"
    if t in ("m", "mo", "mao_obra"):
        return "mao_obra"
    if t in ("e", "eq", "equipamento"):
        return "equipamento"
    if t in ("c", "comp", "composicao"):
        return "composicao"
    return "insumo"


def _parse_analytical_sheet(matrix: list[tuple]) -> dict[str, CompositionOpen]:
    if not matrix:
        return {}

    headers = [str(h or f"col{i}") for i, h in enumerate(matrix[0])]
    headers_norm = [_norm(h) for h in headers]

    comp_code_i = _pick_col(headers_norm, _COMP_CODE_KEYS + _CODE_KEYS)
    comp_desc_i = _pick_col(headers_norm, _COMP_DESC_KEYS + _DESC_KEYS)
    comp_unit_i = _pick_col(headers_norm, _UNIT_KEYS)
    item_code_i = _pick_col(headers_norm, _ITEM_CODE_KEYS + _CODE_KEYS)
    item_desc_i = _pick_col(headers_norm, _ITEM_DESC_KEYS + _DESC_KEYS)
    item_unit_i = _pick_col(headers_norm, _ITEM_UNIT_KEYS + _UNIT_KEYS)
    coef_i = _pick_col(headers_norm, _COEF_KEYS)
    unit_price_i = _pick_col(headers_norm, _UNIT_PRICE_KEYS + _PRICE_KEYS)
    partial_i = _pick_col(headers_norm, _PARTIAL_KEYS)
    type_i = _pick_col(headers_norm, _TYPE_KEYS)
    total_i = _pick_col(headers_norm, ("custo total", "total composição", "total composicao"))

    compositions: dict[str, CompositionOpen] = {}
    current_code = ""
    current_desc = ""
    current_unit = "un"
    current_total = 0.0

    for row in matrix[1:]:
        if not row:
            continue
        cells = list(row)

        comp_code = (
            str(cells[comp_code_i]).strip()
            if comp_code_i is not None and comp_code_i < len(cells) and cells[comp_code_i]
            else ""
        )
        comp_desc = (
            str(cells[comp_desc_i]).strip()
            if comp_desc_i is not None and comp_desc_i < len(cells) and cells[comp_desc_i]
            else ""
        )
        if comp_code:
            current_code = comp_code
            if comp_desc:
                current_desc = comp_desc
            if comp_unit_i is not None and comp_unit_i < len(cells) and cells[comp_unit_i]:
                current_unit = str(cells[comp_unit_i]).strip() or current_unit
            if total_i is not None and total_i < len(cells):
                current_total = _float_cell_legacy(cells[total_i], current_total)

        if not current_code:
            continue

        item_code = (
            str(cells[item_code_i]).strip()
            if item_code_i is not None and item_code_i < len(cells) and cells[item_code_i]
            else ""
        )
        if item_code == current_code and item_desc_i is None:
            continue

        item_desc = (
            str(cells[item_desc_i]).strip()
            if item_desc_i is not None and item_desc_i < len(cells) and cells[item_desc_i]
            else ""
        )
        if not item_code and not item_desc:
            continue
        if item_code == current_code and not item_desc:
            continue

        item_unit = (
            str(cells[item_unit_i]).strip()
            if item_unit_i is not None and item_unit_i < len(cells) and cells[item_unit_i]
            else "un"
        )
        coef = _float_cell_legacy(cells[coef_i] if coef_i is not None and coef_i < len(cells) else 0)
        unit_price = _float_cell_legacy(
            cells[unit_price_i] if unit_price_i is not None and unit_price_i < len(cells) else 0
        )
        partial = _float_cell_legacy(
            cells[partial_i] if partial_i is not None and partial_i < len(cells) else 0
        )
        if partial == 0 and coef and unit_price:
            partial = round(coef * unit_price, 6)

        type_raw = (
            str(cells[type_i]).strip()
            if type_i is not None and type_i < len(cells) and cells[type_i]
            else ""
        )
        item_type = _classify_item_type(type_raw, item_code, item_desc)

        if current_code not in compositions:
            compositions[current_code] = CompositionOpen(
                code=current_code,
                description=current_desc or current_code,
                unit=current_unit,
                total_price=current_total,
                items=[],
            )
        comp = compositions[current_code]
        if comp_desc and not comp.description:
            comp.description = comp_desc
        if current_total and not comp.total_price:
            comp.total_price = current_total

        comp.items.append(
            CompositionItem(
                item_type=item_type,
                code=item_code or item_desc[:20],
                description=item_desc or item_code,
                unit=item_unit,
                coefficient=coef,
                unit_price=unit_price,
                partial_cost=partial,
            )
        )

    return compositions


def parse_sinapi_workbook(path: Path, *, prefer: str = "composicao") -> list[dict[str, Any]]:
    """Compat — retorna composições fechadas como linhas tabulares."""
    bank = parse_sinapi_full_workbook(path)
    if prefer == "insumo":
        return [i.to_dict() for i in bank["insumos"]]
    return [c.to_dict() for c in bank["closed"]]


def parse_sinapi_full_workbook(
    path: Path,
    *,
    uf: str = "SP",
    desonerado: bool = True,
) -> dict[str, Any]:
    """
    Extrai banco completo SINAPI Referência:
    - closed: composições fechadas (CSD/CCD ou CSD/CCD nacional por UF)
    - open: composições abertas analíticas (CPU)
    - insumos: catálogo de insumos (ISD/ICD)

    Formato 2025+: planilha nacional única — informe `uf` para preços regionais.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl necessário: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        csd_name = next((n for n in wb.sheetnames if _norm(n) == "csd"), None)
        if csd_name:
            probe = list(wb[csd_name].iter_rows(values_only=True))
            if _is_national_matrix(probe):
                return _parse_national_workbook(wb, uf=uf.upper())

        return _parse_legacy_workbook(wb, desonerado=desonerado)
    finally:
        wb.close()


def _parse_legacy_workbook(wb: Any, *, desonerado: bool) -> dict[str, Any]:
    closed: list[CompositionClosed] = []
    insumos: list[InsumoRecord] = []
    open_map: dict[str, CompositionOpen] = {}

    for sheet_name in wb.sheetnames:
        key = _norm(sheet_name)
        ws = wb[sheet_name]
        matrix = list(ws.iter_rows(values_only=True))
        if not matrix:
            continue

        headers = [str(h or f"col{i}") for i, h in enumerate(matrix[0])]

        if key in _COMPOSITION_CLOSED_SHEETS:
            for row in _rows_from_matrix(headers, matrix[1:]):
                closed.append(
                    CompositionClosed(
                        code=row["code"],
                        description=row["description"],
                        unit=row["unit"],
                        price=row["price"],
                        price_sem_desoneracao=row["price"],
                    )
                )
        elif key in _INSUMO_SHEETS:
            for row in _rows_from_matrix(headers, matrix[1:]):
                insumos.append(
                    InsumoRecord(
                        code=row["code"],
                        description=row["description"],
                        unit=row["unit"],
                        price=row["price"],
                        price_sem_desoneracao=row["price"],
                        origin=key.upper(),
                    )
                )
        elif key in _ANALYTICAL_SHEETS or "analit" in key:
            open_map.update(_parse_analytical_sheet(matrix))

    if not closed and open_map:
        for code, comp in open_map.items():
            total = comp.total_price or sum(i.partial_cost for i in comp.items)
            closed.append(
                CompositionClosed(
                    code=code,
                    description=comp.description,
                    unit=comp.unit,
                    price=round(total, 4),
                    price_sem_desoneracao=round(total, 4),
                )
            )

    if not open_map and closed:
        for c in closed:
            open_map[c.code] = CompositionOpen(
                code=c.code,
                description=c.description,
                unit=c.unit,
                total_price=c.price,
                items=[],
            )

    return {"closed": closed, "open": open_map, "insumos": insumos, "format": "legacy"}


def export_sinapi_csv(rows: list[dict[str, Any]], dest: Path) -> Path:
    import csv

    dest.parent.mkdir(parents=True, exist_ok=True)
    with open(dest, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=["code", "description", "unit", "price"])
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "code": row.get("code", ""),
                    "description": row.get("description", ""),
                    "unit": row.get("unit", "un"),
                    "price": row.get("price", 0),
                }
            )
    return dest
