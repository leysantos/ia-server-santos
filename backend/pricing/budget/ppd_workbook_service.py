from __future__ import annotations

import logging
import shutil
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from core.workflow.storage.client import get_workflow_storage
from pricing.budget.budget_session import BudgetSession, SESSION_STORE
from pricing.budget.ppd_exporter import sync_session_to_workbook
from pricing.budget.ppd_layout import PPD_TEMPLATE_ID
from pricing.budget.ppd_lo_export import export_workbook_bytes_to_pdf, libreoffice_available, libreoffice_info
from pricing.budget.ppd_sheet_registry import list_exportable_sheets
from pricing.budget.ppd_template_registry import PPD_SEMINF_2026_ID, resolve_template
from pricing.budget.ppd_workbook_init import (
    ensure_base_in_workbook,
    ensure_base_price_sheet,
    is_seminf_2026_workbook,
    workbook_sheetnames,
)
from pricing.budget.ppd_xlsx_assets import merge_workbook_preserving_assets

logger = logging.getLogger(__name__)

WORKBOOK_FILENAME = "workbook.xlsm"

DEFAULT_SEMINF_EXPORTABLE = [
    "MCQ",
    "ORC_SINTETICO",
    "ORC_ANALITICO",
    "CRONOGRAMA",
    "ESP_TECNICA",
]


def workbook_storage_key(session_id: str) -> str:
    return f"budgets/{session_id}/{WORKBOOK_FILENAME}"


def get_workbook_binding(session: BudgetSession) -> dict[str, Any] | None:
    binding = session.intent.get("ppd_workbook")
    return binding if isinstance(binding, dict) else None


def is_ppd_session(session: BudgetSession) -> bool:
    template = (session.project.template or session.intent.get("template") or "").upper()
    return template in (PPD_TEMPLATE_ID, PPD_SEMINF_2026_ID) or bool(get_workbook_binding(session))


def _exportable_sheet_keys(template_path: Path) -> list[str]:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(template_path, read_only=True)
        keys = [s["key"] for s in list_exportable_sheets(wb.sheetnames)]
        wb.close()
        return keys or list(DEFAULT_SEMINF_EXPORTABLE)
    except Exception:
        return list(DEFAULT_SEMINF_EXPORTABLE)


def _ensure_workbook_ready(session_id: str, binding: dict[str, Any] | None) -> dict[str, Any]:
    """Garante workbook ppd_seminf_abril_2026 — migra legado (v8.1/PLANILHA) automaticamente."""
    if not binding or not binding.get("storage_key"):
        return init_workbook(session_id)

    storage = get_workflow_storage()
    key = str(binding["storage_key"])
    if not storage.exists(key):
        return init_workbook(session_id, template_id=binding.get("template_id"))

    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        tmp.write(storage.get_bytes(key))

    try:
        names = workbook_sheetnames(tmp_path)
    finally:
        tmp_path.unlink(missing_ok=True)

    if is_seminf_2026_workbook(names):
        if binding.get("template_id") != PPD_SEMINF_2026_ID:
            binding = dict(binding)
            tmpl = resolve_template(PPD_SEMINF_2026_ID)
            binding["template_id"] = tmpl.id
            binding["template_version"] = tmpl.version
            binding["template_label"] = tmpl.label
            binding["exportable_sheets"] = _exportable_sheet_keys(tmpl.path)
        return binding

    logger.info("Workbook legado (sem ORC_SINTETICO) — migrando sessão %s", session_id[:8])
    return init_workbook(session_id, template_id=PPD_SEMINF_2026_ID, force=True)


def init_workbook(
    session_id: str,
    *,
    template_id: str | None = None,
    force: bool = False,
) -> dict[str, Any]:
    session = SESSION_STORE.get(session_id)
    if not session:
        raise KeyError(f"Sessão não encontrada: {session_id}")

    tmpl = resolve_template(template_id or PPD_SEMINF_2026_ID)
    storage = get_workflow_storage()
    key = workbook_storage_key(session_id)

    if force or not storage.exists(key):
        with tempfile.TemporaryDirectory(prefix="ppd-init-") as tmp:
            local = Path(tmp) / WORKBOOK_FILENAME
            shutil.copy(tmpl.path, local)
            if tmpl.base_template:
                ensure_base_price_sheet(local, base_template=tmpl.base_template)
            storage.put_file(
                key,
                local,
                content_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            )

    now = datetime.now(timezone.utc).isoformat()
    exportable = _exportable_sheet_keys(tmpl.path)
    binding: dict[str, Any] = {
        "template_id": tmpl.id,
        "template_version": tmpl.version,
        "template_label": tmpl.label,
        "storage_key": key,
        "storage_backend": storage.backend,
        "sync_status": "initialized",
        "last_sync_at": None,
        "initialized_at": now,
        "libreoffice_available": libreoffice_available(),
        "exportable_sheets": exportable,
        "is_seminf_2026": tmpl.id == PPD_SEMINF_2026_ID,
    }
    session.intent["ppd_workbook"] = binding
    session.updated_at = now
    return dict(binding)


def sync_workbook(session_id: str) -> dict[str, Any]:
    session = SESSION_STORE.get(session_id)
    if not session:
        raise KeyError(f"Sessão não encontrada: {session_id}")

    binding = get_workbook_binding(session)
    binding = _ensure_workbook_ready(session_id, binding)

    storage = get_workflow_storage()
    key = str(binding["storage_key"])

    for root in session.roots:
        root.recompute_total()

    tmpl = resolve_template(binding.get("template_id"))

    with tempfile.TemporaryDirectory(prefix="ppd-sync-") as tmp:
        local_path = Path(tmp) / WORKBOOK_FILENAME
        asset_source = Path(tmp) / "asset_source.xlsm"
        local_path.write_bytes(storage.get_bytes(key))
        shutil.copy(local_path, asset_source)

        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl necessário para sync PPD") from exc

        wb = openpyxl.load_workbook(local_path, keep_vba=True)
        if tmpl.base_template:
            ensure_base_in_workbook(wb, base_template=tmpl.base_template)

        sync_result = sync_session_to_workbook(
            wb,
            session.roots,
            session.project,
            schedule=session.schedule,
            tech_spec=session.tech_spec,
        )
        wb.save(local_path)
        wb.close()

        merged = merge_workbook_preserving_assets(asset_source.read_bytes(), local_path.read_bytes())
        local_path.write_bytes(merged)

        storage.put_file(
            key,
            local_path,
            content_type="application/vnd.ms-excel.sheet.macroEnabled.12",
        )

    now = datetime.now(timezone.utc).isoformat()
    binding.update(
        {
            "sync_status": "synced",
            "last_sync_at": now,
            "last_sync_rows": _count_sync_rows(session),
            "last_mcq_row": sync_result.get("last_mcq_row"),
            "last_cronograma_row": sync_result.get("last_cronograma_row"),
            "last_esp_row": sync_result.get("last_esp_row"),
            "libreoffice_available": libreoffice_available(),
            "exportable_sheets": binding.get("exportable_sheets") or list(DEFAULT_SEMINF_EXPORTABLE),
            "is_seminf_2026": True,
        }
    )
    session.intent["ppd_workbook"] = binding
    session.updated_at = now
    return dict(binding)


def get_workbook_bytes(session_id: str, *, auto_init: bool = True) -> bytes:
    session = SESSION_STORE.get(session_id)
    if not session:
        raise KeyError(f"Sessão não encontrada: {session_id}")

    binding = get_workbook_binding(session)
    if not binding and auto_init:
        binding = init_workbook(session_id)
    elif binding:
        binding = _ensure_workbook_ready(session_id, binding)
        session.intent["ppd_workbook"] = binding

    storage = get_workflow_storage()
    key = str((binding or {}).get("storage_key") or workbook_storage_key(session_id))
    if not storage.exists(key):
        init_workbook(session_id)
        key = workbook_storage_key(session_id)
    return storage.get_bytes(key)


def export_planilha_pdf(session_id: str, *, sync_first: bool = True) -> bytes:
    return export_workbook_sheet_pdf(session_id, sheet="ORC_SINTETICO", sync_first=sync_first)


def export_mcq_pdf(session_id: str, *, sync_first: bool = True) -> bytes:
    return export_workbook_sheet_pdf(session_id, sheet="MCQ", sync_first=sync_first)


def export_workbook_sheet_pdf(
    session_id: str,
    *,
    sheet: str,
    sync_first: bool = True,
) -> bytes:
    session = SESSION_STORE.get(session_id)
    if sync_first:
        sync_workbook(session_id)
        session = SESSION_STORE.get(session_id)
    data = get_workbook_bytes(session_id, auto_init=True)
    binding = get_workbook_binding(session) if session else {}
    last_row = _last_row_for_sheet(binding or {}, sheet)
    return export_workbook_bytes_to_pdf(data, sheet=sheet, last_data_row=last_row)


def _last_row_for_sheet(binding: dict[str, Any], sheet: str) -> int | None:
    key = sheet.strip().upper()
    mapping = {
        "MCQ": "last_mcq_row",
        "ORC_SINTETICO": "last_mcq_row",
        "PLANILHA": "last_mcq_row",
        "ORC_ANALITICO": "last_mcq_row",
        "CRONOGRAMA": "last_cronograma_row",
        "ESP_TECNICA": "last_esp_row",
    }
    alias = {
        "OR": "last_mcq_row",
        "SINTETICO": "last_mcq_row",
        "ANALITICO": "last_mcq_row",
        "ESP": "last_esp_row",
    }
    field = mapping.get(key) or alias.get(key)
    if not field:
        return None
    try:
        val = binding.get(field)
        return int(val) if val else None
    except (TypeError, ValueError):
        return None


def workbook_status(session_id: str) -> dict[str, Any]:
    session = SESSION_STORE.get(session_id)
    if not session:
        raise KeyError(f"Sessão não encontrada: {session_id}")
    binding = get_workbook_binding(session) or {}
    storage = get_workflow_storage()
    key = binding.get("storage_key") or workbook_storage_key(session_id)
    lo = libreoffice_info()

    exportable = binding.get("exportable_sheets")
    is_seminf = binding.get("is_seminf_2026")
    if storage.exists(str(key)):
        with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp:
            tmp_path = Path(tmp.name)
            tmp.write(storage.get_bytes(str(key)))
        try:
            names = workbook_sheetnames(tmp_path)
            is_seminf = is_seminf_2026_workbook(names)
            if is_seminf:
                exportable = list(DEFAULT_SEMINF_EXPORTABLE)
            elif not exportable:
                exportable = [s["key"] for s in list_exportable_sheets(names)]
        finally:
            tmp_path.unlink(missing_ok=True)

    return {
        "session_id": session_id,
        "is_ppd": is_ppd_session(session),
        "workbook_exists": storage.exists(str(key)),
        "is_seminf_2026": bool(is_seminf),
        "libreoffice_available": lo["available"],
        "libreoffice_path": lo.get("path") or None,
        "libreoffice_version": lo.get("version") or None,
        "exportable_sheets": exportable or list(DEFAULT_SEMINF_EXPORTABLE),
        "binding": binding or None,
    }


def _count_sync_rows(session: BudgetSession) -> int:
    from pricing.budget.ppd_layout import ROW_TYPE_ETAPA, ROW_TYPE_SERVICO, ROW_TYPE_SUB_ETAPA

    count = 0

    def walk(nodes: list) -> None:
        nonlocal count
        for node in nodes:
            if node.row_type in (ROW_TYPE_ETAPA, ROW_TYPE_SUB_ETAPA, ROW_TYPE_SERVICO):
                count += 1
            if node.calculation_note:
                count += 1
            walk(node.children)

    walk(session.roots)
    return count
