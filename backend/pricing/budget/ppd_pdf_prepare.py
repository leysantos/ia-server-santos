from __future__ import annotations

import logging
import shutil
import tempfile
from pathlib import Path

from pricing.budget.ppd_sheet_registry import (
    resolve_workbook_sheet,
    sheet_spec_for,
)
from pricing.budget.ppd_xlsx_assets import merge_workbook_preserving_assets

logger = logging.getLogger(__name__)


def _base_sheet_names(sheetnames: list[str]) -> list[str]:
    return [n for n in sheetnames if n.startswith("Base_")]


def _sheets_to_keep(sheetnames: list[str], target: str, spec) -> set[str]:
    target_name = resolve_workbook_sheet(sheetnames, target)
    if not target_name:
        raise ValueError(f"Aba {target} ausente")
    keep = {target_name}
    upper_map = {n.upper(): n for n in sheetnames}

    if spec and spec.needs_mcq and "MCQ" in upper_map:
        keep.add(upper_map["MCQ"])
    if spec and spec.needs_base:
        keep.update(_base_sheet_names(sheetnames))

    return keep


def _cell_has_content(value) -> bool:
    if value in (None, ""):
        return False
    if isinstance(value, str) and value.startswith("="):
        return False
    return True


def _last_data_row(ws, spec, *, start_row: int = 21) -> int:
    cols = spec.data_cols
    last = start_row
    max_row = min(ws.max_row or start_row, start_row + 2500)
    for row_idx in range(start_row, max_row + 1):
        if any(_cell_has_content(ws.cell(row_idx, col).value) for col in cols):
            last = row_idx
    return last


def _apply_print_setup(ws, spec, sheet_name: str, last_row: int) -> str:
    start_col = spec.print_start_col
    end_col = spec.print_end_col
    footer_pad = spec.footer_pad
    safe_name = sheet_name.replace("'", "''")
    print_area = f"'{safe_name}'!${start_col}$11:${end_col}${last_row + footer_pad}"
    if spec.key in ("CRONOGRAMA", "ESP_TECNICA"):
        print_area = f"'{safe_name}'!${start_col}$1:${end_col}${last_row + footer_pad}"

    ws.print_area = print_area

    if spec.key in ("ORC_SINTETICO", "ORC_ANALITICO", "PLANILHA"):
        ws.page_setup.orientation = "landscape"
    elif spec.key == "MCQ":
        ws.page_setup.orientation = "portrait"
    elif spec.key == "CRONOGRAMA":
        ws.page_setup.orientation = "landscape"

    try:
        from openpyxl.worksheet.properties import PageSetupProperties

        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    except Exception:
        pass
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return print_area


def prepare_single_sheet_pdf_workbook(
    source: Path,
    sheet_name: str,
    *,
    last_data_row: int | None = None,
) -> Path:
    """
    Cópia para PDF: mantém abas referenciadas, preserva logo/cabeçalho (&G),
    orientação paisagem em ORC_* e área de impressão ajustada.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment
    except ImportError as exc:
        raise ImportError("openpyxl necessário") from exc

    asset_source = source.read_bytes()

    tmp_dir = Path(tempfile.mkdtemp(prefix="ppd-pdf-prep-"))
    dest = tmp_dir / "export.xlsm"
    shutil.copy(source, dest)

    wb = openpyxl.load_workbook(dest, keep_vba=True)
    resolved = resolve_workbook_sheet(wb.sheetnames, sheet_name)
    if not resolved:
        wb.close()
        raise ValueError(f"Aba {sheet_name} ausente no workbook")

    spec = sheet_spec_for(resolved)
    if not spec:
        wb.close()
        raise ValueError(f"Aba não suportada para PDF: {sheet_name}")

    keep = _sheets_to_keep(wb.sheetnames, sheet_name, spec)
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]

    ws = wb[resolved]
    start_row = 1 if spec.key in ("CRONOGRAMA", "ESP_TECNICA") else 21
    if last_data_row and last_data_row >= start_row:
        last_row = last_data_row
    else:
        last_row = _last_data_row(ws, spec, start_row=start_row)

    print_area = _apply_print_setup(ws, spec, resolved, last_row)

    if spec.key in ("MCQ", "ORC_SINTETICO", "ORC_ANALITICO", "PLANILHA"):
        wrap = Alignment(wrap_text=True, vertical="top")
        desc_col = 10
        data_start = 21
        for row_idx in range(data_start, last_row + 1):
            cell = ws.cell(row_idx, desc_col)
            if cell.value not in (None, ""):
                cell.alignment = wrap
        if spec.key == "MCQ":
            ws.column_dimensions["J"].width = max(ws.column_dimensions["J"].width or 0, 48)

    for name in wb.sheetnames:
        hidden_ws = wb[name]
        if name != resolved:
            hidden_ws.sheet_state = "hidden"
            hidden_ws.print_area = None

    wb.active = wb.sheetnames.index(resolved)
    modified_path = tmp_dir / "modified.xlsm"
    wb.save(modified_path)
    wb.close()

    merged = merge_workbook_preserving_assets(asset_source, modified_path)
    dest.write_bytes(merged)

    logger.info(
        "PDF prep %s: keep=%s last_row=%s print_area=%s",
        resolved,
        sorted(keep),
        last_row,
        print_area,
    )
    return dest


def cleanup_prepared_workbook(path: Path) -> None:
    try:
        if path.parent.name.startswith("ppd-pdf-prep-"):
            shutil.rmtree(path.parent, ignore_errors=True)
    except OSError as exc:
        logger.debug("cleanup prepared workbook: %s", exc)
