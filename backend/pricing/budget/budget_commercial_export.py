"""Exportação CPQ — proposta comercial com margem (B21)."""

from __future__ import annotations

from typing import Any

from pricing.budget.budget_export_branding import ExportBrandingConfig
from pricing.budget.budget_xlsx_builder import export_budget_document_xlsx
from pricing.budget.budget_pdf_export import export_budget_pdf
from pricing.models.budget_item import BudgetItem
from pricing.models.budget_metadata import BudgetProjectMetadata


def commercial_totals(roots: list[BudgetItem], meta: BudgetProjectMetadata) -> dict[str, float]:
    base = round(sum(r.effective_total() for r in roots), 2)
    margin_pct = max(0.0, float(meta.commercial_margin_pct or 0))
    margin_value = round(base * margin_pct / 100, 2)
    return {
        "cost_total": base,
        "margin_pct": margin_pct,
        "margin_value": margin_value,
        "proposal_total": round(base + margin_value, 2),
    }


def export_proposta_comercial_xlsx(
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata,
    *,
    branding: ExportBrandingConfig | None = None,
    schedule: Any | None = None,
    logo_bytes: bytes | None = None,
    company_profile: Any | None = None,
) -> bytes:
    """Reutiliza orçamento sintético + rodapé de margem comercial."""
    try:
        import io

        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl necessário") from exc

    base_bytes = export_budget_document_xlsx(
        "orc_sintetico",
        roots,
        metadata,
        branding=branding,
        schedule=schedule,
        logo_bytes=logo_bytes,
        company_profile=company_profile,
    )
    totals = commercial_totals(roots, metadata)
    wb = openpyxl.load_workbook(io.BytesIO(base_bytes))
    ws = wb.active
    ws.title = "PROPOSTA"
    row = ws.max_row + 2
    client = (metadata.commercial_client or metadata.empresa or "").strip()
    if client:
        ws.cell(row=row, column=1, value=f"Cliente: {client}")
        row += 1
    ws.cell(row=row, column=1, value=f"Custo direto (sem margem): R$ {totals['cost_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    row += 1
    ws.cell(row=row, column=1, value=f"Margem comercial ({totals['margin_pct']:.2f}%): R$ {totals['margin_value']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    row += 1
    ws.cell(row=row, column=1, value=f"TOTAL PROPOSTA: R$ {totals['proposal_total']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_proposta_comercial_pdf(
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata,
    *,
    branding: ExportBrandingConfig | None = None,
    schedule: Any | None = None,
    logo_bytes: bytes | None = None,
    company_profile: Any | None = None,
) -> bytes:
    """PDF sintético com totais de margem no rodapé via export nativo."""
    return export_budget_pdf(
        "orc_sintetico",
        roots,
        metadata,
        branding=branding,
        schedule=schedule,
        logo_bytes=logo_bytes,
        company_profile=company_profile,
    )
