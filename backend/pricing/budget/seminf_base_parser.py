"""Parser da aba Base das planilhas DP/SEMINF (MC_OR ou Tabela_Preco)."""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from pricing.budget.ppd_parser import extract_price_base_rows, pick_base_sheet_name

_MONTHS_PT: dict[str, int] = {
    "janeiro": 1,
    "jan": 1,
    "fevereiro": 2,
    "fev": 2,
    "marco": 3,
    "março": 3,
    "marc": 3,
    "abril": 4,
    "abr": 4,
    "maio": 5,
    "mai": 5,
    "junho": 6,
    "jun": 6,
    "julho": 7,
    "jul": 7,
    "agosto": 8,
    "ago": 8,
    "setembro": 9,
    "set": 9,
    "outubro": 10,
    "out": 10,
    "novembro": 11,
    "nov": 11,
    "dezembro": 12,
    "dez": 12,
}

_MC_OR_SHEETS = frozenset({"PLANILHA", "MCQ", "ETAPAS", "CURVA_ABC"})


def load_workbook_sheet_names(path: str | Path) -> list[str]:
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def detect_workbook_format(path: Path, sheet_names: list[str]) -> str:
    """
    tabela_preco — planilha leve só com aba(s) Base (ex.: Tabela_Preco-Abril2026.xlsm).
    mc_or — modelo completo de orçamento (PLANILHA, MCQ, ETAPAS…).
    """
    stem = path.stem.lower().replace("-", "_")
    if "tabela_preco" in stem:
        return "tabela_preco"
    if "mc_or" in stem or "mod_mc" in stem:
        return "mc_or"
    upper = {s.upper() for s in sheet_names}
    if upper & _MC_OR_SHEETS:
        return "mc_or"
    return "tabela_preco"


def find_base_sheet_name(sheet_names: list[str]) -> str | None:
    return pick_base_sheet_name(sheet_names)


def parse_base_sheet_period(name: str) -> tuple[int | None, int | None]:
    """Extrai ano/mês de Base_Abril-2026-copia, Tabela_Preco-Abril2026, etc."""
    normalized = unicodedata.normalize("NFKD", (name or "").lower())
    normalized = "".join(c for c in normalized if not unicodedata.combining(c))
    year_match = re.search(r"(20\d{2})", normalized)
    year = int(year_match.group(1)) if year_match else None
    month: int | None = None
    for token, num in _MONTHS_PT.items():
        if token in normalized:
            month = num
            break
    return year, month


def resolve_reference_period(
    path: Path,
    base_sheet: str,
    *,
    year: int | None = None,
    month: int | None = None,
) -> tuple[int | None, int | None]:
    sheet_year, sheet_month = parse_base_sheet_period(base_sheet)
    file_year, file_month = parse_base_sheet_period(path.stem)
    return (
        year or sheet_year or file_year,
        month or sheet_month or file_month,
    )


def infer_seminf_reference(
    path: Path,
    *,
    base_sheet: str | None = None,
    year: int | None = None,
    month: int | None = None,
    source_slug: str = "SEMINF",
) -> str:
    ref_year, ref_month = resolve_reference_period(
        path,
        base_sheet or "",
        year=year,
        month=month,
    )
    if ref_year and ref_month:
        return f"BR-{source_slug}-{ref_year}-{ref_month:02d}"
    return f"BR-{source_slug}-{path.stem}"


def is_seminf_regional_code(code: str) -> bool:
    """Código municipal regional (ex.: 107071.1.9.SEMINF)."""
    from pricing.budget.seminf_open_parser import normalize_seminf_code

    normalized = normalize_seminf_code(code)
    return normalized.endswith(".SEMINF")


def is_sinapi_national_code(code: str) -> bool:
    """Código SINAPI puro (numérico) presente na aba Base mas servido pelo price bank Caixa."""
    raw = (code or "").strip()
    if not raw or is_seminf_regional_code(raw):
        return False
    if raw.isdigit():
        return True
    return bool(re.match(r"^\d+(?:\.\d+)*$", raw))


def tp2_lookup_key(code: str) -> str:
    """Chave para cruzar item de CPU com a coluna tp2 da aba Base."""
    from pricing.budget.seminf_open_parser import normalize_seminf_code

    raw = str(code or "").strip()
    if is_seminf_regional_code(raw):
        return normalize_seminf_code(raw)
    if raw.isdigit():
        return raw
    return raw


def build_tp2_index_from_base_rows(rows: list[dict[str, Any]]) -> dict[str, str]:
    """Mapa código → tp2 (ex. AS) a partir das linhas da aba Base."""
    index: dict[str, str] = {}
    for row in rows:
        meta = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        tp2 = str(meta.get("tp2") or "").strip()
        if not tp2:
            continue
        code = str(row.get("code") or "").strip()
        if not code:
            continue
        index[tp2_lookup_key(code)] = tp2
    return index


def build_tp2_index_from_workbook(path: str | Path, *, sheet_name: str | None = None) -> dict[str, str]:
    rows = extract_price_base_rows(path, sheet_name=sheet_name)
    return build_tp2_index_from_base_rows(rows)


def validate_seminf_bundle_period(
    paths: list[Path],
    *,
    year: int | None,
    month: int | None,
) -> None:
    """Garante que todos os arquivos do lote correspondem ao período selecionado."""
    if not year or not month:
        return
    for path in paths:
        file_year, file_month = resolve_reference_period(path, path.stem)
        if file_year and file_month and (file_year != year or file_month != month):
            raise ValueError(
                f"Arquivo {path.name} é de {file_month:02d}/{file_year}, "
                f"mas o período selecionado é {month:02d}/{year}."
            )


def find_seminf_bundle_in_dir(
    folder: str | Path,
    *,
    year: int | None = None,
    month: int | None = None,
) -> dict[str, Path]:
    """Reexport — detecção inteligente em pricing.budget.seminf_bundle_detect."""
    from pricing.budget.seminf_bundle_detect import find_seminf_bundle_in_dir as _find

    return _find(folder, year=year, month=month)


def extract_seminf_base_compositions(
    path: str | Path,
    *,
    uf: str = "AM",
    only_seminf_codes: bool = True,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """
    Lê composições fechadas da aba Base_* (CODIGO, DESCRICAO, und, ComD, SemD).

    Suporta:
    - Tabela_Preco-*.xlsm (modelo leve, só preços)
    - 00_MOD_MC_OR_*.xlsm (modelo completo — mesma aba Base)

    Por padrão importa só códigos `*.SEMINF` — códigos SINAPI ficam no price bank Caixa.
    """
    path = Path(path)
    sheet_names = load_workbook_sheet_names(path)
    base_sheet = pick_base_sheet_name(sheet_names) or ""
    workbook_format = detect_workbook_format(path, sheet_names)

    if not base_sheet:
        return [], {
            "error": "base_sheet_missing",
            "path": str(path),
            "sheets": sheet_names,
            "workbook_format": workbook_format,
        }

    rows = extract_price_base_rows(path, sheet_name=base_sheet)
    if not rows:
        return [], {
            "error": "base_sheet_empty",
            "path": str(path),
            "base_sheet": base_sheet,
            "workbook_format": workbook_format,
        }

    sheet_year, sheet_month = resolve_reference_period(path, base_sheet)

    closed: list[dict[str, Any]] = []
    skipped_sinapi = 0
    uf_key = (uf or "AM").upper()
    for row in rows:
        code = str(row.get("code") or "").strip()
        if not code:
            continue
        if only_seminf_codes and not is_seminf_regional_code(code):
            if is_sinapi_national_code(code):
                skipped_sinapi += 1
            continue
        from pricing.budget.seminf_open_parser import normalize_seminf_code

        code = normalize_seminf_code(code)
        price_com = float(row.get("price") or 0)
        meta = row.get("metadata") if isinstance(row.get("metadata"), dict) else {}
        price_sem = float(meta.get("price_sem_desoneracao") or price_com)
        tp2 = str(meta.get("tp2") or "").strip()
        closed.append(
            {
                "code": code,
                "description": str(row.get("description") or ""),
                "unit": str(row.get("unit") or "un"),
                "price": price_com,
                "price_sem_desoneracao": price_sem,
                "regional": {uf_key: {"comd": price_com, "semd": price_sem}},
                "tp2": tp2,
            }
        )

    seminf_codes = len(closed)
    metadata: dict[str, Any] = {
        "publisher": "SEMINF-AM",
        "region": "Manaus/Amazonas",
        "workbook_format": workbook_format,
        "base_sheet": base_sheet,
        "base_items": seminf_codes,
        "base_items_total_in_sheet": len(rows),
        "seminf_regional_codes": seminf_codes,
        "sinapi_codes_skipped": skipped_sinapi,
        "import_filter": "seminf_only" if only_seminf_codes else "all",
        "sheet_year": sheet_year,
        "sheet_month": sheet_month,
        "source_file": str(path.resolve()),
        "sheets": sheet_names,
    }
    return closed, metadata
