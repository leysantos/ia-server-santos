"""Parser das planilhas Composicao-Seminf-* (aba CPUs — composições abertas SEMINF)."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from pricing.budget.price_bank_store import CompositionItem, CompositionOpen

_PROPRIO_BANKS = frozenset({"próprio", "proprio"})


def normalize_seminf_code(code: str) -> str:
    """Normaliza códigos regionais (espaços, SEMIINF, ponto antes de SEMINF)."""
    raw = re.sub(r"\s+", "", (code or "").strip().upper())
    raw = raw.replace("SEMIINF", "SEMINF")
    raw = re.sub(r"\.+", ".", raw)
    raw = re.sub(r"(\d)SEMINF$", r"\1.SEMINF", raw)
    if raw.endswith("SEMINF") and not raw.endswith(".SEMINF"):
        raw = f"{raw[:-6]}.SEMINF"
    return raw


def is_seminf_proprio_row(row: tuple[Any, ...] | list[Any]) -> bool:
    """Linha principal de composição regional (Banco = Próprio)."""
    if not row or len(row) < 4:
        return False
    if str(row[0] or "").strip() != "Composição":
        return False
    return str(row[2] or "").strip().lower() in _PROPRIO_BANKS


def detect_seminf_open_desoneracao(path: Path) -> str | None:
    """comd | semd — inferido pelo nome do arquivo ou cabeçalho da aba CPUs."""
    stem = path.stem.lower()
    if "semd" in stem or "sem-d" in stem or "sem_d" in stem:
        return "semd"
    if "comd" in stem or "com-d" in stem or "com_d" in stem:
        return "comd"
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = "CPUs" if "CPUs" in wb.sheetnames else wb.sheetnames[0]
        row2 = next(wb[sheet].iter_rows(min_row=2, max_row=2, values_only=True), None)
        wb.close()
        if row2:
            blob = " ".join(str(c or "") for c in row2).lower()
            if "não desonerado" in blob or "nao desonerado" in blob:
                return "semd"
            if "desonerado" in blob:
                return "comd"
    except Exception:
        pass
    return None


def _float_cell(value: object, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def _classify_aux_item(row: tuple[Any, ...] | list[Any], item_code: str) -> str:
    unit = str(row[6] if len(row) > 6 else "").strip().upper()
    tipo = str(row[4] if len(row) > 4 else "").lower()
    desc = str(row[3] if len(row) > 3 else "").lower()
    if unit == "H" or "mao" in tipo or "encargos" in desc or "pedreiro" in desc or "servente" in desc:
        return "mao_obra"
    if "equip" in tipo or "equip" in desc:
        return "equipamento"
    if item_code.isdigit():
        return "composicao"
    return "insumo"


def parse_seminf_open_workbook(path: str | Path) -> dict[str, dict[str, Any]]:
    """
    Lê aba CPUs — retorna mapa código SEMINF normalizado → composição aberta parcial.

    Estrutura da planilha (Composicao-Seminf-*-ComD/SemD.xlsx):
    - Linha principal: Composição | código | Próprio | descrição | … | und | qtd | V.Unit | Total
    - Itens: Composição Auxiliar | código SINAPI | SINAPI | … | und | coef | V.Unit | Total
    """
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = "CPUs" if "CPUs" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    comps: dict[str, dict[str, Any]] = {}
    current_key = ""

    for row in ws.iter_rows(values_only=True):
        if is_seminf_proprio_row(row):
            code = normalize_seminf_code(str(row[1] or ""))
            if not code.endswith(".SEMINF"):
                current_key = ""
                continue
            current_key = code
            total = _float_cell(row[9] if len(row) > 9 else None)
            if total <= 0:
                total = _float_cell(row[8] if len(row) > 8 else None)
            comps[code] = {
                "code": code,
                "description": str(row[3] or "").strip(),
                "unit": str(row[6] or "un").strip() or "un",
                "total": total,
                "items": [],
            }
            continue

        if not current_key or current_key not in comps:
            continue

        kind = str(row[0] or "").strip()
        if kind not in ("Composição Auxiliar", "Insumo"):
            continue

        item_code = str(row[1] or "").strip()
        coef = _float_cell(row[7] if len(row) > 7 else 0)
        unit_price = _float_cell(row[8] if len(row) > 8 else 0)
        partial = _float_cell(row[9] if len(row) > 9 else 0)
        if partial <= 0 and coef > 0 and unit_price > 0:
            partial = round(coef * unit_price, 4)

        item_type = "insumo" if kind == "Insumo" else _classify_aux_item(row, item_code)
        if kind == "Insumo":
            bank = str(row[2] or "").strip().lower()
            tipo = str(row[4] if len(row) > 4 else "").lower()
            if "equip" in tipo:
                item_type = "equipamento"
            elif bank == "sinapi" and item_code.isdigit():
                item_type = "insumo"

        comps[current_key]["items"].append(
            {
                "item_type": item_type,
                "code": item_code,
                "description": str(row[3] or "").strip(),
                "unit": str(row[6] or "").strip(),
                "coefficient": coef,
                "unit_price": unit_price,
                "partial_cost": partial,
            }
        )

    wb.close()
    return comps


def merge_seminf_open_compositions(
    comd_map: dict[str, dict[str, Any]],
    semd_map: dict[str, dict[str, Any]],
    *,
    tp2_index: dict[str, str] | None = None,
    sinapi_as_index: dict[str, str] | None = None,
) -> dict[str, CompositionOpen]:
    """Combina planilhas ComD + SemD em CompositionOpen com preços por item."""
    from pricing.budget.seminf_base_parser import tp2_lookup_key
    from pricing.budget.tp2_as import AS_MARKER, merge_tp2

    codes = sorted(set(comd_map) | set(semd_map))
    open_map: dict[str, CompositionOpen] = {}
    tp2_lookup = tp2_index or {}
    sinapi_as = sinapi_as_index or {}

    for code in codes:
        comd = comd_map.get(code)
        semd = semd_map.get(code)
        base = comd or semd
        if not base:
            continue

        semd_items = {str(i.get("code", "")).strip(): i for i in (semd or {}).get("items", [])}
        items: list[CompositionItem] = []
        for idx, item in enumerate(base.get("items", [])):
            item_code = str(item.get("code", "")).strip()
            sem_item = semd_items.get(item_code)
            if sem_item is None and idx < len((semd or {}).get("items", [])):
                sem_item = semd["items"][idx]

            unit_com = float(item.get("unit_price") or 0)
            unit_sem = float((sem_item or {}).get("unit_price") or unit_com)
            partial_com = float(item.get("partial_cost") or 0)
            partial_sem = float((sem_item or {}).get("partial_cost") or partial_com)
            if partial_sem <= 0 and item.get("coefficient") and unit_sem:
                partial_sem = round(float(item["coefficient"]) * unit_sem, 4)

            items.append(
                CompositionItem(
                    item_type=str(item.get("item_type") or "insumo"),
                    code=item_code or str(item.get("description", ""))[:20],
                    description=str(item.get("description") or ""),
                    unit=str(item.get("unit") or ""),
                    coefficient=float(item.get("coefficient") or 0),
                    unit_price=unit_com,
                    partial_cost=partial_com,
                    unit_price_sem=unit_sem,
                    partial_cost_sem=partial_sem,
                    tp2=merge_tp2(
                        tp2_lookup.get(tp2_lookup_key(item_code), "")
                        or sinapi_as.get(item_code, ""),
                        1.0 if sinapi_as.get(item_code) == AS_MARKER else 0.0,
                    ),
                )
            )

        total_com = float((comd or {}).get("total") or 0)
        total_sem = float((semd or {}).get("total") or total_com)
        if total_com <= 0 and items:
            total_com = round(sum(i.partial_cost for i in items), 4)
        if total_sem <= 0 and items:
            total_sem = round(sum(i.partial_cost_sem for i in items), 4)

        comp_tp2 = AS_MARKER if any(i.tp2 == AS_MARKER for i in items) else ""

        open_map[code] = CompositionOpen(
            code=code,
            description=str(base.get("description") or code),
            unit=str(base.get("unit") or "un"),
            total_price=total_com,
            total_price_sem=total_sem,
            tp2=comp_tp2,
            items=items,
        )

    return open_map
