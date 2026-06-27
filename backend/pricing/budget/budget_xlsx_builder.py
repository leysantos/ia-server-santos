"""Geração nativa de planilhas Excel (openpyxl) com logo e cabeçalho personalizados."""

from __future__ import annotations

import io
import tempfile
from typing import Any

from core.system.company_profile import CompanyProfile, get_company_profile
from pricing.budget.budget_export_branding import (
    ExportBrandingConfig,
    EXPORT_DOC_TYPES,
    load_logo_bytes,
)
from pricing.budget.budget_export_tables import (
    ExportTableData,
    budget_desoneracao_mode,
    build_export_table,
    _budget_row_tipo,
    _cpu_tipo_label,
    _fetch_open_composition_items,
    _is_service_row,
    _resolve_open_composition_lookup,
)
from pricing.budget.budget_pdf_export import DOC_TITLES
from pricing.budget.budget_pdf_landscape_template import (
    format_obra_meta_obra_lines,
    format_obra_meta_pricing_lines,
)
from pricing.budget.ppd_layout import (
    COL_DESCRIPTION,
    COL_ITEM,
    COL_QUANTITY,
    COL_ROW_TYPE,
    COL_SOURCE_CODE,
    COL_TOTAL_COMD,
    COL_TOTAL_SEMD,
    COL_UNIT,
    COL_UNIT_COST_COMD,
    COL_UNIT_COST_SEMD,
    COL_UNIT_PRICE_BDI_COMD,
    COL_UNIT_PRICE_BDI_SEMD,
    ROW_TYPE_ETAPA,
    ROW_TYPE_SERVICO,
    ROW_TYPE_SUB_ETAPA,
)
from pricing.budget.ppd_exporter import (
    CRONOGRAMA_ETAPA_MAX_ROW,
    CRONOGRAMA_ETAPA_START,
    CRONOGRAMA_ETAPA_STEP,
    _schedule_total_days,
)
from pricing.models.budget_item import BudgetItem
from pricing.models.budget_metadata import BudgetProjectMetadata

_XLSX_HEADER_BLUE = "1F4E79"
_XLSX_FONT_NAME = "Arial Narrow"
_XLSX_NUM_FMT = "#,##0.00"


def _xlsx_font(size: int = 10, *, bold: bool = False, color: str | None = None):
    from openpyxl.styles import Font

    kwargs: dict[str, Any] = {"name": _XLSX_FONT_NAME, "size": size, "bold": bold}
    if color:
        kwargs["color"] = color
    return Font(**kwargs)


def _apply_numeric_cell_format(cell, col_idx: int, right_cols: tuple[int, ...]) -> None:
    if isinstance(cell.value, (int, float)) and (col_idx - 1) in right_cols:
        cell.number_format = _XLSX_NUM_FMT


def _apply_worksheet_arial_narrow(ws) -> None:
    """Garante Arial Narrow em todas as células e formato numérico com 2 decimais."""
    from openpyxl.styles import Font

    for row in ws.iter_rows():
        for cell in row:
            f = cell.font
            cell.font = Font(
                name=_XLSX_FONT_NAME,
                size=f.size or 10,
                bold=f.bold,
                italic=f.italic,
                vertAlign=f.vertAlign,
                underline=f.underline,
                strike=f.strike,
                color=f.color,
            )
            if isinstance(cell.value, (int, float)) and cell.number_format in ("General", "0"):
                cell.number_format = _XLSX_NUM_FMT


def _doc_table_col_count(doc_type: str) -> int:
    return {
        "orc_sintetico": 7,
        "orc_analitico": 8,
        "mcq": 5,
        "cronograma": 3,
        "esp_tecnica": 8,
        "curva_abc": 7,
        "curva_s": 7,
        "histograma": 7,
    }.get(doc_type.strip().lower(), 8)


def _doc_header_col_count(doc_type: str) -> int:
    """Largura do cabeçalho/rodapé institucional (pode ser maior que a tabela de dados)."""
    key = doc_type.strip().lower()
    if key in ("mcq", "esp_tecnica"):
        return 8
    return _doc_table_col_count(key)


_META_OBRA_END_COL = 3  # A:C — bloco obra/objeto (padrão)
_META_PRICE_START_COL = 4  # D — bases de preços e BDI (padrão)
_META_ANALITICO_OBRA_END_COL = 4  # A:D
_META_ANALITICO_PRICE_START_COL = 5  # E
_META_ANALITICO_PRICE_END_COL = 9  # I


def _meta_obra_end_col(table_cols: int, doc_type: str | None = None) -> int:
    if (doc_type or "").strip().lower() == "orc_analitico":
        return _META_ANALITICO_OBRA_END_COL
    if table_cols <= 3:
        return table_cols
    return min(_META_OBRA_END_COL, table_cols - 1)


def _meta_price_start_col(table_cols: int, doc_type: str | None = None) -> int:
    if (doc_type or "").strip().lower() == "orc_analitico":
        return _META_ANALITICO_PRICE_START_COL
    if table_cols <= 3:
        return 1
    return min(_META_PRICE_START_COL, table_cols)


def _meta_price_end_col(table_cols: int, doc_type: str | None = None) -> int:
    if (doc_type or "").strip().lower() == "orc_analitico":
        return _META_ANALITICO_PRICE_END_COL
    return table_cols


def _set_merged_cell(
    ws,
    row: int,
    col_start: int,
    col_end: int,
    value: str | None,
    *,
    font=None,
    alignment=None,
    fill=None,
    border=None,
) -> None:
    from openpyxl.styles import Alignment

    if col_end > col_start:
        ws.merge_cells(
            start_row=row,
            start_column=col_start,
            end_row=row,
            end_column=col_end,
        )
    cell = ws.cell(row=row, column=col_start, value=value if value else None)
    if font is not None:
        cell.font = font
    cell.alignment = alignment or Alignment(vertical="center", wrap_text=True)
    if fill is not None:
        cell.fill = fill
    if border is not None:
        cell.border = border


def _institutional_subtitle_lines(
    brand: ExportBrandingConfig,
    profile: CompanyProfile,
) -> list[str]:
    lines: list[str] = []
    for raw in (
        brand.header_line1 or profile.display_name() or profile.razao_social,
        brand.header_line2 or "",
        brand.header_line3 or "",
    ):
        text = str(raw or "").strip()
        if text:
            lines.append(text)
    if not any("cnpj" in ln.lower() for ln in lines):
        cnpj = str(profile.cnpj or "").strip()
        if cnpj:
            lines.append(f"CNPJ: {cnpj}")
    return lines[:3]


def _footer_content_lines(
    brand: ExportBrandingConfig,
    profile: CompanyProfile,
) -> list[str]:
    lines: list[str] = []
    rt = str(brand.footer_line1 or profile.responsavel_linha() or "").strip()
    rt_contact = str(brand.footer_line2 or profile.rt_contato_linha() or "").strip()
    if rt:
        lines.append(rt)
    if rt_contact:
        lines.append(rt_contact)

    addr = profile.endereco_linha().strip()
    if addr:
        prefix = "Endereço: " if not addr.lower().startswith("endereço") else ""
        lines.append(f"{prefix}{addr}")

    contact_parts: list[str] = []
    if profile.email:
        contact_parts.append(f"E-mail: {profile.email}")
    if profile.telefone:
        contact_parts.append(f"Telefone: {profile.telefone}")
    if contact_parts:
        lines.append(" | ".join(contact_parts))
    return lines


def _apply_export_column_widths(ws, doc_type: str) -> None:
    from openpyxl.utils import get_column_letter

    presets: dict[str, dict[int, float]] = {
        "orc_analitico": {1: 9, 2: 11, 3: 12, 4: 42, 5: 7, 6: 10, 7: 14, 8: 14},
        "orc_sintetico": {1: 9, 2: 12, 3: 44, 4: 7, 5: 10, 6: 14, 7: 14},
        "mcq": {1: 9, 2: 12, 3: 50, 4: 7, 5: 12},
        "cronograma": {1: 10, 2: 44, 3: 16},
        "curva_abc": {1: 6, 2: 10, 3: 38, 4: 14, 5: 12, 6: 12, 7: 8},
        "curva_s": {1: 14, 2: 10, 3: 14, 4: 14, 5: 16, 6: 18, 7: 14},
        "histograma": {1: 14, 2: 10, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14},
    }
    widths = presets.get(doc_type.strip().lower())
    if not widths:
        return
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _apply_portrait_page_setup(ws) -> None:
    ws.page_setup.orientation = "portrait"
    try:
        from openpyxl.worksheet.properties import PageSetupProperties

        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    except Exception:
        pass
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def export_budget_document_xlsx(
    doc_type: str,
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata | None = None,
    *,
    branding: ExportBrandingConfig | None = None,
    schedule: Any | None = None,
    tech_spec: dict[str, Any] | None = None,
    logo_bytes: bytes | None = None,
    company_profile: CompanyProfile | None = None,
) -> bytes:
    """Exporta uma planilha única com o mesmo layout do PDF correspondente."""
    key = doc_type.strip().lower()
    if key not in EXPORT_DOC_TYPES:
        raise ValueError(f"Tipo de documento inválido: {doc_type}")

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl necessário") from exc

    meta = metadata or BudgetProjectMetadata()
    brand = branding or ExportBrandingConfig()
    profile = company_profile or get_company_profile()
    if logo_bytes is None and brand.show_logo:
        logo_bytes = load_logo_bytes(brand.logo_storage_key)

    table_data, extra_line, body_lines = build_export_table(
        key, roots, meta, schedule=schedule, tech_spec=tech_spec
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = key.upper()[:31]

    header_font = _xlsx_font(11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    stripe_fill = PatternFill("solid", fgColor="F2F2F2")
    summary_fill = PatternFill("solid", fgColor="E8EEF4")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    table_cols = _doc_table_col_count(key)
    header_cols = _doc_header_col_count(key)

    row_idx = _write_document_header(
        ws,
        title=DOC_TITLES[key],
        brand=brand,
        profile=profile,
        meta=meta,
        logo_bytes=logo_bytes,
        include_obra_meta=key in (
            "orc_sintetico",
            "orc_analitico",
            "cronograma",
            "curva_abc",
            "curva_s",
            "histograma",
        ),
        table_cols=header_cols,
        doc_type=key,
    )

    if extra_line:
        ws.cell(row=row_idx, column=1, value=extra_line)
        row_idx += 2

    if key == "orc_sintetico":
        row_idx = _write_orc_sintetico_formula_table(
            ws,
            roots,
            meta,
            start_row=row_idx,
            header_font=header_font,
            header_fill=header_fill,
            stripe_fill=stripe_fill,
            summary_fill=summary_fill,
            border=border,
        )
    elif key == "orc_analitico":
        row_idx = _write_orc_analitico_formula_table(
            ws,
            roots,
            meta,
            start_row=row_idx,
            header_font=header_font,
            header_fill=header_fill,
            stripe_fill=stripe_fill,
            summary_fill=summary_fill,
            border=border,
        )
    elif table_data:
        row_idx = _write_export_table(
            ws,
            table_data,
            start_row=row_idx,
            header_font=header_font,
            header_fill=header_fill,
            stripe_fill=stripe_fill,
            summary_fill=summary_fill,
            border=border,
        )
    elif body_lines:
        for line in body_lines:
            cell = ws.cell(row=row_idx, column=1, value=line)
            if row_idx == 8 and line and line == body_lines[0]:
                cell.font = _xlsx_font(12, bold=True)
            row_idx += 1

    footer_row = row_idx + 2
    _write_footer_block(ws, brand, profile, footer_row, table_cols=header_cols)
    if key in (
        "orc_sintetico",
        "orc_analitico",
        "mcq",
        "cronograma",
        "curva_abc",
        "curva_s",
        "histograma",
    ):
        _apply_export_column_widths(ws, key)
    else:
        _autosize_columns(ws)

    if key == "mcq":
        _apply_portrait_page_setup(ws)

    _apply_worksheet_arial_narrow(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_document_header(
    ws,
    *,
    title: str,
    brand: ExportBrandingConfig,
    profile: CompanyProfile,
    meta: BudgetProjectMetadata,
    logo_bytes: bytes | None,
    include_obra_meta: bool,
    table_cols: int,
    doc_type: str | None = None,
) -> int:
    """Cabeçalho institucional: logo no canto A1; linhas 1–4 mescladas A:H; meta obra/bases."""
    from openpyxl.styles import Alignment, Border, Font, Side

    text_end = table_cols
    header_blue = Side(style="thin", color=_XLSX_HEADER_BLUE)
    header_sep_border = Border(bottom=header_blue)

    if logo_bytes and brand.show_logo:
        try:
            from openpyxl.drawing.image import Image

            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp.write(logo_bytes)
                tmp_path = tmp.name
            img = Image(tmp_path)
            img.width = 88
            img.height = 72
            ws.add_image(img, "A1")
        except Exception:
            pass

    for row_num in range(1, 5):
        ws.row_dimensions[row_num].height = 16 if row_num == 1 else 14

    title_font = _xlsx_font(12, bold=True, color=_XLSX_HEADER_BLUE)
    subtitle_font = _xlsx_font(10, color="333333")
    meta_font = _xlsx_font(9, color="333333")

    _set_merged_cell(
        ws,
        1,
        1,
        text_end,
        title or brand.header_title,
        font=title_font,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=header_sep_border if not include_obra_meta else None,
    )

    subtitle_lines = _institutional_subtitle_lines(brand, profile)
    for offset, line in enumerate(subtitle_lines):
        row_num = 2 + offset
        if row_num > 4:
            break
        is_last_inst = row_num == 4 or offset == len(subtitle_lines) - 1
        _set_merged_cell(
            ws,
            row_num,
            1,
            text_end,
            line,
            font=subtitle_font,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=header_sep_border if is_last_inst and not include_obra_meta else None,
        )

    if include_obra_meta:
        obra_l1, obra_l2 = format_obra_meta_obra_lines(meta)
        price_l1, price_l2 = format_obra_meta_pricing_lines(meta)
        obra_end = _meta_obra_end_col(table_cols, doc_type)
        price_start = _meta_price_start_col(table_cols, doc_type)
        price_end = _meta_price_end_col(table_cols, doc_type)
        meta_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        meta_border = Border(bottom=Side(style="thin", color="CCCCCC"))

        if table_cols <= 3:
            _set_merged_cell(
                ws, 6, 1, table_cols, obra_l1, font=meta_font, alignment=meta_align
            )
            _set_merged_cell(
                ws, 7, 1, table_cols, obra_l2 or price_l1, font=meta_font, alignment=meta_align
            )
            _set_merged_cell(
                ws, 8, 1, table_cols, price_l2, font=meta_font, alignment=meta_align, border=meta_border
            )
            return 10

        _set_merged_cell(ws, 6, 1, obra_end, obra_l1, font=meta_font, alignment=meta_align)
        _set_merged_cell(
            ws, 6, price_start, price_end, price_l1, font=meta_font, alignment=meta_align
        )
        if obra_l2:
            _set_merged_cell(ws, 7, 1, obra_end, obra_l2, font=meta_font, alignment=meta_align)
        _set_merged_cell(
            ws,
            7,
            price_start,
            price_end,
            price_l2,
            font=meta_font,
            alignment=meta_align,
            border=meta_border,
        )
        ws.row_dimensions[6].height = 36
        ws.row_dimensions[7].height = 36
        return 9

    return 6


def _write_export_table(
    ws,
    data: ExportTableData,
    *,
    start_row: int,
    header_font,
    header_fill,
    stripe_fill,
    summary_fill,
    border,
) -> int:
    from openpyxl.styles import Alignment, Font

    row_idx = start_row
    for col_idx, label in enumerate(data.headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="right" if (col_idx - 1) in data.right_cols else "center",
            vertical="center",
        )
    row_idx += 1

    body_end = len(data.rows) - data.summary_rows
    for body_idx, row_values in enumerate(data.rows):
        is_summary = body_idx >= body_end
        is_bold = body_idx in data.bold_rows or is_summary
        if not is_summary and body_idx % 2 == 1:
            row_fill = stripe_fill
        elif is_summary:
            row_fill = summary_fill
        else:
            row_fill = None

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value != "" else None)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if (col_idx - 1) in data.right_cols else "left",
                vertical="top",
                wrap_text=col_idx == (4 if len(data.headers) >= 8 else 3),
            )
            if row_fill:
                cell.fill = row_fill
            cell.font = _xlsx_font(10, bold=is_bold)
            _apply_numeric_cell_format(cell, col_idx, data.right_cols)
        row_idx += 1

    return row_idx + 1


_ORC_SINT_COL_ITEM = 1
_ORC_SINT_COL_CODE = 2
_ORC_SINT_COL_DESC = 3
_ORC_SINT_COL_UNIT = 4
_ORC_SINT_COL_QTY = 5
_ORC_SINT_COL_UNIT_PRICE = 6
_ORC_SINT_COL_TOTAL = 7
_ORC_SINT_COL_UNIT_COST = 8  # coluna auxiliar oculta (custo sem BDI)
_ORC_BDI_REF_COL = 26  # Z1
_ORC_BDI_REF_ROW = 1


def _orc_is_group(item: BudgetItem) -> bool:
    return item.row_type in (ROW_TYPE_ETAPA, ROW_TYPE_SUB_ETAPA) or (
        item.item_type.value == "group" and item.level == 0
    )


def _orc_unit_price(item: BudgetItem, mode: str) -> float | None:
    value = item.unit_price_semd if mode == "semd" else item.unit_price
    return float(value) if value else None


def _orc_unit_cost(item: BudgetItem, mode: str, bdi_rate: float) -> float | None:
    cost = item.unit_cost_semd if mode == "semd" else item.unit_cost
    if cost and float(cost) > 0:
        return float(cost)
    unit_price = _orc_unit_price(item, mode)
    if unit_price and bdi_rate >= 0:
        return round(float(unit_price) / (1 + float(bdi_rate)), 6)
    return None


def _cpu_unit_price(cpu_item: dict[str, Any], mode: str) -> float | None:
    if mode == "semd":
        value = cpu_item.get("unit_price_sem")
        if value is None:
            value = cpu_item.get("unit_price")
    else:
        value = cpu_item.get("unit_price")
    return float(value) if value is not None else None


def _cpu_coefficient(cpu_item: dict[str, Any]) -> float | None:
    value = cpu_item.get("coefficient")
    return float(value) if value is not None else None


def _apply_orc_table_row_style(
    ws,
    row: int,
    *,
    col_count: int,
    desc_col: int,
    right_cols: tuple[int, ...],
    border,
    row_fill,
    bold: bool,
) -> None:
    from openpyxl.styles import Alignment

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.border = border
        cell.alignment = Alignment(
            horizontal="right" if (col_idx - 1) in right_cols else "left",
            vertical="top",
            wrap_text=col_idx == desc_col,
        )
        if row_fill:
            cell.fill = row_fill
        if bold:
            cell.font = _xlsx_font(10, bold=True)
        else:
            cell.font = _xlsx_font(10)


def _write_orc_bdi_footer(
    ws,
    *,
    desc_col: int,
    total_col: int,
    col_count: int,
    right_cols: tuple[int, ...],
    row_idx: int,
    service_total_refs: list[str],
    direct_cost_terms: list[str],
    bdi_ref: str,
    bdi_label: str,
    summary_fill,
    border,
) -> int:
    from openpyxl.utils import get_column_letter

    def _write_summary_row(label: str, formula: str) -> int:
        nonlocal row_idx
        r = row_idx
        ws.cell(row=r, column=desc_col, value=label)
        total_cell = ws.cell(row=r, column=total_col, value=formula)
        total_cell.number_format = _XLSX_NUM_FMT
        _apply_orc_table_row_style(
            ws,
            r,
            col_count=col_count,
            desc_col=desc_col,
            right_cols=right_cols,
            border=border,
            row_fill=summary_fill,
            bold=True,
        )
        row_idx += 1
        return r

    if service_total_refs:
        com_formula = f"=ROUND(SUM({','.join(service_total_refs)}),2)"
        if direct_cost_terms:
            sem_formula = f"=ROUND({'+' .join(direct_cost_terms)},2)"
        else:
            sem_formula = f"=ROUND(SUM({','.join(service_total_refs)})/(1+{bdi_ref}),2)"
    else:
        com_formula = "=0"
        sem_formula = "=0"

    sem_row = _write_summary_row("TOTAL SEM BDI", sem_formula)
    com_row = row_idx + 1
    _write_summary_row(
        bdi_label,
        f"={get_column_letter(total_col)}{com_row}-{get_column_letter(total_col)}{sem_row}",
    )
    _write_summary_row("TOTAL COM BDI", com_formula)
    return row_idx + 1


def _price_headers_for_mode(mode: str) -> tuple[str, str]:
    if mode == "semd":
        return "Unit. Sem D", "Total Sem D"
    return "Unit. Com D", "Total Com D"


_ORC_ANAL_COL_ITEM = 1
_ORC_ANAL_COL_TIPO = 2
_ORC_ANAL_COL_CODE = 3
_ORC_ANAL_COL_DESC = 4
_ORC_ANAL_COL_UNIT = 5
_ORC_ANAL_COL_QTY = 6
_ORC_ANAL_COL_UNIT_PRICE = 7
_ORC_ANAL_COL_TOTAL = 8
_ORC_ANAL_COL_UNIT_COST = 9  # coluna auxiliar oculta (custo sem BDI)


def _write_orc_sintetico_formula_table(
    ws,
    roots: list[BudgetItem],
    meta: BudgetProjectMetadata,
    *,
    start_row: int,
    header_font,
    header_fill,
    stripe_fill,
    summary_fill,
    border,
) -> int:
    """Orçamento sintético com fórmulas: total serviço = Qtd×Unit; etapas somam filhos; rodapé BDI."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    adopted_mode = budget_desoneracao_mode(roots)
    unit_hdr, total_hdr = _price_headers_for_mode(adopted_mode)
    bdi_rate = (
        meta.bdi.rate_sem_desoneracao if adopted_mode == "semd" else meta.bdi.rate_com_desoneracao
    )
    bdi_label = f"BDI ({bdi_rate * 100:.2f}%)".replace(".", ",")

    bdi_ref = f"${get_column_letter(_ORC_BDI_REF_COL)}${_ORC_BDI_REF_ROW}"
    ws.cell(row=_ORC_BDI_REF_ROW, column=_ORC_BDI_REF_COL, value=bdi_rate)
    ws.column_dimensions[get_column_letter(_ORC_SINT_COL_UNIT_COST)].hidden = True

    headers = ["Item", "Código", "Descrição", "Un", "Qtd", unit_hdr, total_hdr]
    right_cols = (4, 5, 6)
    row_idx = start_row

    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="right" if (col_idx - 1) in right_cols else "center",
            vertical="center",
        )
    row_idx += 1

    service_total_refs: list[str] = []
    direct_cost_terms: list[str] = []
    body_row = 0

    def write_item(item: BudgetItem, depth: int) -> str | None:
        nonlocal row_idx, body_row

        if item.metadata.get("is_memory_row") or item.row_type == "MEMORIA":
            return None

        is_group = _orc_is_group(item)
        indent = "  " * depth
        current_row = row_idx
        body_row += 1
        row_fill = stripe_fill if body_row % 2 == 0 else None

        ws.cell(row=current_row, column=_ORC_SINT_COL_ITEM, value=item.code or "")
        ws.cell(row=current_row, column=_ORC_SINT_COL_CODE, value="" if is_group else (item.source_code or ""))
        ws.cell(row=current_row, column=_ORC_SINT_COL_DESC, value=f"{indent}{item.name or ''}")

        if is_group:
            _apply_orc_table_row_style(
                ws,
                current_row,
                col_count=7,
                desc_col=_ORC_SINT_COL_DESC,
                right_cols=right_cols,
                border=border,
                row_fill=row_fill,
                bold=True,
            )
            row_idx += 1
            child_refs: list[str] = []
            for child in item.children:
                ref = write_item(child, depth + 1)
                if ref:
                    child_refs.append(ref)
            if child_refs:
                total_cell = ws.cell(
                    row=current_row,
                    column=_ORC_SINT_COL_TOTAL,
                    value=f"=ROUND(SUM({','.join(child_refs)}),2)",
                )
                total_cell.number_format = _XLSX_NUM_FMT
            return f"{get_column_letter(_ORC_SINT_COL_TOTAL)}{current_row}"

        if not _is_service_row(item):
            _apply_orc_table_row_style(
                ws,
                current_row,
                col_count=7,
                desc_col=_ORC_SINT_COL_DESC,
                right_cols=right_cols,
                border=border,
                row_fill=row_fill,
                bold=False,
            )
            row_idx += 1
            return None

        unit_price = _orc_unit_price(item, adopted_mode)
        unit_cost = _orc_unit_cost(item, adopted_mode, bdi_rate)
        qty = float(item.quantity or 0)

        ws.cell(row=current_row, column=_ORC_SINT_COL_UNIT, value=item.unit or "")
        qty_cell = ws.cell(row=current_row, column=_ORC_SINT_COL_QTY, value=qty if qty else None)
        qty_cell.number_format = _XLSX_NUM_FMT
        if unit_price is not None:
            price_cell = ws.cell(row=current_row, column=_ORC_SINT_COL_UNIT_PRICE, value=unit_price)
            price_cell.number_format = _XLSX_NUM_FMT
        if unit_cost is not None:
            cost_cell = ws.cell(row=current_row, column=_ORC_SINT_COL_UNIT_COST, value=unit_cost)
            cost_cell.number_format = _XLSX_NUM_FMT

        total_ref = f"{get_column_letter(_ORC_SINT_COL_TOTAL)}{current_row}"
        qty_ref = f"{get_column_letter(_ORC_SINT_COL_QTY)}{current_row}"
        price_ref = f"{get_column_letter(_ORC_SINT_COL_UNIT_PRICE)}{current_row}"
        total_cell = ws.cell(
            row=current_row,
            column=_ORC_SINT_COL_TOTAL,
            value=f"=ROUND({qty_ref}*{price_ref},2)",
        )
        total_cell.number_format = _XLSX_NUM_FMT

        service_total_refs.append(total_ref)
        if unit_cost is not None:
            cost_ref = f"{get_column_letter(_ORC_SINT_COL_UNIT_COST)}{current_row}"
            direct_cost_terms.append(f"{qty_ref}*{cost_ref}")
        else:
            direct_cost_terms.append(f"{total_ref}/(1+{bdi_ref})")

        _apply_orc_table_row_style(
            ws,
            current_row,
            col_count=7,
            desc_col=_ORC_SINT_COL_DESC,
            right_cols=right_cols,
            border=border,
            row_fill=row_fill,
            bold=False,
        )
        row_idx += 1
        return total_ref

    for root in roots:
        write_item(root, 0)

    return _write_orc_bdi_footer(
        ws,
        desc_col=_ORC_SINT_COL_DESC,
        total_col=_ORC_SINT_COL_TOTAL,
        col_count=7,
        right_cols=right_cols,
        row_idx=row_idx,
        service_total_refs=service_total_refs,
        direct_cost_terms=direct_cost_terms,
        bdi_ref=bdi_ref,
        bdi_label=bdi_label,
        summary_fill=summary_fill,
        border=border,
    )


def _write_orc_analitico_formula_table(
    ws,
    roots: list[BudgetItem],
    meta: BudgetProjectMetadata,
    *,
    start_row: int,
    header_font,
    header_fill,
    stripe_fill,
    summary_fill,
    border,
) -> int:
    """Orçamento analítico com fórmulas: serviço Qtd×Unit; CPU Coef×Unit; etapas somam filhos; rodapé BDI."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    adopted_mode = budget_desoneracao_mode(roots)
    unit_hdr, total_hdr = _price_headers_for_mode(adopted_mode)
    bdi_rate = (
        meta.bdi.rate_sem_desoneracao if adopted_mode == "semd" else meta.bdi.rate_com_desoneracao
    )
    bdi_label = f"BDI ({bdi_rate * 100:.2f}%)".replace(".", ",")

    bdi_ref = f"${get_column_letter(_ORC_BDI_REF_COL)}${_ORC_BDI_REF_ROW}"
    ws.cell(row=_ORC_BDI_REF_ROW, column=_ORC_BDI_REF_COL, value=bdi_rate)
    ws.column_dimensions[get_column_letter(_ORC_ANAL_COL_UNIT_COST)].hidden = True

    headers = ["Item", "Tipo", "Código", "Descrição", "Un", "Qtd", unit_hdr, total_hdr]
    right_cols = (5, 6, 7)
    row_idx = start_row
    comp_cache: dict[tuple[str, str, str], list[dict[str, Any]]] = {}

    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="right" if (col_idx - 1) in right_cols else "center",
            vertical="center",
        )
    row_idx += 1

    service_total_refs: list[str] = []
    direct_cost_terms: list[str] = []
    body_row = 0

    def write_cpu_rows(cpu_items: list[dict[str, Any]], depth: int) -> None:
        nonlocal row_idx, body_row
        indent = "  " * depth
        for cpu_item in cpu_items:
            current_row = row_idx
            body_row += 1
            row_fill = stripe_fill if body_row % 2 == 0 else None

            ws.cell(
                row=current_row,
                column=_ORC_ANAL_COL_TIPO,
                value=_cpu_tipo_label(str(cpu_item.get("item_type") or "")),
            )
            ws.cell(row=current_row, column=_ORC_ANAL_COL_CODE, value=str(cpu_item.get("code") or ""))
            ws.cell(
                row=current_row,
                column=_ORC_ANAL_COL_DESC,
                value=f"{indent}{cpu_item.get('description') or ''}",
            )
            ws.cell(row=current_row, column=_ORC_ANAL_COL_UNIT, value=str(cpu_item.get("unit") or ""))

            coef = _cpu_coefficient(cpu_item)
            unit_p = _cpu_unit_price(cpu_item, adopted_mode)
            if coef is not None:
                coef_cell = ws.cell(row=current_row, column=_ORC_ANAL_COL_QTY, value=coef)
                coef_cell.number_format = _XLSX_NUM_FMT
            if unit_p is not None:
                price_cell = ws.cell(row=current_row, column=_ORC_ANAL_COL_UNIT_PRICE, value=unit_p)
                price_cell.number_format = _XLSX_NUM_FMT

            qty_ref = f"{get_column_letter(_ORC_ANAL_COL_QTY)}{current_row}"
            price_ref = f"{get_column_letter(_ORC_ANAL_COL_UNIT_PRICE)}{current_row}"
            total_cell = ws.cell(
                row=current_row,
                column=_ORC_ANAL_COL_TOTAL,
                value=f"=ROUND({qty_ref}*{price_ref},2)",
            )
            total_cell.number_format = _XLSX_NUM_FMT

            _apply_orc_table_row_style(
                ws,
                current_row,
                col_count=8,
                desc_col=_ORC_ANAL_COL_DESC,
                right_cols=right_cols,
                border=border,
                row_fill=row_fill,
                bold=False,
            )
            row_idx += 1

    def write_item(item: BudgetItem, depth: int) -> str | None:
        nonlocal row_idx, body_row

        if item.metadata.get("is_memory_row") or item.row_type == "MEMORIA":
            return None

        is_group = _orc_is_group(item)
        indent = "  " * depth
        current_row = row_idx
        body_row += 1
        row_fill = stripe_fill if body_row % 2 == 0 else None

        ws.cell(row=current_row, column=_ORC_ANAL_COL_ITEM, value=item.code or "")
        ws.cell(row=current_row, column=_ORC_ANAL_COL_TIPO, value=_budget_row_tipo(item, is_group=is_group))
        ws.cell(row=current_row, column=_ORC_ANAL_COL_CODE, value="" if is_group else (item.source_code or ""))
        ws.cell(row=current_row, column=_ORC_ANAL_COL_DESC, value=f"{indent}{item.name or ''}")

        if is_group:
            _apply_orc_table_row_style(
                ws,
                current_row,
                col_count=8,
                desc_col=_ORC_ANAL_COL_DESC,
                right_cols=right_cols,
                border=border,
                row_fill=row_fill,
                bold=True,
            )
            row_idx += 1
            child_refs: list[str] = []
            for child in item.children:
                ref = write_item(child, depth + 1)
                if ref:
                    child_refs.append(ref)
            if child_refs:
                total_cell = ws.cell(
                    row=current_row,
                    column=_ORC_ANAL_COL_TOTAL,
                    value=f"=ROUND(SUM({','.join(child_refs)}),2)",
                )
                total_cell.number_format = _XLSX_NUM_FMT
            return f"{get_column_letter(_ORC_ANAL_COL_TOTAL)}{current_row}"

        if not _is_service_row(item):
            _apply_orc_table_row_style(
                ws,
                current_row,
                col_count=8,
                desc_col=_ORC_ANAL_COL_DESC,
                right_cols=right_cols,
                border=border,
                row_fill=row_fill,
                bold=False,
            )
            row_idx += 1
            return None

        unit_price = _orc_unit_price(item, adopted_mode)
        unit_cost = _orc_unit_cost(item, adopted_mode, bdi_rate)
        qty = float(item.quantity or 0)
        source_code = (item.source_code or "").strip()

        ws.cell(row=current_row, column=_ORC_ANAL_COL_UNIT, value=item.unit or "")
        qty_cell = ws.cell(row=current_row, column=_ORC_ANAL_COL_QTY, value=qty if qty else None)
        qty_cell.number_format = _XLSX_NUM_FMT
        if unit_price is not None:
            price_cell = ws.cell(row=current_row, column=_ORC_ANAL_COL_UNIT_PRICE, value=unit_price)
            price_cell.number_format = _XLSX_NUM_FMT
        if unit_cost is not None:
            cost_cell = ws.cell(row=current_row, column=_ORC_ANAL_COL_UNIT_COST, value=unit_cost)
            cost_cell.number_format = _XLSX_NUM_FMT

        total_ref = f"{get_column_letter(_ORC_ANAL_COL_TOTAL)}{current_row}"
        qty_ref = f"{get_column_letter(_ORC_ANAL_COL_QTY)}{current_row}"
        price_ref = f"{get_column_letter(_ORC_ANAL_COL_UNIT_PRICE)}{current_row}"
        total_cell = ws.cell(
            row=current_row,
            column=_ORC_ANAL_COL_TOTAL,
            value=f"=ROUND({qty_ref}*{price_ref},2)",
        )
        total_cell.number_format = _XLSX_NUM_FMT

        service_total_refs.append(total_ref)
        if unit_cost is not None:
            cost_ref = f"{get_column_letter(_ORC_ANAL_COL_UNIT_COST)}{current_row}"
            direct_cost_terms.append(f"{qty_ref}*{cost_ref}")
        else:
            direct_cost_terms.append(f"{total_ref}/(1+{bdi_ref})")

        _apply_orc_table_row_style(
            ws,
            current_row,
            col_count=8,
            desc_col=_ORC_ANAL_COL_DESC,
            right_cols=right_cols,
            border=border,
            row_fill=row_fill,
            bold=False,
        )
        row_idx += 1

        if source_code:
            lookup = _resolve_open_composition_lookup(item, meta)
            if lookup:
                cpu_items = _fetch_open_composition_items(source_code, lookup, comp_cache)
                if cpu_items:
                    write_cpu_rows(cpu_items, depth + 1)

        return total_ref

    for root in roots:
        write_item(root, 0)

    return _write_orc_bdi_footer(
        ws,
        desc_col=_ORC_ANAL_COL_DESC,
        total_col=_ORC_ANAL_COL_TOTAL,
        col_count=8,
        right_cols=right_cols,
        row_idx=row_idx,
        service_total_refs=service_total_refs,
        direct_cost_terms=direct_cost_terms,
        bdi_ref=bdi_ref,
        bdi_label=bdi_label,
        summary_fill=summary_fill,
        border=border,
    )


def _write_footer_block(
    ws,
    brand: ExportBrandingConfig,
    profile: CompanyProfile,
    row: int,
    *,
    table_cols: int,
) -> None:
    from openpyxl.styles import Alignment, Font

    lines = _footer_content_lines(brand, profile)
    if not lines:
        return

    footer_font = _xlsx_font(9, color="333333")
    footer_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for line in lines:
        _set_merged_cell(
            ws,
            row,
            1,
            table_cols,
            line,
            font=footer_font,
            alignment=footer_align,
        )
        ws.row_dimensions[row].height = 18
        row += 1


def export_budget_workbook_xlsx(
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata | None = None,
    *,
    branding: ExportBrandingConfig | None = None,
    schedule: Any | None = None,
    tech_spec: dict[str, Any] | None = None,
    logo_bytes: bytes | None = None,
    company_profile: CompanyProfile | None = None,
) -> bytes:
    """Exporta workbook completo (5 abas) — legado; preferir export_budget_document_xlsx."""
    _ = company_profile  # reservado — workbook legado usa branding global
    from pricing.budget.ppd_exporter import (
        CRONOGRAMA_ETAPA_MAX_ROW,
        CRONOGRAMA_ETAPA_START,
        CRONOGRAMA_ETAPA_STEP,
        _schedule_total_days,
        _write_ppd_tree,
    )
    from pricing.budget.ppd_layout import (
        COL_DESCRIPTION,
        COL_ITEM,
        COL_QUANTITY,
        COL_ROW_TYPE,
        COL_SOURCE_CODE,
        COL_TOTAL_COMD,
        COL_TOTAL_SEMD,
        COL_UNIT,
        COL_UNIT_COST_COMD,
        COL_UNIT_COST_SEMD,
        COL_UNIT_PRICE_BDI_COMD,
        COL_UNIT_PRICE_BDI_SEMD,
        ROW_TYPE_ETAPA,
        ROW_TYPE_SERVICO,
        ROW_TYPE_SUB_ETAPA,
    )
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl necessário") from exc

    meta = metadata or BudgetProjectMetadata()
    brand = branding or ExportBrandingConfig()
    if logo_bytes is None and brand.show_logo:
        logo_bytes = load_logo_bytes(brand.logo_storage_key)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheets: list[tuple[str, str]] = [
        ("ORC_SINTETICO", "PLANILHA ORÇAMENTÁRIA - ORÇAMENTO SINTÉTICO"),
        ("ORC_ANALITICO", "PLANILHA ORÇAMENTÁRIA - ORÇAMENTO ANALÍTICO"),
        ("MCQ", "MEMÓRIA DE CÁLCULO QUANTITATIVA"),
        ("CRONOGRAMA", "CRONOGRAMA FÍSICO-FINANCEIRO"),
        ("ESP_TECNICA", "ESPECIFICAÇÃO TÉCNICA"),
    ]

    header_font = _xlsx_font(11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, doc_title in sheets:
        ws = wb.create_sheet(sheet_name)
        _write_branded_header(ws, brand, meta, doc_title, logo_bytes)

        if sheet_name in ("ORC_SINTETICO", "ORC_ANALITICO"):
            _write_column_headers(ws, meta, header_font, header_fill, border)
            row_idx = 21
            for root in roots:
                row_idx = _write_ppd_tree(ws, root, row_idx, meta)
            total = sum(r.total_price for r in roots)
            total_semd = sum(r.total_price_semd for r in roots)
            total_eff = sum(r.effective_total() for r in roots)
            ws.cell(row=row_idx + 1, column=COL_TOTAL_COMD + 1, value=round(total, 2))
            ws.cell(row=row_idx + 1, column=COL_TOTAL_SEMD + 1, value=round(total_semd, 2))
            ws.cell(row=row_idx + 2, column=COL_DESCRIPTION + 1, value="TOTAL EFETIVO (MENOR CUSTO)")
            ws.cell(row=row_idx + 2, column=COL_TOTAL_COMD + 1, value=round(total_eff, 2))

        elif sheet_name == "MCQ":
            _write_mcq_headers(ws, meta, header_font, header_fill, border)
            row_idx = 21
            for root in roots:
                row_idx = _write_mcq_values(ws, root, row_idx)
            _write_legacy_footer_block(ws, brand, row_idx + 3)

        elif sheet_name == "CRONOGRAMA":
            _write_cronograma_sheet(ws, roots, meta, schedule, brand)

        elif sheet_name == "ESP_TECNICA":
            _write_esp_tecnica_sheet(ws, meta, tech_spec, brand)

        _autosize_columns(ws)
        _apply_worksheet_arial_narrow(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_branded_header(
    ws,
    brand: ExportBrandingConfig,
    meta: BudgetProjectMetadata,
    doc_title: str,
    logo_bytes: bytes | None,
) -> None:
    from openpyxl.styles import Alignment

    start_col = 1
    if logo_bytes and brand.show_logo:
        try:
            from openpyxl.drawing.image import Image

            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp.write(logo_bytes)
                tmp_path = tmp.name
            img = Image(tmp_path)
            img.width = 100
            img.height = 50
            ws.add_image(img, "A1")
            start_col = 2
        except Exception:
            pass

    title_cell = ws.cell(row=1, column=start_col, value=doc_title or brand.header_title)
    title_cell.font = _xlsx_font(14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    lines = [
        brand.header_line1,
        brand.header_line2,
        brand.header_line3,
    ]
    row = 2
    for line in lines:
        if line:
            c = ws.cell(row=row, column=start_col, value=line)
            c.font = _xlsx_font(10)
            c.alignment = Alignment(horizontal="center")
            row += 1

    ws.cell(row=row, column=1, value=f"Base: {meta.base_preco}")
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=(
            f"BDI Com D: {meta.bdi.rate_com_desoneracao:.2%} · "
            f"Sem D: {meta.bdi.rate_sem_desoneracao:.2%}"
        ),
    )


def _write_column_headers(ws, meta, header_font, header_fill, border) -> None:
    headers = {
        COL_ROW_TYPE + 1: "TIPO",
        COL_ITEM + 1: "ITEM",
        COL_SOURCE_CODE + 1: "CÓDIGO",
        COL_DESCRIPTION + 1: "DESCRIÇÃO",
        COL_UNIT + 1: "UNID",
        COL_QUANTITY + 1: "QUANT",
        COL_UNIT_COST_COMD + 1: "CUSTO UNIT. (R$) COM D",
        COL_UNIT_PRICE_BDI_COMD + 1: f"PREÇO UNIT. COM BDI DE {meta.bdi.rate_com_desoneracao:.2%}",
        COL_TOTAL_COMD + 1: "TOTAL COM BDI (R$)",
        COL_UNIT_COST_SEMD + 1: "CUSTO UNIT. (R$) SEM D",
        COL_UNIT_PRICE_BDI_SEMD + 1: f"PREÇO UNIT. COM BDI DE {meta.bdi.rate_sem_desoneracao:.2%}",
        COL_TOTAL_SEMD + 1: "TOTAL COM BDI (R$)",
    }
    for col, label in headers.items():
        cell = ws.cell(row=19, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border


def _write_mcq_headers(ws, meta, header_font, header_fill, border) -> None:
    headers = {
        7: "TIPO",
        8: "ITEM",
        9: "CÓDIGO",
        10: "DESCRIÇÃO",
        11: "UNID",
        12: "QUANT",
    }
    for col, label in headers.items():
        cell = ws.cell(row=19, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border


def _write_mcq_values(ws, item: BudgetItem, row_idx: int) -> int:
    if item.metadata.get("is_memory_row") or item.row_type == "MEMORIA":
        ws.cell(row=row_idx, column=10, value=item.calculation_note or item.name)
        return row_idx + 1

    is_etapa = item.row_type == ROW_TYPE_ETAPA or (item.item_type.value == "group" and item.level == 0)
    is_sub = item.row_type == ROW_TYPE_SUB_ETAPA

    if is_etapa or is_sub:
        ws.cell(row=row_idx, column=7, value=ROW_TYPE_ETAPA if is_etapa else ROW_TYPE_SUB_ETAPA)
        ws.cell(row=row_idx, column=8, value=item.code)
        ws.cell(row=row_idx, column=10, value=item.name)
        row_idx += 1
        for child in item.children:
            row_idx = _write_mcq_values(ws, child, row_idx)
        return row_idx

    ws.cell(row=row_idx, column=7, value=ROW_TYPE_SERVICO)
    ws.cell(row=row_idx, column=8, value=item.code)
    ws.cell(row=row_idx, column=9, value=item.source_code)
    ws.cell(row=row_idx, column=10, value=item.name or item.pricing_query)
    ws.cell(row=row_idx, column=11, value=item.unit)
    ws.cell(row=row_idx, column=12, value=item.quantity)
    row_idx += 1

    for child in item.children:
        if child.metadata.get("is_memory_row") or child.row_type == "MEMORIA":
            ws.cell(row=row_idx, column=10, value=child.calculation_note or child.name)
            row_idx += 1
        else:
            row_idx = _write_mcq_values(ws, child, row_idx)

    if item.calculation_note and not any(c.metadata.get("is_memory_row") for c in item.children):
        ws.cell(row=row_idx, column=10, value=item.calculation_note)
        row_idx += 1

    return row_idx


def _write_cronograma_sheet(ws, roots, meta, schedule, brand) -> None:
    ws.cell(row=1, column=1, value=brand.header_line2 or meta.projeto)
    ws.cell(row=2, column=1, value=brand.header_line3 or meta.local)
    prazo = _schedule_total_days(schedule)
    if prazo:
        ws.cell(row=3, column=1, value=f"Prazo de execução: {prazo} dias")

    ws.cell(row=8, column=1, value="ETAPA")
    ws.cell(row=8, column=2, value="DESCRIÇÃO")
    ws.cell(row=8, column=3, value="VALOR (R$)")

    row = CRONOGRAMA_ETAPA_START
    for root in roots:
        if root.row_type != ROW_TYPE_ETAPA and root.level != 0:
            continue
        if row > CRONOGRAMA_ETAPA_MAX_ROW:
            break
        code = root.code if "." in str(root.code) else f"{root.code}.0"
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=root.name)
        total = root.total_price or root.effective_total()
        if total:
            ws.cell(row=row, column=3, value=round(total, 2))
        row += CRONOGRAMA_ETAPA_STEP

    _write_legacy_footer_block(ws, brand, row + 2)


def _write_esp_tecnica_sheet(ws, meta, tech_spec, brand) -> None:
    from pricing.spec.tech_spec_models import TechSpecDocument

    doc = TechSpecDocument.from_dict(tech_spec) if tech_spec else None
    ws.cell(row=1, column=1, value=brand.header_line2 or meta.projeto)
    ws.cell(row=2, column=1, value=doc.title if doc else "ESPECIFICAÇÃO TÉCNICA")

    body = ""
    if doc:
        body = (doc.markdown or "").strip()
        if not body and doc.html_content:
            import re

            text = re.sub(r"<br\s*/?>", "\n", doc.html_content, flags=re.I)
            text = re.sub(r"</p>", "\n", text, flags=re.I)
            text = re.sub(r"<[^>]+>", "", text)
            body = text.strip()

    row = 4
    if not body:
        ws.cell(row=row, column=1, value="(Conteúdo não gerado — use a aba Especificação no orçamento)")
    else:
        for line in body.splitlines():
            if line.strip():
                ws.cell(row=row, column=1, value=line.rstrip())
                row += 1

    _write_legacy_footer_block(ws, brand, row + 2)


def _write_legacy_footer_block(ws, brand: ExportBrandingConfig, row: int) -> None:
    if brand.footer_line1:
        ws.cell(row=row, column=1, value=brand.footer_line1)
        row += 1
    if brand.footer_line2:
        ws.cell(row=row, column=1, value=brand.footer_line2)


def _autosize_columns(ws) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, min(ws.max_column or 1, 20) + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, max_row=min(ws.max_row or 1, 200)):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, min(48, len(str(val)) + 2))
        ws.column_dimensions[letter].width = max_len
