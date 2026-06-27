from __future__ import annotations

import io
from typing import Any

from pricing.budget.ppd_layout import (
    COL_DESCRIPTION,
    COL_ITEM,
    COL_QUANTITY,
    COL_ROW_TYPE,
    COL_SOURCE_CODE,
    COL_TOTAL_COMD,
    COL_TOTAL_SEMD,
    COL_UNIT,
    COL_UNIT_COST_COMD,
    COL_UNIT_COST_SEMD,
    COL_UNIT_PRICE_BDI_COMD,
    COL_UNIT_PRICE_BDI_SEMD,
    ROW_TYPE_ETAPA,
    ROW_TYPE_SERVICO,
    ROW_TYPE_SUB_ETAPA,
)
from pricing.models.budget_item import BudgetItem
from pricing.models.budget_metadata import BudgetProjectMetadata

MCQ_DATA_START_ROW = 21
_MCQ_SYNC_COLS = range(7, 14)  # G-M — colunas de entrada MCQ (sem preços)

_MCQ_DESC_FORMULA = (
    '=IF(G{row}="ETAPA", "Digitar aqui o nome da ETAPA", '
    'IF(G{row}="SUB-ETAPA", "Digitar aqui o nome da  SUB-ETAPA", '
    'IF(I{row}="", "", IFERROR(VLOOKUP(VALUE(I{row}), _BaseAbril2026, 2, 0), '
    'IFERROR(VLOOKUP(TEXT(I{row}, "@"), _BaseAbril2026, 2, 0), '
    '"Código Não Encontrado na Tabela")))))'
)
_MCQ_UNIT_FORMULA = (
    '=IF(I{row}="", "", IFERROR(VLOOKUP(VALUE(I{row}), _BaseAbril2026, 3, 0), '
    'IFERROR(VLOOKUP(TEXT(I{row}, "@"), _BaseAbril2026, 3, 0), "")))'
)


def _uses_base_vlookup(source_code: str) -> bool:
    """Códigos numéricos SINAPI/SICRO usam VLOOKUP na Base; códigos SEMINF usam texto literal."""
    code = (source_code or "").strip()
    if not code:
        return False
    if ".SEMINF" in code.upper() or code.count(".") >= 2:
        return False
    return code.isdigit() or code.replace(".", "", 1).isdigit()


def _safe_set_cell(ws, row: int, column: int, value: Any) -> None:
    from openpyxl.cell.cell import MergedCell

    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return
    try:
        cell.value = value
    except AttributeError:
        return


def export_ppd_xlsx(
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata | None = None,
) -> bytes:
    """Exporta no layout PPD municipal (PLANILHA + MCQ)."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl necessário") from exc

    meta = metadata or BudgetProjectMetadata()
    wb = openpyxl.Workbook()

    ws_plan = wb.active
    ws_plan.title = "PLANILHA"
    ws_mcq = wb.create_sheet("MCQ")

    for ws in (ws_plan, ws_mcq):
        _write_ppd_header(ws, meta)
        _write_ppd_column_headers(ws, meta)
        row_idx = 21
        for root in roots:
            row_idx = _write_ppd_tree(ws, root, row_idx, meta)

        total = sum(r.total_price for r in roots)
        total_semd = sum(r.total_price_semd for r in roots)
        total_effective = sum(r.effective_total() for r in roots)
        ws.cell(row=row_idx + 1, column=COL_TOTAL_COMD + 1, value=round(total, 2))
        ws.cell(row=row_idx + 1, column=COL_TOTAL_SEMD + 1, value=round(total_semd, 2))
        ws.cell(row=row_idx + 2, column=COL_DESCRIPTION + 1, value="TOTAL EFETIVO (MENOR CUSTO)")
        ws.cell(row=row_idx + 2, column=COL_TOTAL_COMD + 1, value=round(total_effective, 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_ppd_header(ws, meta: BudgetProjectMetadata) -> None:
    _safe_set_cell(ws, 6, 10, meta.orgao or "SEMINF")
    _safe_set_cell(ws, 11, 8, "PROJETO:")
    _safe_set_cell(ws, 11, 9, meta.projeto)
    _safe_set_cell(ws, 12, 8, "OBJETO:")
    _safe_set_cell(ws, 12, 9, meta.objeto)
    _safe_set_cell(ws, 13, 8, "LOCAL:")
    _safe_set_cell(ws, 13, 9, meta.local)
    _safe_set_cell(ws, 14, 8, "ORÇAMENTO:")
    _safe_set_cell(ws, 14, 9, meta.orcamento)
    _safe_set_cell(ws, 14, 15, meta.obra_type)
    _safe_set_cell(ws, 16, 8, f"BASE DE PREÇO UTILIZADA: {meta.base_preco}")
    _safe_set_cell(ws, 18, 10, "MEMÓRIA DE CÁLCULO")
    _safe_set_cell(ws, 18, 15, 0)
    _safe_set_cell(ws, 18, 14, meta.bdi.rate_com_desoneracao)
    _safe_set_cell(ws, 18, 18, meta.bdi.rate_sem_desoneracao)


def _write_ppd_column_headers(ws, meta: BudgetProjectMetadata) -> None:
    headers = {
        COL_ROW_TYPE + 1: "TIPO",
        COL_ITEM + 1: "ITEM",
        COL_SOURCE_CODE + 1: "CÓDIGO",
        COL_DESCRIPTION + 1: "DESCRIÇÃO",
        COL_UNIT + 1: "UNID",
        COL_QUANTITY + 1: "QUANT",
        COL_UNIT_COST_COMD + 1: "CUSTO UNIT. (R$) COM D",
        COL_UNIT_PRICE_BDI_COMD + 1: f"PREÇO UNIT. COM BDI DE {meta.bdi.rate_com_desoneracao:.2%}",
        COL_TOTAL_COMD + 1: "TOTAL COM BDI (R$)",
        COL_UNIT_COST_SEMD + 1: "CUSTO UNIT. (R$) SEM D",
        COL_UNIT_PRICE_BDI_SEMD + 1: f"PREÇO UNIT. COM BDI DE {meta.bdi.rate_sem_desoneracao:.2%}",
        COL_TOTAL_SEMD + 1: "TOTAL COM BDI (R$)",
    }
    for col, label in headers.items():
        _safe_set_cell(ws, 19, col, label)


def _clear_mcq_data_rows(ws, start_row: int = MCQ_DATA_START_ROW) -> None:
    from openpyxl.cell.cell import MergedCell

    max_row = max(ws.max_row or start_row, start_row + 2500)
    for row_idx in range(start_row, max_row + 1):
        for col in _MCQ_SYNC_COLS:
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell, MergedCell):
                continue
            try:
                cell.value = None
            except AttributeError:
                continue


def _sync_mcq_metadata(ws, meta: BudgetProjectMetadata) -> None:
    """Atualiza cabeçalho MCQ — preserva fórmulas/títulos das abas ORC."""
    _safe_set_cell(ws, 11, 9, meta.projeto or meta.objeto)
    _safe_set_cell(ws, 12, 9, meta.objeto or meta.projeto)
    _safe_set_cell(ws, 13, 9, meta.local)
    _safe_set_cell(ws, 14, 9, meta.orcamento)
    bdi_tipo = (meta.obra_type or meta.bdi.label or "ED").strip().upper()
    _safe_set_cell(ws, 14, 15, bdi_tipo)


def sync_session_to_mcq_worksheet(
    ws,
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata | None = None,
) -> int:
    """Escreve árvore WBS na aba MCQ (colunas G–M), preservando cabeçalho/rodapé do template."""
    meta = metadata or BudgetProjectMetadata()
    _sync_mcq_metadata(ws, meta)
    _clear_mcq_data_rows(ws)
    row_idx = MCQ_DATA_START_ROW
    for root in roots:
        row_idx = _write_mcq_tree(ws, root, row_idx)
    return row_idx


CRONOGRAMA_ETAPA_START = 9
CRONOGRAMA_ETAPA_STEP = 2
CRONOGRAMA_ETAPA_MAX_ROW = 43


def sync_cronograma_worksheet(
    ws,
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata | None = None,
    schedule: Any | None = None,
) -> int:
    """Preenche cronograma físico-financeiro com etapas e totais do orçamento."""
    meta = metadata or BudgetProjectMetadata()
    _safe_set_cell(ws, 1, 13, meta.objeto or meta.projeto)
    _safe_set_cell(ws, 2, 13, meta.local)
    _safe_set_cell(ws, 3, 13, meta.orcamento)

    prazo_dias = _schedule_total_days(schedule)
    if prazo_dias:
        _safe_set_cell(ws, 4, 12, f"PRAZO DE EXECUÇAO:  {prazo_dias}  DIAS")

    row = CRONOGRAMA_ETAPA_START
    last = row
    for root in roots:
        if root.row_type != ROW_TYPE_ETAPA and root.level != 0:
            continue
        if row > CRONOGRAMA_ETAPA_MAX_ROW:
            break
        code = root.code if "." in str(root.code) else f"{root.code}.0"
        _safe_set_cell(ws, row, 1, code)
        _safe_set_cell(ws, row, 2, root.name)
        total = root.total_price or root.effective_total()
        if total:
            _safe_set_cell(ws, row, 3, round(total, 2))
        last = row
        row += CRONOGRAMA_ETAPA_STEP
    return last


def _schedule_total_days(schedule: Any | None) -> int | None:
    if not schedule:
        return None
    try:
        tasks = getattr(schedule, "tasks", None) or []
        leaves = [t for t in tasks if not getattr(t, "is_summary", False)]
        if leaves:
            return max(int(getattr(t, "duration_days", 0) or 0) for t in leaves) or None
        if getattr(schedule, "project_end", None) and getattr(schedule, "project_start", None):
            from datetime import datetime

            start = datetime.fromisoformat(str(schedule.project_start)[:10])
            end = datetime.fromisoformat(str(schedule.project_end)[:10])
            return max(1, (end - start).days)
    except (TypeError, ValueError, AttributeError):
        return None
    return None


def sync_esp_tecnica_worksheet(
    ws,
    tech_spec: dict[str, Any] | None,
    metadata: BudgetProjectMetadata | None = None,
) -> int:
    """Preenche aba ESP_TECNICA com conteúdo da especificação técnica da sessão."""
    from pricing.spec.tech_spec_models import TechSpecDocument

    meta = metadata or BudgetProjectMetadata()
    doc = TechSpecDocument.from_dict(tech_spec) if tech_spec else None

    _safe_set_cell(ws, 1, 1, meta.projeto or meta.objeto or "PROJETO")
    _safe_set_cell(ws, 2, 1, doc.title if doc else "ESPECIFICAÇÃO TÉCNICA")

    body = ""
    if doc:
        body = (doc.markdown or "").strip() or _html_to_plain(doc.html_content)

    row = 4
    if not body:
        _safe_set_cell(
            ws,
            row,
            1,
            "(Conteúdo não gerado — use a aba Especificação no orçamento para gerar o texto)",
        )
        return row

    for line in body.splitlines():
        text = line.rstrip()
        if text:
            _safe_set_cell(ws, row, 1, text)
            row += 1
    return max(row - 1, 4)


def _html_to_plain(html: str) -> str:
    import re

    text = re.sub(r"<br\s*/?>", "\n", html or "", flags=re.I)
    text = re.sub(r"</p>", "\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def sync_session_to_workbook(
    wb,
    roots: list[BudgetItem],
    metadata: BudgetProjectMetadata | None = None,
    *,
    schedule: Any | None = None,
    tech_spec: dict[str, Any] | None = None,
) -> dict[str, int | None]:
    """Sincroniza MCQ + abas auxiliares presentes no template."""
    meta = metadata or BudgetProjectMetadata()
    sheet_map = {n.upper(): n for n in wb.sheetnames}
    result: dict[str, int | None] = {"last_mcq_row": None, "last_cronograma_row": None, "last_esp_row": None}

    if "MCQ" in sheet_map:
        last_mcq = sync_session_to_mcq_worksheet(wb[sheet_map["MCQ"]], roots, meta)
        result["last_mcq_row"] = last_mcq - 1 if last_mcq else None

    if "CRONOGRAMA" in sheet_map:
        last_cron = sync_cronograma_worksheet(wb[sheet_map["CRONOGRAMA"]], roots, meta, schedule)
        result["last_cronograma_row"] = last_cron

    if "ESP_TECNICA" in sheet_map:
        last_esp = sync_esp_tecnica_worksheet(wb[sheet_map["ESP_TECNICA"]], tech_spec, meta)
        result["last_esp_row"] = last_esp

    return result


def _write_mcq_tree(ws, item: BudgetItem, row_idx: int) -> int:
    """Sync MCQ: apenas colunas de entrada (G–M). Totais ficam na PLANILHA (fórmulas)."""
    if item.metadata.get("is_memory_row") or item.row_type == "MEMORIA":
        _safe_set_cell(ws, row_idx, 10, item.calculation_note or item.name)
        _apply_mcq_wrap(ws, row_idx, 10)
        return row_idx + 1

    is_etapa = item.row_type == ROW_TYPE_ETAPA or (item.item_type.value == "group" and item.level == 0)
    is_sub = item.row_type == ROW_TYPE_SUB_ETAPA

    if is_etapa or is_sub:
        _safe_set_cell(ws, row_idx, 7, ROW_TYPE_ETAPA if is_etapa else ROW_TYPE_SUB_ETAPA)
        _safe_set_cell(ws, row_idx, 8, item.code)
        _safe_set_cell(ws, row_idx, 10, item.name)
        _apply_mcq_wrap(ws, row_idx, 10)
        row_idx += 1
        for child in item.children:
            row_idx = _write_mcq_tree(ws, child, row_idx)
        return row_idx

    source_code = (item.source_code or "").strip()
    _safe_set_cell(ws, row_idx, 7, ROW_TYPE_SERVICO)
    _safe_set_cell(ws, row_idx, 8, item.code)
    _safe_set_cell(ws, row_idx, 9, source_code or None)
    if source_code and _uses_base_vlookup(source_code):
        _safe_set_cell(ws, row_idx, 10, _MCQ_DESC_FORMULA.format(row=row_idx))
        _safe_set_cell(ws, row_idx, 11, _MCQ_UNIT_FORMULA.format(row=row_idx))
    else:
        desc = (item.name or item.pricing_query or "").strip()
        _safe_set_cell(ws, row_idx, 10, desc or None)
        _safe_set_cell(ws, row_idx, 11, item.unit or None)
        _apply_mcq_wrap(ws, row_idx, 10)
    _safe_set_cell(ws, row_idx, 12, item.quantity or None)
    _safe_set_cell(ws, row_idx, 13, item.bdi_label or "BDI1")

    row_idx += 1
    for child in item.children:
        if child.metadata.get("is_memory_row") or child.row_type == "MEMORIA":
            _safe_set_cell(ws, row_idx, 10, child.calculation_note or child.name)
            _apply_mcq_wrap(ws, row_idx, 10)
            row_idx += 1
        else:
            row_idx = _write_mcq_tree(ws, child, row_idx)

    if item.calculation_note and not any(
        c.metadata.get("is_memory_row") for c in item.children
    ):
        _safe_set_cell(ws, row_idx, 10, item.calculation_note)
        _apply_mcq_wrap(ws, row_idx, 10)
        row_idx += 1

    return row_idx


def _apply_mcq_wrap(ws, row: int, col: int) -> None:
    try:
        from openpyxl.styles import Alignment

        cell = ws.cell(row=row, column=col)
        if cell.value in (None, ""):
            return
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        text = str(cell.value)
        line_count = max(1, text.count("\n") + 1, len(text) // 48 + 1)
        ws.row_dimensions[row].height = min(120, max(15, line_count * 15))
    except Exception:
        pass


def _write_ppd_tree(ws, item: BudgetItem, row_idx: int, meta: BudgetProjectMetadata) -> int:
    if item.metadata.get("is_memory_row"):
        _safe_set_cell(ws, row_idx, COL_DESCRIPTION + 1, item.calculation_note or item.name)
        return row_idx + 1

    is_etapa = item.row_type == ROW_TYPE_ETAPA or (item.item_type.value == "group" and item.level == 0)
    is_sub = item.row_type == ROW_TYPE_SUB_ETAPA

    if is_etapa or is_sub:
        _safe_set_cell(ws, row_idx, COL_ROW_TYPE + 1, ROW_TYPE_ETAPA if is_etapa else ROW_TYPE_SUB_ETAPA)
        _safe_set_cell(ws, row_idx, COL_ITEM + 1, item.code)
        _safe_set_cell(ws, row_idx, COL_DESCRIPTION + 1, item.name)
        _safe_set_cell(ws, row_idx, COL_TOTAL_COMD + 1, item.total_price or None)
        row_idx += 1
        for child in item.children:
            row_idx = _write_ppd_tree(ws, child, row_idx, meta)
        return row_idx

    _safe_set_cell(ws, row_idx, COL_ROW_TYPE + 1, ROW_TYPE_SERVICO)
    _safe_set_cell(ws, row_idx, COL_ITEM + 1, item.code)
    _safe_set_cell(ws, row_idx, COL_SOURCE_CODE + 1, item.source_code)
    _safe_set_cell(ws, row_idx, COL_DESCRIPTION + 1, item.name)
    _safe_set_cell(ws, row_idx, COL_UNIT + 1, item.unit)
    _safe_set_cell(ws, row_idx, COL_QUANTITY + 1, item.quantity or None)
    _safe_set_cell(ws, row_idx, COL_UNIT_COST_COMD + 1, item.unit_cost or None)
    _safe_set_cell(ws, row_idx, COL_UNIT_PRICE_BDI_COMD + 1, item.unit_price or None)
    _safe_set_cell(ws, row_idx, COL_TOTAL_COMD + 1, item.total_price or None)
    _safe_set_cell(ws, row_idx, COL_UNIT_COST_SEMD + 1, item.unit_cost_semd or None)
    _safe_set_cell(ws, row_idx, COL_UNIT_PRICE_BDI_SEMD + 1, item.unit_price_semd or None)
    _safe_set_cell(ws, row_idx, COL_TOTAL_SEMD + 1, item.total_price_semd or None)
    _safe_set_cell(ws, row_idx, 13, "BDI1")

    row_idx += 1
    for child in item.children:
        if child.metadata.get("is_memory_row") or child.row_type == "MEMORIA":
            _safe_set_cell(ws, row_idx, COL_DESCRIPTION + 1, child.calculation_note or child.name)
            row_idx += 1
        else:
            row_idx = _write_ppd_tree(ws, child, row_idx, meta)

    if item.calculation_note and not any(
        c.metadata.get("is_memory_row") for c in item.children
    ):
        _safe_set_cell(ws, row_idx, COL_DESCRIPTION + 1, item.calculation_note)
        row_idx += 1

    return row_idx
