from __future__ import annotations

import logging
from pathlib import Path

logger = logging.getLogger(__name__)

_BASE_NAME = "_BaseAbril2026"
_BASE_RANGE = "$A$1:$F$9000"

SEMINF_2026_MARKERS = frozenset({"ORC_SINTETICO", "ORC_ANALITICO", "CRONOGRAMA", "ESP_TECNICA"})


def is_seminf_2026_workbook(sheetnames: list[str]) -> bool:
    upper = {n.upper() for n in sheetnames}
    return "ORC_SINTETICO" in upper and "MCQ" in upper


def workbook_sheetnames(path: Path) -> list[str]:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception as exc:
        logger.debug("workbook_sheetnames falhou: %s", exc)
        return []


def ensure_base_in_workbook(wb, *, base_template: Path) -> bool:
    """Injeta aba Base_* no workbook aberto e corrige named range _BaseAbril2026."""
    if not base_template.exists():
        return False

    existing = next((n for n in wb.sheetnames if n.startswith("Base_")), None)
    from openpyxl.workbook.defined_name import DefinedName

    if existing:
        attr_text = f"'{existing}'!{_BASE_RANGE}"
        current = wb.defined_names.get(_BASE_NAME)
        current_val = getattr(current, "value", None) or str(current or "")
        if "#REF" in str(current_val) or current_val != attr_text:
            wb.defined_names[_BASE_NAME] = DefinedName(name=_BASE_NAME, attr_text=attr_text)
            return True
        return False

    try:
        import openpyxl
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError:
        return False

    wb_src = openpyxl.load_workbook(base_template, keep_vba=True, read_only=True)
    base_name = next((n for n in wb_src.sheetnames if n.startswith("Base_")), None)
    if not base_name:
        wb_src.close()
        return False

    ws_src = wb_src[base_name]
    ws_dst = wb.create_sheet(base_name)
    for row in ws_src.iter_rows(min_row=1, max_row=9000, max_col=6):
        for cell in row:
            if cell.value is not None:
                ws_dst.cell(cell.row, cell.column, cell.value)
    wb_src.close()

    wb.defined_names[_BASE_NAME] = DefinedName(
        name=_BASE_NAME,
        attr_text=f"'{base_name}'!{_BASE_RANGE}",
    )
    logger.info("Base de preços injetada na workbook: %s", base_name)
    return True


def ensure_base_price_sheet(workbook_path: Path, *, base_template: Path) -> bool:
    """Injeta aba Base_* do template v8.1 se ausente (arquivo no disco)."""
    try:
        import openpyxl
    except ImportError:
        return False

    wb = openpyxl.load_workbook(workbook_path, keep_vba=True)
    changed = ensure_base_in_workbook(wb, base_template=base_template)
    if changed or _BASE_NAME not in wb.defined_names:
        wb.save(workbook_path)
    wb.close()
    return changed
