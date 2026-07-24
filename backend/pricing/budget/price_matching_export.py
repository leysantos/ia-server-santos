"""Exportação Excel (fórmulas nativas) e PDF do módulo Lançar Preços."""

from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from pricing.budget.price_matching_hierarchy import ImportRowKind

# Colunas visíveis (1-based): Item, Código, Base, Descrição, Und, Qtd, PU s/BDI, PU c/BDI, Tot s/BDI, Tot c/BDI, Conf.
COL_ITEM = 1
COL_CODIGO = 2
COL_BASE = 3
COL_DESCRICAO = 4
COL_UND = 5
COL_QTD = 6
COL_PU_SEM = 7
COL_PU_COM = 8
COL_TOT_SEM = 9
COL_TOT_COM = 10
COL_CONF = 11
COL_PU_BASE_HIDDEN = 12  # PU base catálogo (coluna oculta — fórmulas)


def _confidence_label(score: float | None) -> str:
    if score is None:
        return "—"
    pct = score * 100 if score <= 1 else score
    return f"{pct:.0f}%"


def _fmt_brl(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_qty(value: float) -> str:
    return f"{value:,.4f}".replace(",", "X").replace(".", ",").replace("X", ".").rstrip("0").rstrip(",")


def _price_bases_header_lines(job: dict[str, Any]) -> list[str]:
    """Lista bases de preço habilitadas com período/referência."""
    bases = [b for b in (job.get("price_bases") or []) if b.get("enabled", True)]
    lines: list[str] = []
    for base in bases:
        label = str(base.get("label") or base.get("source") or "BASE").strip().upper()
        uf = str(base.get("uf") or "").strip()
        ref = str(base.get("reference") or "").strip()
        chunk = label
        if uf:
            chunk += f" · UF {uf}"
        if ref:
            chunk += f" · Período {ref}"
        lines.append(chunk)
    if lines:
        return lines
    seen: set[str] = set()
    for row in job.get("rows") or []:
        base = str(row.get("base") or "").strip()
        ref = str(row.get("reference") or "").strip()
        if not base:
            continue
        key = f"{base}|{ref}"
        if key in seen:
            continue
        seen.add(key)
        chunk = base
        if ref:
            chunk += f" · Período {ref}"
        lines.append(chunk)
    return lines or ["—"]


def build_export_lines(job: dict[str, Any]) -> list[dict[str, Any]]:
    """Monta linhas de exportação na ordem hierárquica (etapa → sub-etapa → serviço)."""
    hierarchy = job.get("hierarchy") or []
    rows_by_item = {str(r.get("item") or ""): r for r in job.get("rows") or []}

    if hierarchy:
        lines: list[dict[str, Any]] = []
        for node in hierarchy:
            item = str(node.get("item") or "")
            row_type = str(node.get("row_type") or ImportRowKind.SERVICO.value)
            priced = rows_by_item.get(item, {})
            is_header = row_type in (ImportRowKind.ETAPA.value, ImportRowKind.SUB_ETAPA.value)
            lines.append(
                {
                    "item": item,
                    "descricao": node.get("descricao") or "",
                    "unidade": node.get("unidade") or "",
                    "quantidade": float(node.get("quantidade") or 0),
                    "row_type": row_type,
                    "is_header": is_header,
                    "codigo_base": priced.get("codigo_base") or node.get("codigo") or "",
                    "base": priced.get("base") or "",
                    "valor_unitario_base": priced.get("valor_unitario_base"),
                    "valor_unitario": priced.get("valor_unitario"),
                    "valor_total": priced.get("valor_total"),
                    "score_confianca": priced.get("score_confianca"),
                }
            )
        return lines

    return [
        {
            "item": str(row.get("item") or ""),
            "descricao": row.get("descricao_original") or "",
            "unidade": row.get("unidade") or "",
            "quantidade": float(row.get("quantidade") or 0),
            "row_type": ImportRowKind.SERVICO.value,
            "is_header": False,
            "codigo_base": row.get("codigo_base") or "",
            "base": row.get("base") or "",
            "valor_unitario_base": row.get("valor_unitario_base"),
            "valor_unitario": row.get("valor_unitario"),
            "valor_total": row.get("valor_total"),
            "score_confianca": row.get("score_confianca"),
        }
        for row in job.get("rows") or []
    ]


def _catalog_base_unit(line: dict[str, Any], increase: float) -> float | None:
    base_unit = line.get("valor_unitario_base")
    if base_unit is None and line.get("valor_unitario") is not None:
        inc = increase or 1.0
        base_unit = float(line.get("valor_unitario") or 0) / inc
    if base_unit is None:
        return None
    return float(base_unit)


def export_price_matching_xlsx(job: dict[str, Any], dest: Path | None = None) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl necessário") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançar Preços"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    meta_font = Font(bold=True)
    etapa_font = Font(bold=True, size=11)
    sub_font = Font(bold=True, size=10, italic=True)
    footer_font = Font(bold=True)

    bdi_decimal = float(job.get("bdi") or 0)
    bdi_pct = bdi_decimal * 100
    increase = float(job.get("increase_index") or 1.0)

    ws["A1"] = "Cliente"
    ws["B1"] = job.get("cliente") or ""
    ws["A2"] = "Obra"
    ws["B2"] = job.get("obra") or ""
    ws["A3"] = "Data"
    ws["B3"] = datetime.now().strftime("%d/%m/%Y")
    ws["A4"] = "BDI (%)"
    ws["B4"] = bdi_pct
    ws["A5"] = "Índice acréscimo"
    ws["B5"] = increase

    base_lines = _price_bases_header_lines(job)
    ws["A6"] = "Bases de preços"
    ws["A6"].font = meta_font
    for i, bl in enumerate(base_lines):
        ws.cell(row=6 + i, column=2, value=bl)

    meta_end_row = 5 + max(1, len(base_lines))
    for cell in ("A1", "A2", "A3", "A4", "A5"):
        ws[cell].font = meta_font

    header_row = meta_end_row + 2
    headers = [
        "Item",
        "Código",
        "Base de preço",
        "Descrição",
        "Unidade",
        "Quantidade",
        "Preço unit. s/ BDI",
        "Preço unit. c/ BDI",
        "Valor total s/ BDI",
        "Valor total c/ BDI",
        "Conf.",
    ]
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=title)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    increase_ref = "$B$5"
    bdi_pct_ref = "$B$4"
    data_start = header_row + 1
    export_lines = build_export_lines(job)
    service_row_nums: list[int] = []

    for i, line in enumerate(export_lines):
        r = data_start + i
        is_header = bool(line.get("is_header"))
        row_type = line.get("row_type") or ImportRowKind.SERVICO.value

        ws.cell(row=r, column=COL_ITEM, value=line.get("item") or "")
        ws.cell(row=r, column=COL_CODIGO, value="" if is_header else (line.get("codigo_base") or ""))
        ws.cell(row=r, column=COL_BASE, value="" if is_header else (line.get("base") or ""))
        ws.cell(row=r, column=COL_DESCRICAO, value=line.get("descricao") or "")
        ws.cell(row=r, column=COL_UND, value="" if is_header else (line.get("unidade") or ""))
        ws.cell(row=r, column=COL_QTD, value="" if is_header else float(line.get("quantidade") or 0))

        if is_header:
            if row_type == ImportRowKind.ETAPA.value:
                ws.cell(row=r, column=COL_DESCRICAO).font = etapa_font
            else:
                ws.cell(row=r, column=COL_DESCRICAO).font = sub_font
        else:
            service_row_nums.append(r)
            base_unit = _catalog_base_unit(line, increase)
            ws.cell(row=r, column=COL_PU_BASE_HIDDEN, value=base_unit if base_unit is not None else "")
            ws.cell(row=r, column=COL_PU_SEM, value=f"={get_column_letter(COL_PU_BASE_HIDDEN)}{r}*{increase_ref}")
            ws.cell(row=r, column=COL_PU_COM, value=f"={get_column_letter(COL_PU_SEM)}{r}*(1+{bdi_pct_ref}/100)")
            ws.cell(row=r, column=COL_TOT_SEM, value=f"={get_column_letter(COL_QTD)}{r}*{get_column_letter(COL_PU_SEM)}{r}")
            ws.cell(row=r, column=COL_TOT_COM, value=f"={get_column_letter(COL_QTD)}{r}*{get_column_letter(COL_PU_COM)}{r}")
            ws.cell(row=r, column=COL_CONF, value=_confidence_label(line.get("score_confianca")))

        for col in range(1, COL_CONF + 1):
            ws.cell(row=r, column=col).border = border

    ws.column_dimensions[get_column_letter(COL_PU_BASE_HIDDEN)].hidden = True

    if service_row_nums:
        first = service_row_nums[0]
        last = service_row_nums[-1]
        fr1 = last + 2
        fr2 = fr1 + 1
        fr3 = fr2 + 1
        tot_sem_col = get_column_letter(COL_TOT_SEM)
        tot_com_col = get_column_letter(COL_TOT_COM)

        ws.cell(row=fr1, column=COL_DESCRICAO, value="Total s/ BDI").font = footer_font
        ws.cell(row=fr1, column=COL_TOT_SEM, value=f"=SUM({tot_sem_col}{first}:{tot_sem_col}{last})").font = footer_font

        ws.cell(row=fr2, column=COL_DESCRICAO, value="Valor BDI").font = footer_font
        ws.cell(row=fr2, column=COL_TOT_SEM, value=f"={tot_sem_col}{fr1}*{bdi_pct_ref}/100").font = footer_font

        ws.cell(row=fr3, column=COL_DESCRICAO, value="Total c/ BDI").font = footer_font
        ws.cell(row=fr3, column=COL_TOT_COM, value=f"={tot_sem_col}{fr1}+{tot_sem_col}{fr2}").font = footer_font

        for row_idx in (fr1, fr2, fr3):
            for col in range(1, COL_CONF + 1):
                ws.cell(row=row_idx, column=col).border = border

    widths = [8, 14, 12, 40, 8, 10, 14, 14, 16, 16, 8]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    if dest:
        dest.write_bytes(data)
    return data


def export_price_matching_pdf(job: dict[str, Any], dest: Path | None = None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []

    bdi = float(job.get("bdi") or 0)
    increase = float(job.get("increase_index") or 1.0)

    title = Paragraph("<b>Lançamento de Preços — IA Server Santos</b>", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 6))

    meta_rows: list[list[str]] = [
        ["Cliente", job.get("cliente") or "—", "Obra", job.get("obra") or "—"],
        ["Data", datetime.now().strftime("%d/%m/%Y"), "BDI", f"{bdi * 100:.2f}%"],
        ["Índice acréscimo", str(increase), "UF", job.get("uf") or "AM"],
    ]
    base_lines = _price_bases_header_lines(job)
    for i, bl in enumerate(base_lines):
        if i == 0:
            meta_rows.append(["Bases de preços", bl, "", ""])
        else:
            meta_rows.append(["", bl, "", ""])

    meta_table = Table(meta_rows, colWidths=[32 * mm, 72 * mm, 32 * mm, 72 * mm])
    meta_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(meta_table)
    story.append(Spacer(1, 8))

    header = [
        "Item",
        "Código",
        "Base",
        "Descrição",
        "Und",
        "Qtd",
        "PU s/ BDI",
        "PU c/ BDI",
        "Tot. s/ BDI",
        "Tot. c/ BDI",
        "Conf.",
    ]
    data = [header]
    subtotal_sem = 0.0

    for line in build_export_lines(job):
        is_header = bool(line.get("is_header"))
        row_type = line.get("row_type") or ImportRowKind.SERVICO.value
        desc = (line.get("descricao") or "")[:55]
        if is_header:
            prefix = "▸ " if row_type == ImportRowKind.SUB_ETAPA.value else "■ "
            data.append(
                [
                    line.get("item") or "",
                    "",
                    "",
                    prefix + desc,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            continue

        qty = float(line.get("quantidade") or 0)
        base_unit = _catalog_base_unit(line, increase) or 0.0
        pu_sem = base_unit * increase
        pu_com = pu_sem * (1 + bdi)
        tot_sem = pu_sem * qty
        tot_com = pu_com * qty
        subtotal_sem += tot_sem

        data.append(
            [
                line.get("item") or "",
                line.get("codigo_base") or "—",
                line.get("base") or "—",
                desc,
                line.get("unidade") or "",
                _fmt_qty(qty),
                _fmt_brl(pu_sem),
                _fmt_brl(pu_com),
                _fmt_brl(tot_sem),
                _fmt_brl(tot_com),
                _confidence_label(line.get("score_confianca")),
            ]
        )

    valor_bdi = subtotal_sem * bdi
    total_com = subtotal_sem + valor_bdi

    data.append(["", "", "", "Total s/ BDI", "", "", "", "", _fmt_brl(subtotal_sem), "", ""])
    data.append(["", "", "", "Valor BDI", "", "", "", "", _fmt_brl(valor_bdi), "", ""])
    data.append(["", "", "", "Total c/ BDI", "", "", "", "", "", _fmt_brl(total_com), ""])

    table = Table(
        data,
        colWidths=[11 * mm, 16 * mm, 12 * mm, 52 * mm, 10 * mm, 12 * mm, 20 * mm, 20 * mm, 22 * mm, 22 * mm, 11 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, -3), (3, -1), "Helvetica-Bold"),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 10))
    story.append(Paragraph("<i>Gerado pelo IA Server Santos</i>", styles["Normal"]))

    doc.build(story)
    data_bytes = buf.getvalue()
    if dest:
        dest.write_bytes(data_bytes)
    return data_bytes
