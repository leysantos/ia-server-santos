"""Estilos corporativos para Excel do histograma MO."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Azul escuro corporativo — padrão construtora
CORPORATE_BLUE = "0B2E4A"
CORPORATE_BLUE_LIGHT = "E8EEF4"
WHITE = "FFFFFF"
BORDER_GRAY = "CCCCCC"
STRIPE_GRAY = "F7F9FB"
FONT_NAME = "Calibri"


@dataclass(frozen=True)
class HistogramStyles:
    title_font: Font
    header_font: Font
    meta_font: Font
    body_font: Font
    total_font: Font
    header_fill: PatternFill
    total_fill: PatternFill
    stripe_fill: PatternFill
    thin_border: Border
    center: Alignment
    left: Alignment
    right: Alignment


def build_histogram_styles() -> HistogramStyles:
    thin = Side(style="thin", color=BORDER_GRAY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return HistogramStyles(
        title_font=Font(name=FONT_NAME, size=14, bold=True, color=WHITE),
        header_font=Font(name=FONT_NAME, size=10, bold=True, color=WHITE),
        meta_font=Font(name=FONT_NAME, size=10, color="333333"),
        body_font=Font(name=FONT_NAME, size=10, color="1A1A1A"),
        total_font=Font(name=FONT_NAME, size=10, bold=True, color=WHITE),
        header_fill=PatternFill("solid", fgColor=CORPORATE_BLUE),
        total_fill=PatternFill("solid", fgColor=CORPORATE_BLUE),
        stripe_fill=PatternFill("solid", fgColor=STRIPE_GRAY),
        thin_border=border,
        center=Alignment(horizontal="center", vertical="center", wrap_text=True),
        left=Alignment(horizontal="left", vertical="center", wrap_text=True),
        right=Alignment(horizontal="right", vertical="center"),
    )


def apply_cell_style(cell, *, font=None, fill=None, alignment=None, border=None) -> None:
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
