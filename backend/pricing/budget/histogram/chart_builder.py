"""Gráfico empilhado por tipo de recurso (uma cor por linha)."""

from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pricing.budget.histogram.section_mapper import HistogramSection
from pricing.budget.histogram.style_builder import build_histogram_styles

HISTOGRAM_ITEM_COLORS = (
    "0B2E4A",
    "F59E0B",
    "10B981",
    "F472B6",
    "A78BFA",
    "FB7185",
    "34D399",
    "60A5FA",
    "FBBF24",
    "C084FC",
    "2DD4BF",
    "F97316",
)


def attach_section_stacked_chart(
    ws_data: Worksheet,
    ws_chart: Worksheet,
    section: HistogramSection,
    *,
    table_header_row: int,
    data_start_row: int,
    first_period_col: int,
) -> None:
    """Gráfico de colunas empilhadas — uma série por item da tabela."""
    period_count = section.period_count
    if period_count <= 0 or not section.items:
        return

    last_period_col = first_period_col + period_count - 1
    cats = Reference(
        ws_data,
        min_col=first_period_col,
        min_row=table_header_row,
        max_col=last_period_col,
        max_row=table_header_row,
    )

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = section.title
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = "Período (dias)"
    chart.style = 2
    chart.legend.position = "b"

    data_end_row = data_start_row + len(section.items) - 1
    for i, item in enumerate(section.items):
        row_num = data_start_row + i
        values = Reference(
            ws_data,
            min_col=first_period_col,
            min_row=row_num,
            max_col=last_period_col,
            max_row=row_num,
        )
        chart.add_data(values, titles_from_data=False)
        if chart.series:
            series = chart.series[-1]
            color = HISTOGRAM_ITEM_COLORS[i % len(HISTOGRAM_ITEM_COLORS)]
            series.graphicalProperties.solidFill = color
            series.title = item.description[:31]

    chart.set_categories(cats)
    chart.width = 22
    chart.height = 12
    ws_chart.add_chart(chart, "B2")

    styles = build_histogram_styles()
    ws_chart["A1"] = section.title
    ws_chart["A1"].font = styles.title_font
    ws_chart.merge_cells("A1:H1")
    ws_chart["A2"] = (
        f"{len(section.items)} tipo(s) · períodos "
        f"{get_column_letter(first_period_col)}–{get_column_letter(last_period_col)}"
    )
