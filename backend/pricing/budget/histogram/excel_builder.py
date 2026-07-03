"""Montagem Excel — seções MO e equipamentos conforme planilha-exemplo."""

from __future__ import annotations

import io
import tempfile
from datetime import date

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.system.company_profile import CompanyProfile
from pricing.budget.budget_export_branding import ExportBrandingConfig
from pricing.budget.histogram.chart_builder import attach_section_stacked_chart
from pricing.budget.histogram.section_mapper import HistogramReport, HistogramSection
from pricing.budget.histogram.style_builder import apply_cell_style, build_histogram_styles


def _write_section_header(
    ws: Worksheet,
    report: HistogramReport,
    section: HistogramSection,
    *,
    profile: CompanyProfile,
    brand: ExportBrandingConfig,
    logo_bytes: bytes | None,
    col_count: int,
) -> int:
    styles = build_histogram_styles()
    empresa = report.empresa or profile.company_name or brand.header_line1 or "Empresa"

    if logo_bytes and brand.show_logo and ws.max_row == 1:
        try:
            from openpyxl.drawing.image import Image

            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                tmp.write(logo_bytes)
                tmp_path = tmp.name
            img = Image(tmp_path)
            img.width = 72
            img.height = 56
            ws.add_image(img, "A1")
        except Exception:
            pass

    row = max(ws.max_row, 1) + 1 if ws.max_row > 1 else 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=section.title)
    apply_cell_style(cell, font=styles.title_font, fill=styles.header_fill, alignment=styles.center)
    ws.row_dimensions[row].height = 28
    row += 1

    cliente_line = f"CLIENTE: {report.cliente} — {report.obra}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=cliente_line)
    apply_cell_style(cell, font=styles.meta_font, alignment=styles.left)
    row += 2

    emission = date.today().strftime("%d/%m/%Y")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=f"Empresa: {empresa} · Emissão: {emission}")
    apply_cell_style(cell, font=styles.meta_font, alignment=styles.left)
    return row + 2


def _write_section_table(
    ws: Worksheet,
    section: HistogramSection,
    start_row: int,
) -> tuple[int, int, int]:
    styles = build_histogram_styles()
    period_count = section.period_count
    header_row = start_row

    headers = ["ITENS", "DISCRIMINAÇÃO", *[str(d) for d in section.period_labels], "TOTAL"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        apply_cell_style(
            cell,
            font=styles.header_font,
            fill=styles.header_fill,
            alignment=styles.center if col > 2 else styles.left,
            border=styles.thin_border,
        )

    data_start = header_row + 1
    first_value_col = 3
    last_value_col = first_value_col + period_count - 1
    total_col = last_value_col + 1

    for i, item in enumerate(section.items):
        r = data_start + i
        stripe = i % 2 == 1
        fill = styles.stripe_fill if stripe else None

        c_item = ws.cell(row=r, column=1, value=item.index)
        c_desc = ws.cell(row=r, column=2, value=item.description)
        apply_cell_style(
            c_item, font=styles.body_font, fill=fill, alignment=styles.center, border=styles.thin_border
        )
        apply_cell_style(
            c_desc, font=styles.body_font, fill=fill, alignment=styles.left, border=styles.thin_border
        )

        for j in range(period_count):
            col = first_value_col + j
            val = item.values[j] if j < len(item.values) else 0
            cell = ws.cell(row=r, column=col, value=round(val, 2) if val else None)
            apply_cell_style(
                cell, font=styles.body_font, fill=fill, alignment=styles.center, border=styles.thin_border
            )
            cell.number_format = "0.00"

        first_l = get_column_letter(first_value_col)
        last_l = get_column_letter(last_value_col)
        total_cell = ws.cell(row=r, column=total_col, value=f"=SUM({first_l}{r}:{last_l}{r})")
        apply_cell_style(
            total_cell, font=styles.body_font, fill=fill, alignment=styles.center, border=styles.thin_border
        )
        total_cell.number_format = "0.00"

    total_row = data_start + len(section.items)
    total_label = ws.cell(row=total_row, column=2, value="TOTAL")
    apply_cell_style(
        total_label,
        font=styles.total_font,
        fill=styles.total_fill,
        alignment=styles.left,
        border=styles.thin_border,
    )
    ws.cell(row=total_row, column=1, value="")
    apply_cell_style(
        ws.cell(row=total_row, column=1),
        font=styles.total_font,
        fill=styles.total_fill,
        border=styles.thin_border,
    )

    for j in range(period_count):
        col = first_value_col + j
        col_l = get_column_letter(col)
        cell = ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({col_l}{data_start}:{col_l}{total_row - 1})",
        )
        apply_cell_style(
            cell,
            font=styles.total_font,
            fill=styles.total_fill,
            alignment=styles.center,
            border=styles.thin_border,
        )
        cell.number_format = "0.00"

    grand = ws.cell(
        row=total_row,
        column=total_col,
        value=f"=SUM({get_column_letter(first_value_col)}{total_row}:{get_column_letter(last_value_col)}{total_row})",
    )
    apply_cell_style(
        grand,
        font=styles.total_font,
        fill=styles.total_fill,
        alignment=styles.center,
        border=styles.thin_border,
    )
    grand.number_format = "0.00"

    return header_row, data_start, total_row


def _apply_section_layout(ws: Worksheet, period_count: int, *, table_header_row: int) -> None:
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    for i in range(period_count):
        ws.column_dimensions[get_column_letter(3 + i)].width = 9
    ws.column_dimensions[get_column_letter(3 + period_count)].width = 10
    ws.freeze_panes = f"C{table_header_row + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1


def _write_section_to_sheet(
    ws: Worksheet,
    report: HistogramReport,
    section: HistogramSection,
    *,
    profile: CompanyProfile,
    brand: ExportBrandingConfig,
    logo_bytes: bytes | None,
) -> tuple[int, int, int]:
    col_count = 2 + section.period_count + 1
    table_start = _write_section_header(
        ws, report, section, profile=profile, brand=brand, logo_bytes=logo_bytes, col_count=col_count
    )
    header_row, data_start, total_row = _write_section_table(ws, section, table_start)
    _apply_section_layout(ws, section.period_count, table_header_row=header_row)
    return header_row, data_start, total_row


def build_histogram_report_workbook_bytes(
    report: HistogramReport,
    *,
    profile: CompanyProfile,
    brand: ExportBrandingConfig,
    logo_bytes: bytes | None = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sections: list[tuple[str, str, HistogramSection]] = []
    if report.mao_obra and report.mao_obra.items:
        sections.append(("HISTOGRAMA MO", "GRAFICO MO", report.mao_obra))
    if report.equipamento and report.equipamento.items:
        sections.append(("HISTOGRAMA EQ", "GRAFICO EQ", report.equipamento))

    if not sections:
        ws = wb.create_sheet("HISTOGRAMA")
        ws["A1"] = "Sem dados de mão de obra ou equipamentos — sincronize cronograma e CPUs."
    else:
        for sheet_name, chart_sheet_name, section in sections:
            ws = wb.create_sheet(sheet_name[:31])
            header_row, data_start, total_row = _write_section_to_sheet(
                ws, report, section, profile=profile, brand=brand, logo_bytes=logo_bytes
            )
            ws_chart = wb.create_sheet(chart_sheet_name[:31])
            attach_section_stacked_chart(
                ws,
                ws_chart,
                section,
                table_header_row=header_row,
                data_start_row=data_start,
                first_period_col=3,
            )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
