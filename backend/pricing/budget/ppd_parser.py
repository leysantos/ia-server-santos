from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from pricing.budget.ppd_layout import (
    COL_DESCRIPTION,
    COL_ITEM,
    COL_QUANTITY,
    COL_ROW_TYPE,
    COL_SOURCE_CODE,
    COL_TOTAL_COMD,
    COL_UNIT,
    COL_UNIT_COST_COMD,
    COL_UNIT_COST_SEMD,
    COL_UNIT_PRICE_BDI_COMD,
    COL_UNIT_PRICE_BDI_SEMD,
    COL_TOTAL_SEMD,
    DATA_START_ROW,
    HEADER_LABELS,
    PPD_TEMPLATE_ID,
    ROW_TYPE_ETAPA,
    ROW_TYPE_SERVICO,
    ROW_TYPE_SUB_ETAPA,
)
from pricing.budget.bdi_types import detect_obra_type, normalize_obra_type
from pricing.models.budget_item import BudgetItem, BudgetItemType
from pricing.models.budget_metadata import BdiConfig, BudgetProjectMetadata


def _cell(row: tuple, idx: int) -> Any:
    return row[idx] if len(row) > idx else None


def _str_val(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _float_val(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(str(val).replace(",", "."))
    except ValueError:
        return 0.0


def _level_from_item_code(code: str) -> int:
    if not code or code.isdigit():
        return 0
    parts = code.split(".")
    return max(0, len(parts) - 1)


def parse_project_metadata(ws) -> BudgetProjectMetadata:
    meta = BudgetProjectMetadata(template=PPD_TEMPLATE_ID)
    obra_type: str | None = None

    for row in ws.iter_rows(values_only=True, max_row=20):
        cells = [_str_val(c) for c in row]
        joined = " ".join(c for c in cells if c)

        for i, cell in enumerate(cells):
            upper = cell.upper().rstrip(":")
            if upper in ("PROJETO", "PROJETO:"):
                meta.projeto = cells[i + 1] if i + 1 < len(cells) else ""
            elif upper in ("OBJETO", "OBJETO:"):
                meta.objeto = cells[i + 1] if i + 1 < len(cells) else ""
            elif upper in ("LOCAL", "LOCAL:"):
                meta.local = cells[i + 1] if i + 1 < len(cells) else ""
            elif upper.startswith("ORÇAMENTO") or upper.startswith("ORCAMENTO"):
                meta.orcamento = cells[i + 1] if i + 1 < len(cells) else ""
            elif "BASE DE PREÇO" in upper or "BASE DE PRECO" in upper:
                val = cells[i + 1] if i + 1 < len(cells) and cells[i + 1] else cell
                if val and val != cell:
                    meta.base_preco = val
                elif "UTILIZADA" in cell.upper():
                    meta.base_preco = cell.split(":", 1)[-1].strip() if ":" in cell else cell
                else:
                    meta.base_preco = cell
            elif "TIPO DE" in upper and "BDI" in upper:
                # célula seguinte = RF, ED, etc.
                nxt = cells[i + 1] if i + 1 < len(cells) else ""
                if nxt.upper() in ("ED", "RF", "FIE", "IE", "EI", "OPMF", "SEE", "AG"):
                    obra_type = normalize_obra_type(nxt)

        if "SEMINF" in joined and not meta.orgao:
            meta.orgao = "SEMINF"
        if "PROCESSO" in joined.upper():
            m = re.search(r"(\d{4}/\d+/\d+/\d+)", joined)
            if m:
                meta.processo = m.group(1)

    if not obra_type:
        obra_type = detect_obra_type(
            orcamento=meta.orcamento,
            objeto=meta.objeto or meta.projeto,
        )

    meta.obra_type = obra_type
    meta.bdi = BdiConfig.from_obra_type(obra_type)
    if not meta.projeto and meta.objeto:
        meta.projeto = meta.objeto
    return meta


def parse_ppd_sheet(ws, sheet_name: str = "PLANILHA") -> tuple[BudgetProjectMetadata, list[BudgetItem]]:
    """Extrai árvore de orçamento de aba PLANILHA ou MCQ."""
    metadata = parse_project_metadata(ws)
    rows = list(ws.iter_rows(values_only=True))
    items: list[BudgetItem] = []
    stack: list[BudgetItem] = []
    pending_memory: str | None = None
    seen_codes: set[str] = set()

    def _unique_code(item_code: str, parent: str, row_idx: int) -> str:
        if item_code not in seen_codes:
            seen_codes.add(item_code)
            return item_code
        alt = f"{parent}-{item_code}" if parent else f"{item_code}-r{row_idx}"
        seen_codes.add(alt)
        return alt

    for row_idx, row in enumerate(rows):
        if row_idx < DATA_START_ROW:
            continue

        row_type = _str_val(_cell(row, COL_ROW_TYPE)).upper()
        item_code = _str_val(_cell(row, COL_ITEM))
        source_code = _str_val(_cell(row, COL_SOURCE_CODE))
        description = _str_val(_cell(row, COL_DESCRIPTION))
        unit = _str_val(_cell(row, COL_UNIT))
        quantity = _float_val(_cell(row, COL_QUANTITY))

        if row_type in ("EXEMPLO", "ITEM", "CÓDIGO", "CODIGO"):
            continue

        if not description and not item_code and not row_type:
            continue

        if row_type == ROW_TYPE_ETAPA or (item_code and item_code.isdigit() and not source_code and description):
            node = BudgetItem(
                row_id=f"ppd-{row_idx}",
                code=item_code or str(len(items) + 1),
                name=description,
                level=0,
                quantity=0,
                unit="",
                unit_cost=0,
                unit_price=0,
                total_price=_float_val(_cell(row, COL_TOTAL_COMD)),
                row_type=ROW_TYPE_ETAPA,
                item_type=BudgetItemType.GROUP,
                metadata={"sheet": sheet_name},
            )
            items.append(node)
            stack = [node]
            pending_memory = None
            continue

        if row_type == ROW_TYPE_SUB_ETAPA:
            node = BudgetItem(
                row_id=f"ppd-{row_idx}",
                code=item_code,
                name=description,
                level=1,
                quantity=0,
                unit="",
                unit_cost=0,
                unit_price=0,
                total_price=_float_val(_cell(row, COL_TOTAL_COMD)),
                row_type=ROW_TYPE_SUB_ETAPA,
                item_type=BudgetItemType.GROUP,
                parent_code=stack[-1].code if stack else None,
                metadata={"sheet": sheet_name},
            )
            if stack:
                stack[-1].children.append(node)
            else:
                items.append(node)
            stack = stack[:1] + [node] if stack else [node]
            pending_memory = None
            continue

        if row_type == ROW_TYPE_SERVICO or (item_code and "." in item_code and description):
            unit_cost = _float_val(_cell(row, COL_UNIT_COST_COMD))
            unit_price = _float_val(_cell(row, COL_UNIT_PRICE_BDI_COMD))
            total = _float_val(_cell(row, COL_TOTAL_COMD))
            unit_cost_semd = _float_val(_cell(row, COL_UNIT_COST_SEMD))
            unit_price_semd = _float_val(_cell(row, COL_UNIT_PRICE_BDI_SEMD))
            total_semd = _float_val(_cell(row, COL_TOTAL_SEMD))

            if not unit_price and unit_cost:
                unit_price = metadata.bdi.price_with_bdi(unit_cost, True)
            if not total and quantity and unit_price:
                total = round(quantity * unit_price, 2)

            node = BudgetItem(
                row_id=f"ppd-{row_idx}",
                code=_unique_code(
                    item_code,
                    stack[-1].code if stack else "",
                    row_idx,
                ),
                name=description,
                level=_level_from_item_code(item_code) + 1,
                quantity=quantity,
                unit=unit,
                unit_cost=unit_cost,
                unit_cost_semd=unit_cost_semd or unit_cost,
                unit_price=unit_price,
                unit_price_semd=unit_price_semd,
                total_price=total,
                total_price_semd=total_semd,
                source_code=source_code.strip(),
                source_base=_service_source_base(source_code.strip(), metadata),
                row_type=ROW_TYPE_SERVICO,
                item_type=BudgetItemType.COMPOSITION,
                calculation_note=pending_memory or "",
                bdi_rate=metadata.bdi.rate_com_desoneracao,
                bdi_label=metadata.bdi.label,
                parent_code=stack[-1].code if stack else None,
                metadata={"sheet": sheet_name},
            )
            if stack:
                stack[-1].children.append(node)
            else:
                items.append(node)
            pending_memory = None
            continue

        if description and not item_code and not row_type:
            pending_memory = description
            mem = BudgetItem(
                row_id=f"ppd-{row_idx}-m",
                code=f"{stack[-1].code}.m{len(stack[-1].children)+1}" if stack else f"m{row_idx}",
                name=description,
                level=(stack[-1].level + 1) if stack else 1,
                quantity=0,
                unit="",
                unit_cost=0,
                unit_price=0,
                total_price=0,
                row_type="MEMORIA",
                item_type=BudgetItemType.INPUT,
                calculation_note=description,
                parent_code=stack[-1].code if stack else None,
                metadata={"sheet": sheet_name, "is_memory_row": True},
            )
            if stack:
                stack[-1].children.append(mem)

    for root in items:
        root.recompute_total()

    return metadata, items


def parse_ppd_workbook(path: str | Path) -> tuple[BudgetProjectMetadata, list[BudgetItem], dict[str, Any]]:
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info: dict[str, Any] = {"sheets": wb.sheetnames, "path": str(path)}

    sheet_plan = "PLANILHA" if "PLANILHA" in wb.sheetnames else None
    sheet_mcq = "MCQ" if "MCQ" in wb.sheetnames else None

    if sheet_plan:
        metadata, items = parse_ppd_sheet(wb[sheet_plan], sheet_plan)
    elif sheet_mcq:
        metadata, items = parse_ppd_sheet(wb[sheet_mcq], sheet_mcq)
    else:
        metadata, items = parse_ppd_sheet(wb[wb.sheetnames[0]], wb.sheetnames[0])

    if sheet_mcq and sheet_plan:
        _merge_mcq_memories(wb[sheet_mcq], items)

    base_count = 0
    base_sheet = pick_base_sheet_name(wb.sheetnames)
    if base_sheet:
        base_count = _load_base_sheet(wb[base_sheet])
        info["base_sheet"] = base_sheet
        info["base_items"] = base_count

    wb.close()
    return metadata, items, info


def _merge_mcq_memories(ws, roots: list[BudgetItem]) -> None:
    """Anexa memórias de cálculo da aba MCQ aos serviços correspondentes."""
    rows = list(ws.iter_rows(values_only=True))
    service_map: dict[str, BudgetItem] = {}

    def _walk(item: BudgetItem) -> None:
        if item.row_type == ROW_TYPE_SERVICO:
            service_map[item.code] = item
        for c in item.children:
            _walk(c)

    for root in roots:
        _walk(root)

    pending_svc: BudgetItem | None = None
    for row_idx, row in enumerate(rows):
        if row_idx < DATA_START_ROW:
            continue
        row_type = _str_val(_cell(row, COL_ROW_TYPE)).upper()
        item_code = _str_val(_cell(row, COL_ITEM))
        description = _str_val(_cell(row, COL_DESCRIPTION))

        if row_type == ROW_TYPE_SERVICO and item_code:
            pending_svc = service_map.get(item_code)
            continue

        if description and not item_code and not row_type and pending_svc:
            note = description
            pending_svc.calculation_note = note
            mem = BudgetItem(
                row_id=f"ppd-mcq-{row_idx}-m",
                code=f"{pending_svc.code}.m{len(pending_svc.children)+1}",
                name=note,
                level=pending_svc.level + 1,
                quantity=0,
                unit="",
                unit_cost=0,
                unit_price=0,
                total_price=0,
                row_type="MEMORIA",
                item_type=BudgetItemType.INPUT,
                calculation_note=note,
                parent_code=pending_svc.code,
                metadata={"is_memory_row": True, "sheet": "MCQ"},
            )
            pending_svc.children.append(mem)


def _load_base_sheet(ws) -> int:
    """Extrai insumos da aba Base (CODIGO, DESCRICAO, und, preços)."""
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    for row in rows:
        code = _str_val(row[0] if row else "")
        if not _is_valid_price_base_code(code):
            continue
        count += 1
    return count


_PRICE_BASE_CODE_RE = re.compile(
    r"^\d+(?:\.\d+)*(?:\.SEMINF)?$",
    re.IGNORECASE,
)


def _is_valid_price_base_code(code: str) -> bool:
    """Aceita código SINAPI nacional (105114) e composições regionais (107071.1.9.SEMINF)."""
    raw = (code or "").strip()
    if not raw:
        return False
    if raw.isdigit():
        return True
    return bool(_PRICE_BASE_CODE_RE.match(raw))


def _service_source_base(source_code: str, metadata: BudgetProjectMetadata) -> str:
    """Fonte da composição no orçamento — regional SEMINF não usa banco SINAPI."""
    from pricing.budget.seminf_base_parser import is_seminf_regional_code

    code = (source_code or "").strip()
    if is_seminf_regional_code(code):
        return "DP-SEMINF"
    if metadata.base_preco:
        first = metadata.base_preco.split("/")[0].strip()
        if first:
            return first
    return "SINAPI"


def pick_base_sheet_name(sheet_names: list[str]) -> str | None:
    """Escolhe aba Base — prefere cópia de trabalho sobre ORIGINAL."""
    bases = [n for n in sheet_names if n.lower().startswith("base")]
    if not bases:
        return None
    lowered = {n: n.lower() for n in bases}
    for name, low in lowered.items():
        if "copia" in low or "copy" in low:
            return name
    non_original = [n for n in bases if "original" not in lowered[n]]
    if non_original:
        return non_original[0]
    return bases[0]


def _find_tp2_col(headers: list[str]) -> int | None:
    for idx, cell in enumerate(headers):
        if _str_val(cell).lower() == "tp2":
            return idx
    return None


def extract_price_base_rows(path: str | Path, *, sheet_name: str | None = None) -> list[dict]:
    """Retorna linhas da aba Base para carregar no provider SINAPI."""
    import openpyxl

    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    base_name = sheet_name or pick_base_sheet_name(wb.sheetnames)
    base_ws = wb[base_name] if base_name and base_name in wb.sheetnames else None
    if not base_ws:
        wb.close()
        return []

    rows = list(base_ws.iter_rows(values_only=True))
    header_idx = _find_base_header_row(rows)
    data_rows = rows[header_idx + 1 :] if header_idx is not None else rows
    header_row = rows[header_idx] if header_idx is not None else ()
    headers = [_str_val(c) for c in header_row[:12]]
    tp2_i = _find_tp2_col(headers)

    items: list[dict] = []
    for row in data_rows:
        code = _str_val(row[0] if row else "")
        if not _is_valid_price_base_code(code):
            continue
        desc = _str_val(row[1] if len(row) > 1 else "")
        unit = _str_val(row[2] if len(row) > 2 else "un")
        price_com = _float_val(row[3] if len(row) > 3 else 0)
        price_sem = _float_val(row[4] if len(row) > 4 else price_com)
        tp2 = ""
        if tp2_i is not None and len(row) > tp2_i:
            tp2 = _str_val(row[tp2_i])
        items.append(
            {
                "code": code,
                "description": desc,
                "unit": unit,
                "price": price_com,
                "metadata": {
                    "price_sem_desoneracao": price_sem,
                    "source": "PPD_BASE",
                    "base_sheet": base_name,
                    "tp2": tp2,
                },
            }
        )
    wb.close()
    return items


def _find_base_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    for idx, row in enumerate(rows[:30]):
        cells = [_str_val(c).upper() for c in row[:8]]
        if not cells:
            continue
        if cells[0] in ("CODIGO", "CÓDIGO") and any("DESCR" in c for c in cells):
            return idx
    return None
