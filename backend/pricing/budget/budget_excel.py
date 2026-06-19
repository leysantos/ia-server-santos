from __future__ import annotations

import io
from typing import Any

from pricing.models.budget_item import BudgetItem


def export_budget_xlsx(
    roots: list[BudgetItem],
    title: str = "Orçamento",
    metadata: dict[str, Any] | None = None,
) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ImportError("openpyxl necessário: pip install openpyxl") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Código",
        "Descrição",
        "Qtd",
        "Un",
        "Preço Unit.",
        "Total",
        "Base",
        "Cód. Base",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    for root in roots:
        row_idx = _append_item(ws, root, row_idx)

    ws.append([])
    ws.append(["", "TOTAL GERAL", "", "", "", sum(r.total_price for r in roots)])

    if metadata:
        ws_meta = wb.create_sheet("Metadados")
        for key, val in metadata.items():
            ws_meta.append([str(key), str(val)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _append_item(ws, item: BudgetItem, row_idx: int) -> int:
    indent = "  " * item.level
    ws.append(
        [
            item.code,
            f"{indent}{item.name}",
            item.quantity if item.quantity else "",
            item.unit,
            item.unit_price if item.unit_price else "",
            item.total_price if item.total_price else "",
            item.source_base,
            item.source_code,
        ]
    )
    row_idx += 1
    for child in item.children:
        row_idx = _append_item(ws, child, row_idx)
    return row_idx
