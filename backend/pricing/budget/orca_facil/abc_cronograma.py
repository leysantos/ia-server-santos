"""OF8 — Preenche CURVA_ABC e CRONOGRAMA na cópia do modelo (paridade Cursor).

Espelha CONT_DREN `_fill_abc_crono.py` + `_fix_cronograma.py` / `_fix_crono_final.py`:
- CURVA_ABC: cenário adotado = menor total entre ComD e SemD (paridade MCQ!V14/V15);
  serviços ordenados por total c/ BDI + fórmulas % / classe A-B-C
- CRONOGRAMA: administração em todos os meses; demais etapas em barras Gantt
  ao longo do prazo; valores via fórmulas PLANILHA (ComD/SemD conforme MCQ!V15)
"""

from __future__ import annotations

import logging
import math
import re
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.page import PageMargins

from pricing.budget.bdi_types import get_obra_bdi
from pricing.budget.orca_facil.base_index import ModelPriceBaseIndex

logger = logging.getLogger(__name__)

ABC_START = 20
ABC_CLEAR_END = 261
CRONO_FIRST_ROW = 10
CRONO_TOTAL_ROW = 172
CRONO_ACUM_ROW = 173
IL_COL = 246  # coluna IL — pesos físicos MÊS1…
MES_WIDTH = 6
MES_FIRST_COL = 5  # E = início do bloco MÊS 1
ORANGE = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
MAX_MESES = 24  # template SEMINF costuma ir além de 6


def _thin_border() -> Border:
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _trunc2(value: float) -> float:
    return math.trunc(float(value) * 100) / 100.0


def _is_admin_stage(name: str) -> bool:
    n = re.sub(r"\s+", " ", (name or "").upper())
    return "ADMINISTRA" in n


def _prazo_meses(premissas: dict[str, Any] | None, default: int = 6) -> int:
    raw = (premissas or {}).get("prazo_meses")
    try:
        n = int(float(raw))
    except (TypeError, ValueError):
        n = default
    return max(1, min(n, MAX_MESES))


def _mes_starts(n_meses: int) -> list[int]:
    return [MES_FIRST_COL + i * MES_WIDTH for i in range(n_meses)]


def _valor_cols(n_meses: int) -> list[str]:
    return [get_column_letter(c) for c in _mes_starts(n_meses)]


def _perc_cols(n_meses: int) -> list[str]:
    return [get_column_letter(c + 4) for c in _mes_starts(n_meses)]


def build_gantt_weights(stage_names: list[str], n_meses: int) -> list[list[int]]:
    """
    Administração → [1]*n_meses.
    Demais etapas → barras contíguas escalonadas ao longo do prazo (Gantt).
    """
    n = len(stage_names)
    weights: list[list[int] | None] = [None] * n
    non_admin: list[int] = []
    for i, name in enumerate(stage_names):
        if _is_admin_stage(name):
            weights[i] = [1] * n_meses
        else:
            non_admin.append(i)

    k = len(non_admin)
    if k == 0:
        return [w or [0] * n_meses for w in weights]

    # Barras contíguas: span = n_meses - k + 1 → cobre o prazo com Gantt (sobreposição
    # quando há menos etapas que meses; 1 mês cada quando etapas ≈ meses).
    span = max(1, min(n_meses, n_meses - k + 1))

    for j, si in enumerate(non_admin):
        w = [0] * n_meses
        if k == 1:
            for m in range(n_meses):
                w[m] = 1
        else:
            start = int(round(j * (n_meses - span) / (k - 1)))
            start = max(0, min(start, n_meses - span))
            for m in range(start, start + span):
                if span >= 3 and m not in (start, start + span - 1):
                    w[m] = 2
                else:
                    w[m] = 1
        weights[si] = w

    return [w or [0] * n_meses for w in weights]


def _collect_servicos(plan: dict[str, Any]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for st in plan.get("stages") or []:
        items = list(st.get("items") or [])
        for sub in st.get("subetapas") or []:
            items.extend(list(sub.get("items") or []))
        for it in items:
            code = str(it.get("code") or "").strip()
            if not code:
                continue
            try:
                qty = float(it.get("qty") if it.get("qty") is not None else 0)
            except (TypeError, ValueError):
                qty = 0.0
            out.append(
                {
                    "code": code,
                    "qty": qty,
                    "description": str(it.get("description") or ""),
                    "unit": str(it.get("unit") or ""),
                    "price_comd": it.get("price_comd"),
                    "price_semd": it.get("price_semd"),
                }
            )
    return out


def fill_curva_abc(
    wb,
    *,
    plan: dict[str, Any],
    base_index: ModelPriceBaseIndex | None,
    obra_type: str | None,
) -> dict[str, Any]:
    """
    Preenche CURVA_ABC com o cenário adotado = menor total entre ComD e SemD
    (paridade macro MCQ!V14/V15: marca \"X\" no menor).
    """
    if "CURVA_ABC" not in wb.sheetnames:
        return {"skipped": True, "reason": "sem aba CURVA_ABC"}

    ws = wb["CURVA_ABC"]
    rates = get_obra_bdi(obra_type)
    bdi_comd = float(rates.rate_com_desoneracao)
    bdi_semd = float(rates.rate_sem_desoneracao)

    for r in range(ABC_START, ABC_CLEAR_END + 1):
        for c in range(1, 14):
            ws.cell(r, c).value = None

    # Monta linhas com ambos os cenários
    rows_raw: list[dict[str, Any]] = []
    total_comd = 0.0
    total_semd = 0.0
    for raw in _collect_servicos(plan):
        code = raw["code"]
        brow = base_index.get_by_code(code) if base_index else None
        desc = raw["description"]
        und = raw["unit"]
        pc = float(raw["price_comd"]) if raw.get("price_comd") is not None else None
        ps = float(raw["price_semd"]) if raw.get("price_semd") is not None else None
        if brow:
            if pc is None:
                pc = float(brow.price_comd)
            if ps is None:
                ps = float(brow.price_semd)
            if not desc:
                desc = brow.description
            if not und:
                und = brow.unit
        if pc is None and ps is None:
            continue
        if pc is None:
            pc = ps
        if ps is None:
            ps = pc
        qty = float(raw["qty"] or 0)
        preco_c = _trunc2(float(pc) * (1.0 + bdi_comd))
        preco_s = _trunc2(float(ps) * (1.0 + bdi_semd))
        tot_c = _trunc2(qty * preco_c)
        tot_s = _trunc2(qty * preco_s)
        total_comd += tot_c
        total_semd += tot_s
        rows_raw.append(
            {
                "cod": code,
                "desc": desc,
                "und": und or "UN",
                "qty": qty,
                "custo_comd": float(pc),
                "custo_semd": float(ps),
                "preco_comd": preco_c,
                "preco_semd": preco_s,
                "total_comd": tot_c,
                "total_semd": tot_s,
            }
        )

    if not rows_raw:
        return {"n_itens": 0, "total": 0.0, "bdi_rate": bdi_comd, "adopted": "comd"}

    # Cenário adotado = menor total (empate → ComD / desonerado)
    if total_semd < total_comd - 1e-9:
        adopted = "semd"
        bdi = bdi_semd
        label = "SEM D"
    else:
        adopted = "comd"
        bdi = bdi_comd
        label = "COM D"

    itens: list[dict[str, Any]] = []
    for it in rows_raw:
        if adopted == "semd":
            itens.append(
                {
                    "cod": it["cod"],
                    "desc": it["desc"],
                    "und": it["und"],
                    "qty": it["qty"],
                    "custo": it["custo_semd"],
                    "preco": it["preco_semd"],
                    "total": it["total_semd"],
                }
            )
        else:
            itens.append(
                {
                    "cod": it["cod"],
                    "desc": it["desc"],
                    "und": it["und"],
                    "qty": it["qty"],
                    "custo": it["custo_comd"],
                    "preco": it["preco_comd"],
                    "total": it["total_comd"],
                }
            )

    itens.sort(key=lambda x: x["total"], reverse=True)

    # Cabeçalho F19 — refletir cenário adotado
    try:
        ws.cell(19, 6).value = f"CUSTO UNIT. (R$) {label}"
    except Exception:
        pass

    border = _thin_border()
    font = Font(name="Arial", size=8)
    start = ABC_START
    last = start + len(itens) - 1

    for i, it in enumerate(itens):
        r = start + i
        ws.cell(r, 1).value = i + 1
        cod = it["cod"]
        try:
            if "SEMINF" not in cod and str(cod).replace(".", "", 1).isdigit():
                f = float(cod)
                ws.cell(r, 2).value = int(f) if f == int(f) else f
            else:
                ws.cell(r, 2).value = cod
        except Exception:
            ws.cell(r, 2).value = cod
        ws.cell(r, 3).value = it["desc"]
        ws.cell(r, 4).value = it["und"]
        ws.cell(r, 5).value = it["qty"]
        ws.cell(r, 6).value = it["custo"]
        ws.cell(r, 7).value = "BDI1"
        ws.cell(r, 8).value = it["preco"]
        ws.cell(r, 9).value = it["total"]
        ws.cell(r, 10).value = f"=IF(SUM($I${start}:$I${last})=0,0,I{r}/SUM($I${start}:$I${last}))"
        ws.cell(r, 11).value = f"=J{r}" if i == 0 else f"=K{r - 1}+J{r}"
        ws.cell(r, 12).value = f'=IF(K{r}<=0.8,"A",IF(K{r}<=0.95,"B","C"))'
        ws.cell(r, 13).value = f'=IF(L{r}="A","Alta",IF(L{r}="B","Média","Baixa"))'
        for c in range(1, 14):
            cell = ws.cell(r, c)
            cell.font = font
            cell.border = border
            if c in (5, 6, 8, 9):
                cell.number_format = "#,##0.00"
            if c in (10, 11):
                cell.number_format = "0.00%"

    try:
        if "TotalCurvaABC" in wb.defined_names:
            del wb.defined_names["TotalCurvaABC"]
    except Exception:
        pass
    ref = f"'CURVA_ABC'!$I${start}:$I${last}"
    wb.defined_names.add(DefinedName(name="TotalCurvaABC", attr_text=ref))

    total = round(sum(i["total"] for i in itens), 2)
    logger.info(
        "CURVA_ABC: %s itens · adotado=%s · total R$ %.2f · BDI %.4f "
        "(ComD=%.2f SemD=%.2f)",
        len(itens),
        adopted,
        total,
        bdi,
        total_comd,
        total_semd,
    )
    return {
        "n_itens": len(itens),
        "total": total,
        "total_comd": round(total_comd, 2),
        "total_semd": round(total_semd, 2),
        "bdi_rate": bdi,
        "adopted": adopted,
        "rows": f"{start}-{last}",
    }


def _safe_set(ws, row: int, col: int, value: Any) -> None:
    """Evita erro em MergedCell (template CRONOGRAMA tem merges)."""
    cell = ws.cell(row, col)
    try:
        cell.value = value
    except AttributeError:
        # célula mesclada secundária — ignora
        return


def _safe_fill(ws, row: int, col: int, fill: PatternFill) -> None:
    cell = ws.cell(row, col)
    try:
        cell.fill = fill
    except AttributeError:
        return


def _paint_gantt(ws, row: int, pesos: list[int]) -> None:
    for m, w in enumerate(pesos):
        c0 = MES_FIRST_COL + m * MES_WIDTH
        for c in range(c0, c0 + MES_WIDTH):
            _safe_fill(ws, row, c, ORANGE if w else PatternFill(fill_type=None))


def _set_month_formulas(ws, row: int, pesos: list[int], n_meses: int) -> None:
    valor_cols = _valor_cols(n_meses)
    perc_cols = _perc_cols(n_meses)
    ativos = [i for i, w in enumerate(pesos) if w]
    n = len(ativos)
    if n == 0:
        for col in valor_cols + perc_cols:
            _safe_set(ws, row, _col_idx(col), 0)
        return

    if n == 1:
        m = ativos[0]
        for i, col in enumerate(valor_cols):
            _safe_set(ws, row, _col_idx(col), f"=$C{row}" if i == m else 0)
        for i, col in enumerate(perc_cols):
            _safe_set(ws, row, _col_idx(col), f"=$D{row}" if i == m else 0)
    else:
        parts_v = "+".join(f"{valor_cols[i]}{row}" for i in ativos[:-1])
        parts_p = "+".join(f"{perc_cols[i]}{row}" for i in ativos[:-1])
        wsum = sum(pesos[a] for a in ativos) or n
        for j, i in enumerate(ativos):
            if j < n - 1:
                wi = pesos[i]
                if wsum and wi != 1:
                    _safe_set(ws, row, _col_idx(valor_cols[i]), f"=ROUND($C{row}*{wi}/{wsum},2)")
                    _safe_set(ws, row, _col_idx(perc_cols[i]), f"=ROUND($D{row}*{wi}/{wsum},2)")
                else:
                    _safe_set(ws, row, _col_idx(valor_cols[i]), f"=ROUND($C{row}/{n},2)")
                    _safe_set(ws, row, _col_idx(perc_cols[i]), f"=ROUND($D{row}/{n},2)")
            else:
                _safe_set(ws, row, _col_idx(valor_cols[i]), f"=$C{row}-({parts_v})")
                _safe_set(ws, row, _col_idx(perc_cols[i]), f"=$D{row}-({parts_p})")
        for i in range(n_meses):
            if i not in ativos:
                _safe_set(ws, row, _col_idx(valor_cols[i]), 0)
                _safe_set(ws, row, _col_idx(perc_cols[i]), 0)

    for col in valor_cols:
        try:
            ws[f"{col}{row}"].number_format = "#,##0.00"
        except Exception:
            pass
    for col in perc_cols:
        try:
            ws[f"{col}{row}"].number_format = "0.00"
        except Exception:
            pass


def _col_idx(letter: str) -> int:
    from openpyxl.utils import column_index_from_string

    return column_index_from_string(letter)


def fill_cronograma(
    wb,
    *,
    etapa_rows: list[dict[str, Any]],
    premissas: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if "CRONOGRAMA" not in wb.sheetnames:
        return {"skipped": True, "reason": "sem aba CRONOGRAMA"}
    if not etapa_rows:
        return {"n_etapas": 0}

    ws = wb["CRONOGRAMA"]
    n_meses = _prazo_meses(premissas)
    prazo_dias = n_meses * 30
    names = [str(e.get("name") or "") for e in etapa_rows]
    weights = build_gantt_weights(names, n_meses)

    ws["A1"] = prazo_dias
    ws["C1"] = "PRAZO:"
    ws["D1"] = f"{prazo_dias} dias / {n_meses} meses"
    ws["A2"] = "PROJETO:"
    ws["B2"] = '=IF(MCQ!K11="","",MCQ!K11)'
    ws["A3"] = "OBJETO:"
    ws["B3"] = '=IF(MCQ!K12="","",MCQ!K12)'
    ws["E2"] = "LOCAL:"
    ws["F2"] = '=IF(MCQ!K13="","",MCQ!K13)'
    ws["E3"] = "ORÇAMENTO:"
    ws["F3"] = '=IF(MCQ!K14="","",MCQ!K14)'
    ws["H2"] = "TOTAL ORÇAMENTO:"
    ws["I2"] = '=IF(MCQ!$V$15="X",PLANILHA!W1217,PLANILHA!S1217)'
    ws["J2"] = '=IF(MCQ!$V$15="X","(NÃO DESONERADO)","(DESONERADO)")'
    ws["I2"].number_format = '"R$" #,##0.00'
    ws["H2"].font = Font(name="Arial", size=9, bold=True)
    ws["I2"].font = Font(name="Arial", size=10, bold=True, color="C00000")
    ws["J2"].font = Font(name="Arial", size=8, italic=True, color="C00000")
    for addr in ("A2", "A3", "E2", "E3", "C1"):
        ws[addr].font = Font(name="Arial", size=9, bold=True)
    for addr in ("B2", "B3", "F2", "F3", "D1"):
        ws[addr].font = Font(name="Arial", size=9)
        ws[addr].alignment = Alignment(wrap_text=True, vertical="center")
    ws["A4"] = (
        "SECRETARIA MUNICIPAL DE INFRAESTRUTURA - SEMINF\n"
        "DEPARTAMENTO DE PROJETOS"
    )
    ws["A6"] = "CRONOGRAMA FÍSICO  / FINANCEIRO"
    ws["A6"].font = Font(name="Arial", size=14, bold=True)

    # Cabeçalhos de mês (pesos IL..) + limpa slots antigos
    for m in range(MAX_MESES):
        _safe_set(ws, 7, IL_COL + m, f"MÊS{m + 1}")
        _safe_set(ws, 8, IL_COL + m, (m + 1) * 30)

    # Limpa linhas de etapa antigas (pares 10..170)
    for r in range(CRONO_FIRST_ROW, CRONO_TOTAL_ROW, 2):
        for c in (1, 2, 3, 4):
            _safe_set(ws, r, c, None)
        for c in range(IL_COL, IL_COL + MAX_MESES):
            _safe_set(ws, r, c, None)
        for m in range(MAX_MESES):
            c0 = MES_FIRST_COL + m * MES_WIDTH
            for c in range(c0, c0 + MES_WIDTH):
                _safe_set(ws, r, c, None)
                _safe_fill(ws, r + 1, c, PatternFill(fill_type=None))

    stage_rows: list[int] = []
    font_b = Font(name="Arial", size=9, bold=True)
    font = Font(name="Arial", size=9)

    for i, meta in enumerate(etapa_rows):
        row = CRONO_FIRST_ROW + i * 2
        if row >= CRONO_TOTAL_ROW - 2:
            logger.warning("CRONOGRAMA: truncando etapas em %s (limite template)", i)
            break
        stage_rows.append(row)
        nome = str(meta.get("name") or f"ETAPA {i + 1}")
        pl_row = int(meta.get("row") or 21)
        pesos = weights[i] if i < len(weights) else [0] * n_meses

        _safe_set(ws, row, 1, f"{i + 1}.")
        _safe_set(ws, row, 2, nome)
        _safe_set(ws, row, 3, f'=IF(MCQ!$V$15="X",PLANILHA!W{pl_row},PLANILHA!S{pl_row})')
        try:
            ws.cell(row, 3).number_format = '"R$" #,##0.00'
            ws.cell(row, 1).font = font_b
            ws.cell(row, 2).font = font_b
            ws.cell(row, 3).font = font
        except Exception:
            pass

        for m, w in enumerate(pesos):
            _safe_set(ws, row, IL_COL + m, int(w) if w else None)

        _safe_set(
            ws,
            row,
            4,
            f'=IF($C${CRONO_TOTAL_ROW}<>0,ROUND(C{row}*100/$C${CRONO_TOTAL_ROW},2),0)',
        )
        try:
            ws.cell(row, 4).number_format = "0.00"
        except Exception:
            pass
        _set_month_formulas(ws, row, pesos, n_meses)
        _paint_gantt(ws, row + 1, pesos)
        ws.row_dimensions[row + 1].height = 9
        ws.row_dimensions[row + 1].hidden = False

    if not stage_rows:
        return {"n_etapas": 0, "n_meses": n_meses}

    # % do último item fecha 100%
    if len(stage_rows) > 1:
        prev = "+".join(f"D{r}" for r in stage_rows[:-1])
        ws.cell(stage_rows[-1], 4).value = f"=ROUND(100-({prev}),2)"
    else:
        ws.cell(stage_rows[0], 4).value = 100

    # Totais
    valor_cols = _valor_cols(n_meses)
    perc_cols = _perc_cols(n_meses)
    tr = CRONO_TOTAL_ROW
    ar = CRONO_ACUM_ROW
    ws[f"C{tr}"] = f"=SUM({','.join(f'C{r}' for r in stage_rows)})"
    ws[f"C{tr}"].number_format = '"R$" #,##0.00'
    ws[f"D{tr}"] = 100
    ws[f"D{tr}"].number_format = "0.00"
    ws[f"B{tr}"] = "VALOR TOTAL"
    ws[f"B{ar}"] = "TOTAL ACUMULADO"
    ws[f"B{tr}"].font = Font(name="Arial", size=9, bold=True)
    ws[f"B{ar}"].font = Font(name="Arial", size=9, bold=True)

    for col in valor_cols + perc_cols:
        refs = ",".join(f"{col}{r}" for r in stage_rows)
        ws[f"{col}{tr}"] = f"=SUM({refs})"
        ws[f"{col}{tr}"].number_format = "#,##0.00" if col in valor_cols else "0.00"

    # Acumulado: meses 1..n-1 progressivos; último = total / 100%
    pairs = list(zip(valor_cols, perc_cols))
    if pairs:
        v0, p0 = pairs[0]
        ws[f"{v0}{ar}"] = f"={v0}{tr}"
        ws[f"{p0}{ar}"] = f"={p0}{tr}"
        for i in range(1, len(pairs) - 1):
            vcol, pcol = pairs[i]
            pv, pp = pairs[i - 1]
            ws[f"{vcol}{ar}"] = f"=ROUND({pv}{ar}+{vcol}{tr},2)"
            ws[f"{pcol}{ar}"] = f"=ROUND({pp}{ar}+{pcol}{tr},2)"
        if len(pairs) >= 2:
            vlast, plast = pairs[-1]
            ws[f"{vlast}{ar}"] = f"=$C{tr}"
            ws[f"{plast}{ar}"] = 100
        for col in valor_cols:
            ws[f"{col}{ar}"].number_format = "#,##0.00"
        for col in perc_cols:
            ws[f"{col}{ar}"].number_format = "0.00"

    # Ocultar linhas vazias entre última etapa+2 e totais
    last_data = stage_rows[-1] + 1
    for r in range(last_data + 1, tr):
        ws.row_dimensions[r].hidden = True
    for r in list(range(1, last_data + 1)) + [tr, ar, ar + 1]:
        ws.row_dimensions[r].hidden = False

    # Impressão A4 paisagem
    last_mes_col = MES_FIRST_COL + n_meses * MES_WIDTH - 1
    last_letter = get_column_letter(last_mes_col)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    try:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.35, bottom=0.35, header=0.1, footer=0.1)
    ws.print_area = f"A1:{last_letter}{last_data},A{tr}:{last_letter}{ar}"
    ws.page_setup.horizontalCentered = True

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 9
    for c0 in _mes_starts(n_meses):
        for c in range(c0, c0 + MES_WIDTH):
            letter = get_column_letter(c)
            ws.column_dimensions[letter].hidden = False
            ws.column_dimensions[letter].width = 7.2
    # Oculta meses além do prazo (até ~col 120 da área financeira)
    for c in range(MES_FIRST_COL + n_meses * MES_WIDTH, 120):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    logger.info(
        "CRONOGRAMA: %s etapas · %s meses · admin_all=%s",
        len(stage_rows),
        n_meses,
        any(_is_admin_stage(n) for n in names),
    )
    return {
        "n_etapas": len(stage_rows),
        "n_meses": n_meses,
        "prazo_dias": prazo_dias,
        "weights": [{"name": names[i], "pesos": weights[i]} for i in range(len(stage_rows))],
    }


def fill_abc_and_cronograma(
    wb,
    *,
    plan: dict[str, Any],
    etapa_rows: list[dict[str, Any]],
    base_index: ModelPriceBaseIndex | None = None,
    project_info: dict[str, Any] | None = None,
    premissas: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Preenche CURVA_ABC + CRONOGRAMA no workbook já aberto (keep_vba).

    MCQ!V14/V15 já têm fórmulas (=IF(R14<R15,\"X\",\"\") / inverso) que
    marcam o menor entre ComD e SemD — não sobrescrever (openpyxl não
    recalcula; Excel/LibreOffice fará ao abrir). CURVA_ABC usa valores
    estáticos: preenche com o mesmo cenário adotado (menor total).
    """
    info = project_info or {}
    prem = premissas or {}
    obra_type = str(info.get("obra_type") or prem.get("obra_type") or "ED")
    abc = fill_curva_abc(wb, plan=plan, base_index=base_index, obra_type=obra_type)
    crono = fill_cronograma(wb, etapa_rows=etapa_rows, premissas=prem)
    return {"curva_abc": abc, "cronograma": crono}
