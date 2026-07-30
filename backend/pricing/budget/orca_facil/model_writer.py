"""Escreve orçamento na CÓPIA da planilha modelo (paridade Cursor / _build_orcamento.py)."""

from __future__ import annotations

import logging
import shutil
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Layout MCQ do modelo Maio/2026 (colunas I..O) — igual ao case ouro CONT_DREN
PRAZO_CELL = "$Z$23"

L_FORMULA = (
    '=IF(I{r}="META", "Digitar a Meta", IF(I{r}="ETAPA", "Digitar a ETAPA", '
    'IF(I{r}="SUB-ETAPA", "Digitar a SUB-ETAPA", IF(K{r}="", "", '
    'IFERROR(VLOOKUP(VALUE(K{r}), {base}, 2, 0), '
    'IFERROR(VLOOKUP(TEXT(K{r}, "@"), {base}, 2, 0), '
    '"Código Não Encontrado na Tabela"))))))'
)
M_FORMULA = (
    '=IF(K{r}="", "", IFERROR(VLOOKUP(VALUE(K{r}), {base}, 3, 0), '
    'IFERROR(VLOOKUP(TEXT(K{r}, "@"), {base}, 3, 0), "")))'
)
O_FORMULA = '=IF(I{r}="S","BDI1","")'
J_FORMULA = "=NUN_ESTRUTURADO(I{r}:I{r}, J{r}:J{r})"


def detect_base_named_range(wb: openpyxl.Workbook) -> str:
    """Detecta named range da base (_BaseMaio2026 etc.)."""
    try:
        names = list(wb.defined_names)
    except Exception:
        names = []
    for n in names:
        low = str(n).lower()
        if "base" in low and not low.startswith("_xlnm"):
            return str(n)
    # fallback comum SEMINF
    return "_BaseMaio2026"


def _qty_formula(qty: Any) -> str:
    try:
        v = float(qty)
        # evitar notação científica
        if abs(v - round(v)) < 1e-9:
            return f"=TRUNC(({int(round(v))}),2)" if abs(v) >= 1 or v == 0 else f"=TRUNC(({v}),2)"
        return f"=TRUNC(({v}),2)"
    except (TypeError, ValueError):
        s = str(qty or "1").strip()
        if s.startswith("="):
            return s
        return "=TRUNC((1),2)"


def _write_etapa(ws, r: int, nome: str, base_name: str) -> int:
    ws.cell(r, 9).value = "ETAPA"
    ws.cell(r, 10).value = J_FORMULA.format(r=r)
    ws.cell(r, 11).value = None
    ws.cell(r, 12).value = nome
    ws.cell(r, 13).value = M_FORMULA.format(r=r, base=base_name)
    ws.cell(r, 14).value = "=TRUNC((0*0),2)"
    ws.cell(r, 15).value = O_FORMULA.format(r=r)
    return r + 1


def _write_subetapa(ws, r: int, nome: str, base_name: str) -> int:
    ws.cell(r, 9).value = "SUB-ETAPA"
    ws.cell(r, 10).value = J_FORMULA.format(r=r)
    ws.cell(r, 11).value = None
    ws.cell(r, 12).value = nome
    ws.cell(r, 13).value = M_FORMULA.format(r=r, base=base_name)
    ws.cell(r, 14).value = "=TRUNC((0*0),2)"
    ws.cell(r, 15).value = O_FORMULA.format(r=r)
    return r + 1


def _write_serv(ws, r: int, codigo: str, qty: Any, memoria: str, base_name: str) -> int:
    ws.cell(r, 9).value = "S"
    ws.cell(r, 10).value = J_FORMULA.format(r=r)
    ws.cell(r, 11).value = str(codigo)
    ws.cell(r, 12).value = L_FORMULA.format(r=r, base=base_name)
    ws.cell(r, 13).value = M_FORMULA.format(r=r, base=base_name)
    ws.cell(r, 14).value = _qty_formula(qty)
    ws.cell(r, 15).value = O_FORMULA.format(r=r)

    m = r + 1
    ws.cell(m, 9).value = None
    ws.cell(m, 10).value = J_FORMULA.format(r=m)
    ws.cell(m, 11).value = None
    ws.cell(m, 12).value = memoria or ""
    ws.cell(m, 13).value = M_FORMULA.format(r=m, base=base_name)
    ws.cell(m, 14).value = None
    ws.cell(m, 15).value = O_FORMULA.format(r=m)
    return r + 2


def _clear_block(ws, start: int = 21, end: int = 250, base_name: str = "_BaseMaio2026") -> None:
    for r in range(start, end + 1):
        ws.cell(r, 9).value = None
        ws.cell(r, 11).value = None
        ws.cell(r, 12).value = None
        ws.cell(r, 14).value = None
        ws.cell(r, 10).value = J_FORMULA.format(r=r)
        ws.cell(r, 13).value = M_FORMULA.format(r=r, base=base_name)
        ws.cell(r, 15).value = O_FORMULA.format(r=r)


def write_plan_to_model_copy(
    *,
    modelo_path: Path,
    dest_path: Path,
    plan: dict[str, Any],
    project_info: dict[str, Any],
    premissas: dict[str, Any] | None = None,
    base_index: Any = None,
) -> dict[str, Any]:
    """
    Copia o .xlsm modelo e preenche MCQ + CURVA_ABC + CRONOGRAMA.
    Retorna estatísticas da escrita.
    """
    from pricing.budget.orca_facil.abc_cronograma import fill_abc_and_cronograma
    from pricing.budget.orca_facil.base_index import build_base_index_from_model

    modelo_path = Path(modelo_path)
    dest_path = Path(dest_path)
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(modelo_path, dest_path)

    wb = openpyxl.load_workbook(dest_path, keep_vba=True)
    if "MCQ" not in wb.sheetnames:
        wb.close()
        raise ValueError("Planilha modelo sem aba MCQ")

    ws = wb["MCQ"]
    base_name = detect_base_named_range(wb)
    info = project_info or {}
    prem = premissas or {}

    # Cabeçalho (K11–K14, Q14) — igual Cursor
    ws["K11"] = str(info.get("projeto") or "CONTENÇÕES")
    ws["K12"] = str(info.get("objeto") or info.get("projeto") or "")
    ws["K13"] = str(info.get("local") or info.get("endereco") or "")
    ws["K14"] = str(info.get("orcamento") or info.get("objeto") or "")[:120]
    obra_type = str(info.get("obra_type") or prem.get("obra_type") or "ED").upper()
    ws["Q14"] = obra_type

    prazo = prem.get("prazo_meses")
    if prazo is not None:
        ws["Z22"] = "Prazo da Obra (meses)"
        try:
            ws["Z23"] = float(prazo)
        except (TypeError, ValueError):
            ws["Z23"] = 6

    _clear_block(ws, 21, 280, base_name=base_name)

    r = 21
    n_etapas = 0
    n_servicos = 0
    n_mem = 0
    missing_code = 0
    etapa_rows: list[dict[str, Any]] = []

    for stage in plan.get("stages") or []:
        name = str(stage.get("name") or "").strip()
        if not name:
            continue
        etapa_rows.append({"name": name, "row": r})
        r = _write_etapa(ws, r, name, base_name)
        n_etapas += 1

        def _add_items(items: list[dict[str, Any]]) -> None:
            nonlocal r, n_servicos, n_mem, missing_code
            for raw in items or []:
                code = str(raw.get("code") or "").strip()
                if not code:
                    missing_code += 1
                    continue
                memory = str(raw.get("memory") or "").strip()
                if not memory:
                    memory = (
                        f"{raw.get('description') or code}\n"
                        f"{raw.get('qty_basis') or 'Quantitativo a revisar'}\n"
                        f"Total = {raw.get('qty') or 1} {raw.get('unit') or ''}"
                    ).strip()
                r = _write_serv(ws, r, code, raw.get("qty"), memory, base_name)
                n_servicos += 1
                n_mem += 1

        _add_items(list(stage.get("items") or []))
        for sub in stage.get("subetapas") or []:
            sub_name = str(sub.get("name") or "").strip()
            if sub_name:
                r = _write_subetapa(ws, r, sub_name, base_name)
            _add_items(list(sub.get("items") or []))
        r += 1  # linha em branco entre etapas

    # Índice da base (para CURVA_ABC) — reutilizar se já veio do pipeline
    idx = base_index
    if idx is None:
        try:
            idx = build_base_index_from_model(modelo_path)
        except Exception as exc:
            logger.warning("Base index indisponível p/ ABC: %s", exc)
            idx = None

    abc_crono: dict[str, Any] = {}
    try:
        abc_crono = fill_abc_and_cronograma(
            wb,
            plan=plan,
            etapa_rows=etapa_rows,
            base_index=idx,
            project_info=info,
            premissas=prem,
        )
    except Exception as exc:
        logger.exception("Falha ao preencher CURVA_ABC/CRONOGRAMA: %s", exc)
        abc_crono = {"error": str(exc)}

    wb.save(dest_path)
    wb.close()
    logger.info(
        "MCQ+ABC+CRONO escritos em %s (etapas=%s serviços=%s base=%s)",
        dest_path.name,
        n_etapas,
        n_servicos,
        base_name,
    )
    return {
        "path": str(dest_path),
        "base_named_range": base_name,
        "n_etapas": n_etapas,
        "n_servicos": n_servicos,
        "n_memorias": n_mem,
        "skipped_without_code": missing_code,
        "last_row": r,
        "etapa_rows": etapa_rows,
        "abc_cronograma": abc_crono,
    }
