"""Extrai árvore etapa→serviços da planilha exemplo (few-shot estruturado)."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Callable

logger = logging.getLogger(__name__)


def extract_example_tree(exemplo_path: Path | None) -> dict[str, Any]:
    """Retorna {stages:[{name, items:[{code, description, unit, qty}]}]} do exemplo."""
    if not exemplo_path or not Path(exemplo_path).is_file():
        return {"stages": [], "source": None}
    try:
        import openpyxl

        wb = openpyxl.load_workbook(exemplo_path, data_only=True, keep_vba=True)
        if "MCQ" not in wb.sheetnames:
            wb.close()
            return {"stages": [], "source": str(exemplo_path)}
        ws = wb["MCQ"]
        stages: list[dict[str, Any]] = []
        current: dict[str, Any] | None = None

        def _start_etapa(name: Any) -> None:
            nonlocal current
            current = {"name": str(name or "ETAPA").strip(), "items": []}
            stages.append(current)

        for r in range(1, min(ws.max_row or 1, 400) + 1):
            t_i = ws.cell(r, 9).value  # I — layout Maio
            t_g = ws.cell(r, 7).value  # G — layout antigo
            if t_i == "ETAPA":
                _start_etapa(ws.cell(r, 12).value)
            elif t_g == "ETAPA":
                _start_etapa(ws.cell(r, 10).value or ws.cell(r, 12).value)
            elif t_i == "S" and current is not None:
                code = ws.cell(r, 11).value
                desc = ws.cell(r, 12).value
                unit = ws.cell(r, 13).value
                qty = ws.cell(r, 14).value
                # se L é fórmula, descrição pode estar vazia no data_only
                if isinstance(desc, str) and desc.startswith("="):
                    desc = ""
                current["items"].append(
                    {
                        "code": str(code).strip() if code is not None else None,
                        "description": str(desc or "").strip()[:200],
                        "unit": str(unit or "").strip() if unit and not str(unit).startswith("=") else "",
                        "qty": float(qty) if isinstance(qty, (int, float)) else qty,
                    }
                )
            elif t_g == "S" and current is not None:
                code = ws.cell(r, 9).value  # I no layout antigo
                desc = ws.cell(r, 10).value
                unit = ws.cell(r, 11).value
                qty = ws.cell(r, 12).value
                if isinstance(desc, str) and desc.startswith("="):
                    desc = ""
                current["items"].append(
                    {
                        "code": str(code).strip() if code is not None else None,
                        "description": str(desc or "").strip()[:200],
                        "unit": str(unit or "").strip() if unit and not str(unit).startswith("=") else "",
                        "qty": float(qty) if isinstance(qty, (int, float)) else qty,
                    }
                )
        wb.close()
        return {
            "stages": stages,
            "source": Path(exemplo_path).name,
            "n_etapas": len(stages),
            "n_servicos": sum(len(s.get("items") or []) for s in stages),
        }
    except Exception as exc:
        logger.warning("Falha ao extrair árvore do exemplo: %s", exc)
        return {"stages": [], "source": str(exemplo_path), "error": str(exc)}


def map_example_codes_to_model_base(
    example_tree: dict[str, Any],
    *,
    get_by_code: Callable[[str], Any],
    search_base: Callable[[str, int], list],
) -> dict[str, Any]:
    """Tenta alinhar códigos do exemplo à base do modelo (mesmo código ou busca por descrição)."""
    mapped_stages = []
    for stage in example_tree.get("stages") or []:
        items = []
        for raw in stage.get("items") or []:
            code = str(raw.get("code") or "").strip()
            desc = str(raw.get("description") or "").strip()
            resolved = None
            if code and get_by_code(code):
                row = get_by_code(code)
                resolved = {
                    "code": row.code,
                    "description": row.description[:160],
                    "unit": row.unit,
                    "qty": raw.get("qty"),
                    "from_example": True,
                    "mapped_how": "exact_code",
                }
            elif desc:
                hits = search_base(desc, 3)
                if hits:
                    row, score = hits[0]
                    if score >= 3.0:
                        resolved = {
                            "code": row.code,
                            "description": row.description[:160],
                            "unit": row.unit,
                            "qty": raw.get("qty"),
                            "from_example": True,
                            "mapped_how": f"search:{score:.1f}",
                            "example_code": code or None,
                        }
            if resolved:
                items.append(resolved)
            else:
                items.append(
                    {
                        "code": code or None,
                        "description": desc,
                        "unit": raw.get("unit") or "",
                        "qty": raw.get("qty"),
                        "from_example": True,
                        "mapped_how": "unresolved",
                        "needs_match": True,
                    }
                )
        mapped_stages.append({"name": stage.get("name"), "items": items})
    return {
        "stages": mapped_stages,
        "source": example_tree.get("source"),
        "n_etapas": len(mapped_stages),
        "n_servicos": sum(len(s.get("items") or []) for s in mapped_stages),
    }
