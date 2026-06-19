"""Indexador regional de modelos de orçamento — SEMINF/Manaus (Amazonas).

Parsing resiliente a cabeçalhos flutuantes, linhas mescladas e abas heterogêneas
típicas de planilhas exportadas da SEINFRA/SEMINF-AM.
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

PUBLISHER = "SEMINF-AM"
REGION = "Manaus/Amazonas"

_ENGINEERING_TAG_RULES: tuple[tuple[tuple[str, ...], tuple[str, ...]], ...] = (
    (("escava", "escavacao", "escavação"), ("terraplanagem", "infraestrutura", "solo")),
    (("aterro", "compacta"), ("terraplanagem", "solo", "compactacao")),
    (("concreto", "concretagem"), ("estrutura", "concreto", "fundacao")),
    (("form", "fôrma", "forma"), ("estrutura", "concreto", "forma")),
    (("armadura", "aco", "aço"), ("estrutura", "aco", "armadura")),
    (("demoli",), ("demolicao", "reforma", "remocao")),
    (("pintura", "tinta"), ("acabamento", "pintura", "protecao")),
    (("impermeabil",), ("impermeabilizacao", "acabamento", "umidade")),
    (("asfalt", "paviment"), ("pavimentacao", "rodovia", "revestimento")),
    (("drenagem", "dreno", "barbac"), ("drenagem", "hidraulica", "infraestrutura")),
    (("sinaliz", "placa"), ("sinalizacao", "transito", "seguranca")),
    (("andaime",), ("seguranca", "acesso", "estrutura-temporaria")),
    (("container", "canteiro", "mobiliz"), ("canteiro", "servicos-preliminares", "locacao")),
    (("energia", "eletric", "entrada provis"), ("instalacoes", "eletrica", "canteiro")),
    (("agua", "água", "esgoto", "sanitar"), ("instalacoes", "hidraulica", "canteiro")),
    (("junta", "dilatacao", "dilatação"), ("pavimento", "juntas", "manutencao")),
    (("metal", "metálic", "solda"), ("estrutura-metalica", "ponte", "reforma")),
    (("limpeza final",), ("servicos-finais", "entrega", "obra")),
)

_SERVICE_CODE_RE = re.compile(r"^\d+(?:\.\d+)*$")
_ITEM_CODE_RE = re.compile(r"^\d+(?:[.-]\d+)*$")


def normalize_service_code(raw: str) -> str:
    """Remove ruído e padroniza códigos SINAPI/SEMINF/WBS."""
    code = unicodedata.normalize("NFKC", (raw or "").strip())
    code = re.sub(r"\s+", "", code)
    code = re.sub(r"\.{2,}", ".", code)
    code = code.upper()
    if code.endswith(".SEMINF"):
        return code
    if _ITEM_CODE_RE.match(code) and not re.search(r"[A-Z]", code.replace(".", "")):
        return code
    return code


def generate_engineering_tags(description: str) -> list[str]:
    text = unicodedata.normalize("NFKD", (description or "").lower())
    text = "".join(c for c in text if not unicodedata.combining(c))
    tags: list[str] = []
    seen: set[str] = set()
    for triggers, bucket in _ENGINEERING_TAG_RULES:
        if any(t in text for t in triggers):
            for tag in bucket:
                if tag not in seen:
                    seen.add(tag)
                    tags.append(tag)
    return tags


def is_amazonas_budget_workbook(path: Path, sheet_names: list[str] | None = None) -> bool:
    name = path.name.lower()
    if any(k in name for k in ("seminf", "seinfra", "ppd", "mc_or", "mc-or", "manaus")):
        return True
    if sheet_names:
        lower = {s.lower() for s in sheet_names}
        if lower & {"planilha", "mcq", "curva_abc", "cronograma"}:
            return True
    return path.suffix.lower() in (".xlsm", ".xlsx", ".xls")


def _cell(row: tuple[Any, ...] | list[Any], idx: int) -> Any:
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _str_val(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _float_val(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if re.search(r",\d{1,2}$", text):
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _norm_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _str_val(value).lower())
    text = "".join(c for c in text if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text).strip()


def _load_workbook_matrix(path: Path) -> tuple[list[str], dict[str, list[tuple[Any, ...]]], dict[str, Any]]:
    """Carrega todas as abas; suporta .xls (xlrd) e .xlsx/.xlsm (openpyxl)."""
    suffix = path.suffix.lower()
    info: dict[str, Any] = {"path": str(path), "sheets": []}

    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd necessário para .xls") from exc
        book = xlrd.open_workbook(str(path))
        info["sheets"] = book.sheet_names()
        sheets = {
            name: [
                tuple(book.sheet_by_name(name).row_values(r))
                for r in range(book.sheet_by_name(name).nrows)
            ]
            for name in book.sheet_names()
        }
        return info["sheets"], sheets, info

    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    info["sheets"] = wb.sheetnames
    sheets: dict[str, list[tuple[Any, ...]]] = {}
    for name in wb.sheetnames:
        sheets[name] = [tuple(r) for r in wb[name].iter_rows(values_only=True)]
    wb.close()
    return info["sheets"], sheets, info


def _detect_ppd_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    """Localiza linha de cabeçalho PPD (col G = tipo, H = item)."""
    for idx, row in enumerate(rows[:80]):
        g = _norm_header(_cell(row, 6))
        h = _norm_header(_cell(row, 7))
        if g in ("tipo", "item", "codigo", "código") or h in ("item", "codigo", "código"):
            return idx
        joined = " ".join(_norm_header(c) for c in row if c)
        if "tipo" in joined and ("item" in joined or "codigo" in joined):
            return idx
    return None


def _detect_tabular_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    """Cabeçalho flutuante genérico (relatórios SEINFRA)."""
    code_keys = ("codigo", "code", "item", "cod")
    desc_keys = ("descricao", "descrição", "description", "servico", "serviço")
    unit_keys = ("unidade", "unit", "und", "un")
    price_keys = ("preco", "preço", "price", "valor unit", "custo unit", "pu", "unitario")

    for idx, row in enumerate(rows[:120]):
        headers = [_norm_header(c) for c in row]
        if not any(headers):
            continue
        cols: dict[str, int] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            if any(k in h for k in code_keys) and "fonte" not in h:
                cols.setdefault("code", i)
            elif any(k in h for k in desc_keys):
                cols.setdefault("description", i)
            elif any(k in h for k in unit_keys):
                cols.setdefault("unit", i)
            elif any(k in h for k in price_keys):
                cols.setdefault("unit_price", i)
        if "description" in cols and ("code" in cols or "unit" in cols):
            return idx, cols
    return None


def _parse_ppd_layout_rows(
    rows: list[tuple[Any, ...]],
    *,
    data_start: int,
    sheet_name: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Parser resiliente do layout PPD (colunas fixas G–N) com skip de linhas vazias."""
    from pricing.budget.ppd_layout import (
        COL_DESCRIPTION,
        COL_ITEM,
        COL_ROW_TYPE,
        COL_SOURCE_CODE,
        COL_UNIT,
        COL_UNIT_PRICE_BDI_COMD,
        ROW_TYPE_ETAPA,
        ROW_TYPE_SERVICO,
        ROW_TYPE_SUB_ETAPA,
    )

    etapas: list[dict[str, Any]] = []
    current_etapa: dict[str, Any] | None = None
    stats = {"rows_scanned": 0, "rows_skipped_blank": 0, "sheet": sheet_name}

    for row_idx, row in enumerate(rows):
        if row_idx < data_start:
            continue
        stats["rows_scanned"] += 1
        if not any(c is not None and str(c).strip() for c in row):
            stats["rows_skipped_blank"] += 1
            continue

        row_type = _str_val(_cell(row, COL_ROW_TYPE)).upper()
        item_code = normalize_service_code(_str_val(_cell(row, COL_ITEM)))
        source_code = normalize_service_code(_str_val(_cell(row, COL_SOURCE_CODE)))
        description = _str_val(_cell(row, COL_DESCRIPTION))
        unit = _str_val(_cell(row, COL_UNIT)).upper()
        unit_price = _float_val(_cell(row, COL_UNIT_PRICE_BDI_COMD))

        if row_type in ("EXEMPLO", "ITEM", "CÓDIGO", "CODIGO", "TIPO"):
            continue

        if row_type in (ROW_TYPE_ETAPA, ROW_TYPE_SUB_ETAPA) and description:
            etapa_code = item_code or str(len(etapas) + 1)
            current_etapa = {"code": etapa_code, "name": description, "services": []}
            etapas.append(current_etapa)
            continue

        is_service = row_type == ROW_TYPE_SERVICO
        if not is_service and description and item_code:
            base = item_code.split("-")[0]
            is_service = bool(_ITEM_CODE_RE.match(base))

        if is_service:
            if not description:
                continue
            if current_etapa is None:
                current_etapa = {"code": "1", "name": "GERAL", "services": []}
                etapas.append(current_etapa)
            svc_code = item_code or f"{current_etapa['code']}.{len(current_etapa['services']) + 1}"
            sinapi = source_code or svc_code
            current_etapa["services"].append(
                {
                    "code": svc_code,
                    "name": description,
                    "sinapi_code": sinapi,
                    "unit": unit,
                    "unit_price": unit_price,
                    "tags": generate_engineering_tags(description),
                }
            )

    return etapas, stats


def _parse_heuristic_sheet(rows: list[tuple[Any, ...]], sheet_name: str) -> list[dict[str, Any]]:
    header = _detect_tabular_header(rows)
    if not header:
        return []
    header_row, cols = header
    etapas: list[dict[str, Any]] = [{"code": "1", "name": sheet_name, "services": []}]
    current = etapas[0]

    for row in rows[header_row + 1 :]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        desc = _str_val(_cell(row, cols.get("description", -1)))
        if not desc or len(desc) < 4:
            continue
        raw_code = _str_val(_cell(row, cols.get("code", -1))) if "code" in cols else ""
        unit = _str_val(_cell(row, cols.get("unit", -1))).upper()
        unit_price = (
            _float_val(_cell(row, cols.get("unit_price", -1))) if "unit_price" in cols else None
        )
        code = (
            normalize_service_code(raw_code)
            if raw_code
            else f"{current['code']}.{len(current['services']) + 1}"
        )

        if not raw_code and len(desc) < 80 and desc.isupper() and not unit:
            current = {"code": str(len(etapas) + 1), "name": desc, "services": []}
            etapas.append(current)
            continue

        sinapi = code if _SERVICE_CODE_RE.match(code.split(".")[0]) or ".SEMINF" in code else ""
        current["services"].append(
            {
                "code": code,
                "name": desc,
                "sinapi_code": sinapi,
                "unit": unit,
                "unit_price": unit_price,
                "tags": generate_engineering_tags(desc),
            }
        )
    return [e for e in etapas if e["services"]]


def _build_summary_text(
    projeto: str,
    obra_type: str,
    objeto: str,
    etapas: list[dict[str, Any]],
) -> str:
    lines = [
        f"MODELO DE ORÇAMENTO REGIONAL ({PUBLISHER}): {projeto}",
        f"Região: {REGION}",
        f"Tipo de obra: {obra_type or 'RF'}",
        f"Objeto: {objeto}",
        "ESTRUTURA WBS:",
    ]
    for etapa in etapas:
        lines.append(f"\nETAPA {etapa['code']} — {etapa['name']}")
        for svc in etapa.get("services") or []:
            line = f"  S {svc['code']} {svc['name']}"
            if svc.get("sinapi_code"):
                line += f" [SINAPI {svc['sinapi_code']}]"
            if svc.get("unit"):
                line += f" ({svc['unit']})"
            price = svc.get("unit_price")
            if price is not None:
                line += f" PU={price:.2f}"
            if svc.get("tags"):
                line += f" #{' #'.join(svc['tags'])}"
            lines.append(line)
    return "\n".join(lines)


def _try_ppd_parser(path: Path) -> dict[str, Any] | None:
    """Reutiliza parse_ppd_workbook quando PLANILHA/MCQ existem."""
    from pricing.budget.ppd_parser import parse_ppd_workbook

    metadata, roots, info = parse_ppd_workbook(path)
    etapas: list[dict[str, Any]] = []
    for root in roots:
        svc_list = []
        for child in root.children:
            raw_code = child.source_code or child.code
            svc_list.append(
                {
                    "code": normalize_service_code(child.code),
                    "name": child.name,
                    "sinapi_code": normalize_service_code(raw_code),
                    "unit": (child.unit or "").upper(),
                    "unit_price": child.unit_price or None,
                    "tags": generate_engineering_tags(child.name),
                }
            )
        etapas.append({"code": root.code, "name": root.name, "services": svc_list})

    service_count = sum(len(e["services"]) for e in etapas)
    if service_count == 0:
        return None

    projeto = metadata.projeto or path.stem
    summary = _build_summary_text(
        projeto,
        metadata.obra_type or "RF",
        metadata.objeto or "",
        etapas,
    )
    return {
        "format": "ppd",
        "publisher": PUBLISHER,
        "region": REGION,
        "obra_type": metadata.obra_type or "RF",
        "projeto": projeto,
        "objeto": metadata.objeto or "",
        "etapas": etapas,
        "service_count": service_count,
        "summary_text": summary,
        "import_info": info,
    }


def extract_regional_budget_model(path: Path) -> dict[str, Any]:
    """
    Ponto de entrada síncrono — chamável via run_sync().
    Retorna sidecar compatível com BudgetModelCatalog e budget_models_index.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix not in (".xlsm", ".xlsx", ".xls"):
        raise ValueError(f"Formato não suportado para indexador regional: {suffix}")

    sheet_names, sheets, info = _load_workbook_matrix(path)
    if not is_amazonas_budget_workbook(path, sheet_names):
        return {
            "format": "regional",
            "error": "workbook_nao_regional",
            "etapas": [],
            "service_count": 0,
        }

    ppd_result = _try_ppd_parser(path)
    if ppd_result and ppd_result.get("service_count", 0) > 0:
        ppd_result.setdefault("publisher", PUBLISHER)
        ppd_result.setdefault("region", REGION)
        return ppd_result

    etapas: list[dict[str, Any]] = []
    parse_stats: list[dict[str, Any]] = []
    preferred = [s for s in sheet_names if s.upper() in ("PLANILHA", "MCQ")] + [
        s for s in sheet_names if not s.lower().startswith("base")
    ]

    for sheet_name in preferred:
        rows = sheets.get(sheet_name) or []
        if not rows:
            continue
        header_row = _detect_ppd_header_row(rows)
        if header_row is not None:
            data_start = max(header_row + 1, 19)
            parsed, stats = _parse_ppd_layout_rows(
                rows, data_start=data_start, sheet_name=sheet_name
            )
            parse_stats.append(stats)
            if parsed:
                etapas.extend(parsed)
                continue
        heuristic = _parse_heuristic_sheet(rows, sheet_name)
        if heuristic:
            etapas.extend(heuristic)

    service_count = sum(len(e.get("services") or []) for e in etapas)
    projeto = path.stem
    obra_type = "RF"
    summary = _build_summary_text(projeto, obra_type, "", etapas) if etapas else ""

    result: dict[str, Any] = {
        "format": "regional",
        "publisher": PUBLISHER,
        "region": REGION,
        "obra_type": obra_type,
        "projeto": projeto,
        "etapas": etapas,
        "service_count": service_count,
        "summary_text": summary or f"Modelo regional {path.name} (0 serviços detectados)",
        "import_info": {**info, "parse_stats": parse_stats},
    }
    if service_count == 0:
        result["error"] = "nenhum_servico_detectado"
    return result


def extract_regional_budget_model_sync(path: str | Path) -> dict[str, Any]:
    """Alias explícito para run_sync() — I/O de disco + parsing."""
    return extract_regional_budget_model(Path(path))
