"""Preservação de assets (logo) ao salvar workbook via openpyxl."""

import sys
import tempfile
import zipfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "planilhas-exemplos"
    / "ppd_seminf_abril_2026.xlsm"
)


@pytest.mark.skipif(not TEMPLATE.exists(), reason="template seminf ausente")
def test_merge_preserves_header_media():
    import openpyxl
    import shutil

    from pricing.budget.ppd_xlsx_assets import merge_workbook_preserving_assets

    with tempfile.TemporaryDirectory() as tmp:
        src = Path(tmp) / "orig.xlsm"
        mod = Path(tmp) / "mod.xlsm"
        shutil.copy(TEMPLATE, src)
        shutil.copy(TEMPLATE, mod)
        wb = openpyxl.load_workbook(mod, keep_vba=True)
        wb["MCQ"].cell(11, 9, "TESTE LOGO")
        wb.save(mod)
        wb.close()

        merged = merge_workbook_preserving_assets(src, mod)
        out = Path(tmp) / "merged.xlsm"
        out.write_bytes(merged)

        with zipfile.ZipFile(out) as z:
            names = z.namelist()
            assert any(n.startswith("xl/media/") for n in names)
            assert any("vmlDrawing" in n for n in names)

        wb2 = openpyxl.load_workbook(out, read_only=True, data_only=True)
        assert wb2["MCQ"]["I11"].value == "TESTE LOGO"
        wb2.close()
