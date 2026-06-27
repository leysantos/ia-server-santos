"""Testes de exportação nativa Excel/PDF."""

from __future__ import annotations

import pytest

from pricing.budget.budget_export_branding import ExportBrandingConfig
from pricing.budget.budget_export_service import (
    export_session_pdf,
    export_session_workbook_xlsx,
    export_session_xlsx,
)
from pricing.budget.budget_session import SESSION_STORE
from pricing.budget.ppd_template import create_empty_ppd_metadata, create_empty_ppd_tree
from pricing.models.budget_item import BudgetItem, BudgetItemType


@pytest.fixture
def export_session():
    meta = create_empty_ppd_metadata(projeto="Obra Teste")
    meta.empresa = "Empresa XYZ"
    etapa = BudgetItem(
        code="1.0",
        name="ETAPA GERAL",
        item_type=BudgetItemType.GROUP,
        row_type="ETAPA",
        level=0,
        quantity=0,
        unit="",
        unit_price=0,
        total_price=1000.0,
        total_price_semd=950.0,
    )
    servico = BudgetItem(
        code="1.1",
        name="Serviço teste",
        item_type=BudgetItemType.COMPOSITION,
        row_type="SERVICO",
        level=1,
        quantity=10.0,
        unit="m²",
        unit_price=100.0,
        source_code="12345",
        total_price=1000.0,
        unit_price_semd=95.0,
        total_price_semd=950.0,
    )
    etapa.children = [servico]
    session = SESSION_STORE.create(
        roots=[etapa],
        title="Obra Teste",
        intent={},
        project=meta,
    )
    branding = ExportBrandingConfig.from_project(meta, session.id)
    session.intent["export_branding_legacy"] = branding.to_dict()
    yield session


def test_export_workbook_xlsx_has_five_sheets(export_session):
    data = export_session_workbook_xlsx(export_session.id)
    assert len(data) > 5000
    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    assert "ORC_SINTETICO" in wb.sheetnames
    assert "ORC_ANALITICO" in wb.sheetnames
    assert "MCQ" in wb.sheetnames
    assert "CRONOGRAMA" in wb.sheetnames
    assert "ESP_TECNICA" in wb.sheetnames
    wb.close()


@pytest.mark.parametrize(
    "doc_type",
    [
        "orc_sintetico",
        "orc_analitico",
        "mcq",
        "cronograma",
        "esp_tecnica",
        "curva_abc",
    ],
)
def test_export_document_xlsx_single_sheet(export_session, doc_type):
    data = export_session_xlsx(export_session.id, doc_type)
    assert len(data) > 2000
    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == doc_type.upper()[:31]
    wb.close()


def test_orc_sintetico_xlsx_service_and_footer_formulas(export_session):
    import io

    import openpyxl

    data = export_session_xlsx(export_session.id, "orc_sintetico")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active

    service_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=2).value == "12345"
    ]
    assert len(service_rows) == 1
    svc_row = service_rows[0]
    service_total = ws.cell(row=svc_row, column=7).value
    assert isinstance(service_total, str)
    assert service_total.startswith("=ROUND(")
    assert f"*F{svc_row}" in service_total

    etapa_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "1.0"
    ]
    assert len(etapa_rows) == 1
    etapa_total = ws.cell(row=etapa_rows[0], column=7).value
    assert isinstance(etapa_total, str)
    assert etapa_total.startswith("=ROUND(SUM(")

    labels = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert "TOTAL SEM BDI" in labels
    assert "TOTAL COM BDI" in labels
    assert any(isinstance(v, str) and v.startswith("BDI (") for v in labels)

    sem_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=3).value == "TOTAL SEM BDI")
    com_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=3).value == "TOTAL COM BDI")
    bdi_row = next(
        r
        for r in range(1, ws.max_row + 1)
        if isinstance(ws.cell(row=r, column=3).value, str)
        and ws.cell(row=r, column=3).value.startswith("BDI (")
    )
    assert sem_row < bdi_row < com_row
    assert str(ws.cell(row=bdi_row, column=7).value).startswith("=G")
    wb.close()


def test_orc_analitico_xlsx_service_cpu_and_footer_formulas(export_session):
    from unittest.mock import patch

    import io

    import openpyxl

    fake_cpu = {
        "items": [
            {
                "item_type": "insumo",
                "code": "88316",
                "description": "Cimento Portland",
                "unit": "kg",
                "coefficient": 5.2,
                "unit_price": 1.5,
                "partial_cost": 7.8,
            }
        ]
    }

    with patch(
        "pricing.tools.budget_pricing_tools.BudgetPricingTools.get_open_composition",
        return_value=fake_cpu,
    ):
        data = export_session_xlsx(export_session.id, "orc_analitico")

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active

    service_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=3).value == "12345"
    ]
    assert len(service_rows) == 1
    svc_row = service_rows[0]
    service_total = ws.cell(row=svc_row, column=8).value
    assert isinstance(service_total, str)
    assert service_total.startswith("=ROUND(")
    assert f"*G{svc_row}" in service_total

    cpu_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=3).value == "88316"
    ]
    assert len(cpu_rows) == 1
    cpu_row = cpu_rows[0]
    cpu_total = ws.cell(row=cpu_row, column=8).value
    assert isinstance(cpu_total, str)
    assert cpu_total.startswith("=ROUND(")
    assert f"F{cpu_row}*G{cpu_row}" in cpu_total.replace(" ", "")

    etapa_rows = [
        r
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "1.0"
    ]
    assert len(etapa_rows) == 1
    etapa_total = ws.cell(row=etapa_rows[0], column=8).value
    assert isinstance(etapa_total, str)
    assert etapa_total.startswith("=ROUND(SUM(")

    labels = [ws.cell(row=r, column=4).value for r in range(1, ws.max_row + 1)]
    assert "TOTAL SEM BDI" in labels
    assert "TOTAL COM BDI" in labels
    wb.close()


def test_mcq_xlsx_no_bdi_column_portrait_layout(export_session):
    import io

    import openpyxl

    data = export_session_xlsx(export_session.id, "mcq")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active

    headers = [ws.cell(row=6, column=c).value for c in range(1, 6)]
    assert headers == ["Item", "Código", "Descrição", "Un", "Qtd"]
    assert ws.cell(row=6, column=6).value is None

    merged = {str(r) for r in ws.merged_cells.ranges}
    assert any("A1:H1" in m for m in merged)

    assert ws.page_setup.orientation == "portrait"
    wb.close()


def test_mcq_pdf_column_widths_and_qty_format():
    from pricing.budget.budget_pdf_portrait_template import (
        MCQ_COL_FRACS,
        TEMPLATE_ID,
        build_mcq_table,
        mcq_col_widths,
    )
    from pricing.budget.budget_pdf_landscape_template import fmt_qty
    from pricing.models.budget_item import BudgetItem, BudgetItemType

    assert TEMPLATE_ID == "portrait_budget_v1"
    assert MCQ_COL_FRACS == (0.07, 0.18, 0.54, 0.07, 0.14)

    widths = mcq_col_widths(100.0)
    assert widths[2] == pytest.approx(54.0)
    assert widths[3] == pytest.approx(7.0)
    assert widths[4] == pytest.approx(14.0)
    assert widths[2] > widths[1]  # descrição > código

    assert fmt_qty(10) == "10,00"
    assert fmt_qty(10.5) == "10,50"

    servico = BudgetItem(
        code="1.1",
        name="Serviço",
        item_type=BudgetItemType.COMPOSITION,
        row_type="SERVICO",
        level=1,
        quantity=12.345,
        unit="m²",
        unit_price=0,
        total_price=0,
        source_code="12345",
    )
    etapa = BudgetItem(
        code="1.0",
        name="ETAPA",
        item_type=BudgetItemType.GROUP,
        row_type="ETAPA",
        level=0,
        quantity=0,
        unit="",
        unit_price=0,
        total_price=0,
        children=[servico],
    )
    table = build_mcq_table([etapa], usable_width=500.0)[0]
    row = table._cellvalues[2]  # serviço (header + etapa + serviço)
    assert row[4] == "12,35"


def test_mcq_pdf_portrait_orientation(export_session):
    from io import BytesIO

    from pypdf import PdfReader

    pdf = export_session_pdf(export_session.id, "mcq")
    page = PdfReader(BytesIO(pdf)).pages[0]
    width = float(page.mediabox.width)
    height = float(page.mediabox.height)
    assert height > width

    text = page.extract_text() or ""
    assert "Obra: Obra Teste" in text
    assert "Empresa: Empresa XYZ" in text
    assert "Base de preços:" in text
    assert "BDI Com D:" in text


def test_orc_analitico_xlsx_merged_header_and_footer(export_session):
    import io

    import openpyxl

    data = export_session_xlsx(export_session.id, "orc_analitico")
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active

    merged = {str(r) for r in ws.merged_cells.ranges}
    assert any("A1:H1" in m for m in merged)
    assert not any(m.startswith("A1:B") for m in merged)
    assert any("A6:D6" in m for m in merged)
    assert any("E6:I6" in m for m in merged)
    assert any("A7:D7" in m for m in merged)
    assert any("E7:I7" in m for m in merged)
    assert ws.cell(row=9, column=1).value == "Item"
    assert ws.cell(row=1, column=1).font.name == "Arial Narrow"
    svc_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=3).value == "12345"
    )
    assert ws.cell(row=svc_row, column=6).number_format == "#,##0.00"
    wb.close()


def test_export_pdf_orc_sintetico(export_session):
    pdf = export_session_pdf(export_session.id, "orc_sintetico")
    assert pdf[:4] == b"%PDF"


def test_orc_analitico_col_widths_favor_code_and_description():
    from pricing.budget.budget_pdf_export import _orc_col_widths

    width = 100.0
    analitico = _orc_col_widths(width, analitico=True)
    sintetico = _orc_col_widths(width, analitico=False)

    assert len(analitico) == 8
    assert analitico[1] >= width * 0.06  # Tipo
    assert analitico[2] > width * 0.10  # Código
    assert analitico[3] > width * 0.35  # Descrição
    assert sum(analitico) == pytest.approx(width)
    assert len(sintetico) == 7


def test_orc_sintetico_col_widths_favor_code_and_description():
    from pricing.budget.budget_pdf_export import _orc_col_widths

    width = 100.0
    sintetico = _orc_col_widths(width, analitico=False)

    assert sintetico[1] > width * 0.08
    assert sintetico[2] > width * 0.38
    assert sintetico[5] < width * 0.18
    assert sintetico[6] < width * 0.18
    assert sum(sintetico) == pytest.approx(width)


def test_orc_sintetico_meta_lines_split_obra_and_pricing():
    from pricing.budget.budget_pdf_landscape_template import (
        format_obra_meta_obra_lines,
        format_obra_meta_pricing_lines,
    )
    from pricing.models.budget_metadata import BudgetProjectMetadata

    meta = BudgetProjectMetadata(
        projeto="Ponte Rio Negro",
        empresa="SEMINF",
        objeto="Novo projeto",
        local="Manaus/AM",
        orcamento="ORC0001-06/2026",
        obra_type="RF",
        price_bases=[
            {
                "enabled": True,
                "label": "SINAPI",
                "uf": "AM",
                "reference": "BR-2026-05",
            },
            {
                "enabled": True,
                "label": "SEMINF",
                "uf": "AM",
                "reference": "BR-DP-SEMINF-2026-05",
            },
        ],
    )
    obra_l1, obra_l2 = format_obra_meta_obra_lines(meta)
    price_l1, price_l2 = format_obra_meta_pricing_lines(meta)

    assert "Obra: Ponte Rio Negro" in obra_l1
    assert "Empresa: SEMINF" in obra_l1
    assert "Local: Manaus/AM" in obra_l1
    assert "Objeto: Novo projeto" in obra_l2
    assert "Orçamento: ORC0001-06/2026" in obra_l2
    assert "Objeto:" not in obra_l1
    assert price_l1.startswith("Base de preços:")
    assert "SINAPI/AM (BR-2026-05)" in price_l1
    assert "SEMINF/AM (BR-DP-SEMINF-2026-05)" in price_l1
    assert "\n" not in price_l1
    assert "Tipo de obra: Rodovias e Ferrovias (RF)" in price_l2
    assert "BDI Com D:" in price_l2
    assert "BDI Sem D:" in price_l2


def test_landscape_budget_template_covers_sintetico_analitico_cronograma():
    from pricing.budget.budget_pdf_landscape_template import LANDSCAPE_BUDGET_DOC_TYPES, TEMPLATE_ID

    assert TEMPLATE_ID == "landscape_budget_v1"
    assert LANDSCAPE_BUDGET_DOC_TYPES == frozenset({"orc_sintetico", "orc_analitico", "cronograma"})


def test_export_branding_persists_header_and_footer(export_session, tmp_path, monkeypatch):
    from core.system.export_branding_store import get_global_export_branding
    from pricing.budget.budget_export_service import get_export_branding, update_export_branding

    monkeypatch.setattr(
        "core.system.export_branding_store.BRANDING_PATH",
        tmp_path / "export_branding.json",
    )
    sid = export_session.id
    updated = update_export_branding(
        sid,
        {
            "header_line1": "Empresa Custom",
            "header_line2": "Obra Custom",
            "header_line3": "Manaus/AM · ORC-TEST",
            "footer_line1": "Eng. Responsável",
            "footer_line2": "Processo 123",
        },
    )
    assert updated.header_line1 == "Empresa Custom"
    loaded = get_export_branding(sid)
    assert loaded.header_line1 == "Empresa Custom"
    assert loaded.footer_line2 == "Processo 123"
    assert get_global_export_branding().header_line2 == "Obra Custom"
    assert not export_session.intent.get("export_branding_customized")


def test_orc_sintetico_skips_memory_rows():
    from pricing.budget.budget_pdf_export import _flatten_budget_rows
    from pricing.models.budget_item import BudgetItem, BudgetItemType

    servico = BudgetItem(
        code="1.1",
        name="Serviço teste",
        item_type=BudgetItemType.COMPOSITION,
        row_type="SERVICO",
        level=1,
        quantity=10.0,
        unit="m²",
        unit_price=100.0,
        source_code="12345",
        total_price=1000.0,
    )
    memoria = BudgetItem(
        code="",
        name="quantidade = 10 m² · área = comprimento × largura",
        item_type=BudgetItemType.COMPOSITION,
        row_type="MEMORIA",
        level=2,
        quantity=0,
        unit="",
        unit_price=0,
        total_price=0,
        metadata={"is_memory_row": True},
        calculation_note="quantidade = 10 m² · área = comprimento × largura",
    )
    servico.children = [memoria]
    etapa = BudgetItem(
        code="1.0",
        name="ETAPA",
        item_type=BudgetItemType.GROUP,
        row_type="ETAPA",
        level=0,
        quantity=0,
        unit="",
        unit_price=0,
        total_price=1000.0,
        children=[servico],
    )

    sintetico_rows = _flatten_budget_rows([etapa], include_memory=False)
    analitico_rows = _flatten_budget_rows([etapa], include_memory=False)

    assert [item.row_type for item, _ in sintetico_rows] == ["ETAPA", "SERVICO"]
    assert [item.row_type for item, _ in analitico_rows] == ["ETAPA", "SERVICO"]


def test_analitico_pdf_expands_open_composition_items():
    from unittest.mock import patch

    from pricing.budget.budget_pdf_export import _build_orc_table
    from pricing.budget.ppd_template import create_empty_ppd_metadata
    from pricing.models.budget_item import BudgetItem, BudgetItemType

    meta = create_empty_ppd_metadata(projeto="Obra")
    meta.price_bases = [
        {"source": "sinapi", "label": "SINAPI", "enabled": True, "uf": "AM", "reference": "BR-2026-05"}
    ]
    servico = BudgetItem(
        code="1.1",
        name="Serviço teste",
        item_type=BudgetItemType.COMPOSITION,
        row_type="SERVICO",
        level=1,
        quantity=10.0,
        unit="m²",
        unit_price=100.0,
        source_code="12345",
        total_price=1000.0,
        unit_price_semd=95.0,
        total_price_semd=950.0,
    )
    etapa = BudgetItem(
        code="1.0",
        name="ETAPA",
        item_type=BudgetItemType.GROUP,
        row_type="ETAPA",
        level=0,
        quantity=0,
        unit="",
        unit_price=0,
        total_price=1000.0,
        children=[servico],
    )
    fake_cpu = {
        "items": [
            {
                "item_type": "insumo",
                "code": "88316",
                "description": "Cimento Portland",
                "unit": "kg",
                "coefficient": 5.2,
                "unit_price": 1.5,
                "partial_cost": 7.8,
                "unit_price_sem": 1.4,
                "partial_cost_sem": 7.28,
            }
        ]
    }

    with patch(
        "pricing.tools.budget_pricing_tools.BudgetPricingTools.get_open_composition",
        return_value=fake_cpu,
    ):
        story = _build_orc_table([etapa], meta=meta, analitico=True, usable_width=500.0)
    table = story[0]
    rows = table._cellvalues
    header = [str(c) for c in rows[0]]
    assert any("Tipo" in h for h in header)
    assert not any("Unit. Sem D" in h for h in header)  # ComD vence (1000 vs 950 na etapa)
    assert any("Unit. Com D" in h for h in header)
    assert len(rows) >= 4
    cpu_row = rows[3]
    assert "Material" in str(cpu_row[1])
    assert "88316" in str(cpu_row[2])
    assert "Cimento" in str(cpu_row[3])


def test_analitico_pdf_adopted_semd_mode_single_price_columns():
    from pricing.budget.budget_pdf_export import _build_orc_table
    from pricing.budget.ppd_template import create_empty_ppd_metadata
    from pricing.models.budget_item import BudgetItem, BudgetItemType

    meta = create_empty_ppd_metadata(projeto="Obra")
    servico = BudgetItem(
        code="1.1",
        name="Serviço",
        item_type=BudgetItemType.COMPOSITION,
        row_type="SERVICO",
        level=1,
        quantity=1.0,
        unit="m²",
        unit_price=120.0,
        source_code="99999",
        total_price=120.0,
        unit_price_semd=100.0,
        total_price_semd=100.0,
    )
    etapa = BudgetItem(
        code="1.0",
        name="ETAPA",
        item_type=BudgetItemType.GROUP,
        row_type="ETAPA",
        level=0,
        quantity=0,
        unit="",
        unit_price=0,
        total_price=120.0,
        total_price_semd=100.0,
        children=[servico],
    )
    story = _build_orc_table([etapa], meta=meta, analitico=True, usable_width=500.0)
    header = [str(c) for c in story[0]._cellvalues[0]]
    assert any("Sem D" in h for h in header)
    assert not any("Unit. Com D" in h for h in header)
    service_row = story[0]._cellvalues[2]
    assert "100,00" in str(service_row[6]) or "100.00" in str(service_row[6])


def test_folha_page_label_format():
    from pricing.budget.budget_pdf_landscape_template import folha_page_label

    assert folha_page_label(1, 3) == "Pagina 01/03"
    assert folha_page_label(12, 12) == "Pagina 12/12"


def test_landscape_pdf_no_trailing_blank_page():
    from io import BytesIO

    from pypdf import PdfReader
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    from pricing.budget.budget_pdf_landscape_template import build_landscape_context, render_landscape_pdf
    from pricing.budget.budget_export_branding import ExportBrandingConfig
    from core.system.company_profile import get_company_profile

    profile = get_company_profile()
    brand = ExportBrandingConfig()
    ctx = build_landscape_context(
        title="Teste",
        brand=brand,
        profile=profile,
        logo_path=None,
        brasao_path=None,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(f"Linha {i}", styles["Normal"]) for i in range(120)]
    pdf = render_landscape_pdf(story, ctx=ctx)
    pages = PdfReader(BytesIO(pdf)).pages
    assert len(pages) >= 1
    last_text = (pages[-1].extract_text() or "").strip()
    assert "Pagina" in last_text
    assert "Linha" in last_text or "Orçamento" in last_text or "Teste" in last_text


@pytest.mark.parametrize("doc_type", ["mcq", "orc_analitico", "cronograma", "esp_tecnica", "curva_abc"])
def test_export_pdf_all_docs(export_session, doc_type):
    pdf = export_session_pdf(export_session.id, doc_type)
    assert pdf[:4] == b"%PDF"


@pytest.fixture
def export_session_with_schedule(export_session):
    from pricing.schedule.schedule_models import ProjectSchedule, ScheduleTask

    servico = export_session.roots[0].children[0]
    export_session.schedule = ProjectSchedule(
        project_start="2026-06-01",
        project_end="2026-12-31",
        tasks=[
            ScheduleTask(
                task_id="t1",
                budget_row_id=servico.row_id,
                budget_code="1.1",
                name=servico.name,
                row_type="S",
                duration_days=30,
                early_start="2026-06-01",
                early_finish="2026-06-30",
            )
        ],
    )
    return export_session


def test_export_curva_s_pdf(export_session_with_schedule):
    pdf = export_session_pdf(export_session_with_schedule.id, "curva_s")
    assert pdf[:4] == b"%PDF"


def test_export_histograma_pdf(export_session_with_schedule):
    pdf = export_session_pdf(export_session_with_schedule.id, "histograma")
    assert pdf[:4] == b"%PDF"


def test_export_curva_s_xlsx(export_session_with_schedule):
    data = export_session_xlsx(export_session_with_schedule.id, "curva_s")
    assert len(data) > 2000


def test_export_curva_s_without_schedule_raises(export_session):
    with pytest.raises(ValueError, match="Cronograma"):
        export_session_pdf(export_session.id, "curva_s")
