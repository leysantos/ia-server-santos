"""Testes preparação PDF PPD — aba isolada + área de impressão."""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "planilhas-exemplos"
    / "19_PPD_MC_OR_R01-Nivel-1-2-Marco2026-14-05-2026.xlsm"
)


@pytest.mark.skipif(not TEMPLATE.exists(), reason="template exemplo ausente")
def test_prepare_single_sheet_workbook_exists():
    from pricing.budget.ppd_pdf_prepare import cleanup_prepared_workbook, prepare_single_sheet_pdf_workbook

    prepared = prepare_single_sheet_pdf_workbook(TEMPLATE, "PLANILHA")
    try:
        assert prepared.exists()
        import openpyxl

        wb = openpyxl.load_workbook(prepared, read_only=True)
        names_upper = {n.upper() for n in wb.sheetnames}
        assert "PLANILHA" in names_upper
        assert "MCQ" in names_upper
        assert any(n.startswith("Base_") for n in wb.sheetnames)
        assert wb.active.title.upper() == "PLANILHA"
        wb.close()
    finally:
        cleanup_prepared_workbook(prepared)


@pytest.mark.skipif(not TEMPLATE.exists(), reason="template exemplo ausente")
def test_planilha_pdf_page_count_when_lo_available():
    from pricing.budget.ppd_lo_export import export_sheet_pdf, libreoffice_available

    if not libreoffice_available():
        pytest.skip("LibreOffice indisponível")

    from pypdf import PdfReader
    import io

    pdf = export_sheet_pdf(TEMPLATE, sheet="PLANILHA")
    count = len(PdfReader(io.BytesIO(pdf)).pages)
    assert count < 20, f"PLANILHA PDF com páginas demais: {count}"
