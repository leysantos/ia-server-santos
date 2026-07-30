"""OF8 — testes CURVA_ABC / CRONOGRAMA (pesos Gantt + cenário adotado)."""

from __future__ import annotations

from openpyxl import Workbook

from pricing.budget.orca_facil.abc_cronograma import (
    build_gantt_weights,
    fill_curva_abc,
    _is_admin_stage,
)
from pricing.budget.orca_facil.base_index import BaseRow, ModelPriceBaseIndex


def test_admin_detection():
    assert _is_admin_stage("ADMINISTRAÇÃO DA OBRA")
    assert _is_admin_stage("Administracao local")
    assert not _is_admin_stage("DRENAGEM")


def test_gantt_admin_all_months():
    names = ["ADMINISTRAÇÃO DA OBRA", "A", "B", "C"]
    w = build_gantt_weights(names, 6)
    assert w[0] == [1, 1, 1, 1, 1, 1]
    # 3 não-admin em 6 meses → span = 6-3+1 = 4
    assert sum(1 for x in w[1] if x) == 4
    assert any(w[-1])


def test_gantt_covers_timeline_when_stages_eq_months():
    names = ["ADMIN"] + [f"E{i}" for i in range(6)]
    # ADMIN matches ADMINISTRA? "ADMIN" alone — check
    names[0] = "ADMINISTRAÇÃO"
    w = build_gantt_weights(names, 6)
    assert w[0] == [1] * 6
    # cada etapa não-admin 1 mês
    for row in w[1:]:
        assert sum(1 for x in row if x) == 1
    # juntos cobrem todos os meses
    covered = [0] * 6
    for row in w[1:]:
        for m, v in enumerate(row):
            if v:
                covered[m] += 1
    assert all(covered)


def _index_with(*rows: BaseRow) -> ModelPriceBaseIndex:
    return ModelPriceBaseIndex(rows=list(rows))


def test_curva_abc_adopts_menor_semd():
    """Paridade macro MCQ: se SemD+BDI < ComD+BDI, ABC usa SemD."""
    wb = Workbook()
    wb.create_sheet("CURVA_ABC")
    idx = _index_with(
        BaseRow(
            code="1",
            description="A",
            unit="UN",
            price_comd=100.0,
            price_semd=80.0,
        )
    )
    plan = {
        "stages": [
            {
                "name": "X",
                "items": [{"code": "1", "qty": 10, "description": "A", "unit": "UN"}],
            }
        ]
    }
    out = fill_curva_abc(wb, plan=plan, base_index=idx, obra_type="ED")
    assert out["adopted"] == "semd"
    assert out["total_semd"] < out["total_comd"]
    assert out["total"] == out["total_semd"]
    assert "SEM D" in str(wb["CURVA_ABC"].cell(19, 6).value)


def test_curva_abc_adopts_comd_when_cheaper_or_tie():
    wb = Workbook()
    wb.create_sheet("CURVA_ABC")
    idx = _index_with(
        BaseRow(
            code="2",
            description="B",
            unit="UN",
            price_comd=50.0,
            price_semd=90.0,
        )
    )
    plan = {
        "stages": [
            {
                "name": "Y",
                "items": [{"code": "2", "qty": 1, "description": "B", "unit": "UN"}],
            }
        ]
    }
    out = fill_curva_abc(wb, plan=plan, base_index=idx, obra_type="ED")
    assert out["adopted"] == "comd"
    assert out["total"] == out["total_comd"]
    assert "COM D" in str(wb["CURVA_ABC"].cell(19, 6).value)
